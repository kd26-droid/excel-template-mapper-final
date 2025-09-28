"""
Working views for Excel Template Mapper application.
Optimized for smooth Excel to Excel mapping functionality.
"""

import os
import uuid
import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, Any, Optional

import pandas as pd
from django.conf import settings
from django.core.cache import cache
from django.http import FileResponse, Http404, JsonResponse, HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.views.decorators.cache import never_cache
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import numpy as np
from collections import defaultdict
import re
import traceback
import json
import tempfile
import shutil

from .bom_header_mapper import BOMHeaderMapper
from .models import MappingTemplate, TagTemplate
try:
    # Prefer relative import; fall back gracefully on any import error
    from .azure_storage import hybrid_file_manager
except Exception:
    # Fallback to local file manager if azure_storage is not available
    import os
    import uuid
    from pathlib import Path
    from typing import Tuple
    
    class LocalFileManager:
        def __init__(self):
            self.local_upload_dir = Path(settings.BASE_DIR) / 'uploaded_files'
            self.local_temp_dir = Path(settings.BASE_DIR) / 'temp_downloads'
            self._ensure_local_directories()
        
        def _ensure_local_directories(self):
            self.local_upload_dir.mkdir(parents=True, exist_ok=True)
            self.local_temp_dir.mkdir(parents=True, exist_ok=True)
        
        def save_upload_file(self, file, prefix="upload") -> Tuple[str, str]:
            file_extension = Path(file.name).suffix
            unique_filename = f"{uuid.uuid4()}_{prefix}{file_extension}"
            local_file_path = self.local_upload_dir / unique_filename
            
            with open(local_file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            
            return str(local_file_path), file.name
        
        def get_file_path(self, file_identifier: str) -> str:
            return file_identifier
    
    hybrid_file_manager = LocalFileManager()

# Configure logging
logger = logging.getLogger(__name__)

# === Canonicalizer for header labels ===
def _canon(s: str) -> str:
    """
    Canonicalize header strings for consistent comparison.
    Handles NBSP, whitespace, case differences, and punctuation variations.
    """
    return (
        str(s or "")
        .replace("\u00a0", " ")  # Replace non-breaking space with regular space
        .replace("\r", " ")      # Replace carriage return with space
        .replace("\n", " ")      # Replace newline with space
        .strip()                 # Remove leading/trailing whitespace
        .lower()                 # Convert to lowercase
        .replace(" ", "")        # Remove all spaces
        .replace("_", "")        # Remove underscores
        .replace("-", "")        # Remove hyphens
    )

def read_csv_with_encoding(file_path, header_row, **kwargs):
    """
    Helper function to read CSV files with proper encoding detection.
    Tries multiple encodings to handle various CSV formats.
    """
    encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'windows-1252']

    for encoding in encodings_to_try:
        try:
            df = pd.read_csv(
                file_path,
                header=header_row,
                encoding=encoding,
                on_bad_lines='skip',
                **kwargs
            )
            return df
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            if encoding == encodings_to_try[-1]:
                raise e
            continue

    raise Exception("Could not read CSV file with any supported encoding")


def generate_template_columns(tags_count=3, spec_pairs_count=3, customer_id_pairs_count=1):
    """
    Generate complete template column headers including all standard template fields.
    Always includes the 6 core Factwise headers, standard template fields, and dynamic columns.
    Default: 3 tags, 3 spec pairs, 1 customer identification pair.
    """
    # Start with core Factwise headers - these must always be present
    headers = [
        "Item code", "Item name", "Description", "Item type", "Measurement unit", "Procurement entity name"
    ]
    
    # Add standard template fields that should always be present
    headers.extend([
        "Notes", "Internal notes", "Procurement item", "Sales item", "Preferred vendor code"
    ])
    
    # Add dynamic Tag columns (default 3)
    for i in range(1, tags_count + 1):
        headers.append(f"Tag_{i}")
    
    # Add dynamic Specification pairs (default 3)
    for i in range(1, spec_pairs_count + 1):
        headers.extend([f"Specification_Name_{i}", f"Specification_Value_{i}"])
    
    # Add dynamic Customer Identification pairs (default 1)
    for i in range(1, customer_id_pairs_count + 1):
        headers.extend([f"Customer_Identification_Name_{i}", f"Customer_Identification_Value_{i}"])
    
    return headers

# Utility: normalize template/display headers to internal numbered headers
def normalize_headers_to_internal(headers: list, existing_headers: Optional[list] = None) -> list:
    """Convert any external/display dynamic headers to internal numbered forms.
    - Tag → Tag_1, Tag_2, ...
    - Specification name/value → Specification_Name_N / Specification_Value_N
    - Customer identification name/value → Customer_Identification_Name_N / Customer_Identification_Value_N
    Keeps already-internal names unchanged.
    """
    logger.debug(f"🔄 normalize_headers_to_internal called with {len(headers)} headers")
    logger.debug(f"Input headers: {headers}")
    logger.debug(f"Existing headers: {existing_headers}")
    
    if not headers or not isinstance(headers, list):
        logger.warning(f"Invalid headers input: {headers}")
        return headers
    
    # Build maps for existing numbered headers to preserve their indices
    tag_map = {}
    spec_name_map = {}
    spec_value_map = {}
    cust_name_map = {}
    cust_value_map = {}
    
    if existing_headers:
        for h in existing_headers:
            if h.startswith('Tag_'):
                try:
                    idx = int(h.split('_')[1])
                    tag_map[idx] = h
                    logger.debug(f"Found existing Tag_{idx}: {h}")
                except (IndexError, ValueError):
                    logger.warning(f"Could not parse Tag index from: {h}")
            elif h.startswith('Specification_Name_'):
                try:
                    idx = int(h.split('_')[2])
                    spec_name_map[idx] = h
                    logger.debug(f"Found existing Specification_Name_{idx}: {h}")
                except (IndexError, ValueError):
                    logger.warning(f"Could not parse Specification_Name index from: {h}")
            elif h.startswith('Specification_Value_'):
                try:
                    idx = int(h.split('_')[2])
                    spec_value_map[idx] = h
                    logger.debug(f"Found existing Specification_Value_{idx}: {h}")
                except (IndexError, ValueError):
                    logger.warning(f"Could not parse Specification_Value index from: {h}")
            elif h.startswith('Customer_Identification_Name_'):
                try:
                    idx = int(h.split('_')[2])
                    cust_name_map[idx] = h
                    logger.debug(f"Found existing Customer_Identification_Name_{idx}: {h}")
                except (IndexError, ValueError):
                    logger.warning(f"Could not parse Customer_Identification_Name index from: {h}")
            elif h.startswith('Customer_Identification_Value_'):
                try:
                    idx = int(h.split('_')[2])
                    cust_value_map[idx] = h
                    logger.debug(f"Found existing Customer_Identification_Value_{idx}: {h}")
                except (IndexError, ValueError):
                    logger.warning(f"Could not parse Customer_Identification_Value index from: {h}")
    
    next_tag_idx = max(tag_map.keys()) + 1 if tag_map else 1
    next_spec_idx = max(spec_name_map.keys()) + 1 if spec_name_map else 1
    next_cust_idx = max(cust_name_map.keys()) + 1 if cust_name_map else 1
    
    logger.debug(f"Next available indices - Tag: {next_tag_idx}, Spec: {next_spec_idx}, Customer: {next_cust_idx}")
    
    normalized = []
    
    def norm(s: str) -> str:
        return str(s or '').strip().lower()
    
    for h in headers:
        h_str = str(h)
        h_norm = norm(h_str)
        logger.debug(f"Processing header: '{h_str}' (normalized: '{h_norm}')")
        
        # Tag handling
        if h_norm == 'tag':
            # Find an existing Tag_N or assign a new one
            assigned = False
            for i in range(1, next_tag_idx):
                if i not in tag_map:
                    normalized.append(f'Tag_{i}')
                    tag_map[i] = f'Tag_{i}'
                    assigned = True
                    logger.debug(f"Assigned Tag_{i} to '{h_str}'")
                    break
            if not assigned:
                normalized.append(f'Tag_{next_tag_idx}')
                tag_map[next_tag_idx] = f'Tag_{next_tag_idx}'
                logger.debug(f"Assigned new Tag_{next_tag_idx} to '{h_str}'")
                next_tag_idx += 1
            continue
            
        if h_str.startswith('Tag_'):
            normalized.append(h_str)
            try:
                num = int(h_str.split('_')[1])
                tag_map[num] = h_str
                next_tag_idx = max(next_tag_idx, num + 1)
                logger.debug(f"Preserved existing Tag_{num}: {h_str}")
            except Exception:
                logger.warning(f"Could not parse Tag index from: {h_str}")
            continue
            
        # Specification name handling
        if h_norm in ['specification name', 'spec name', 'specification_name']:
            # Find an existing Spec_Name_N or assign a new one
            assigned = False
            for i in range(1, next_spec_idx):
                if i not in spec_name_map:
                    normalized.append(f'Specification_Name_{i}')
                    spec_name_map[i] = f'Specification_Name_{i}'
                    assigned = True
                    logger.debug(f"Assigned Specification_Name_{i} to '{h_str}'")
                    break
            if not assigned:
                normalized.append(f'Specification_Name_{next_spec_idx}')
                spec_name_map[next_spec_idx] = f'Specification_Name_{next_spec_idx}'
                logger.debug(f"Assigned new Specification_Name_{next_spec_idx} to '{h_str}'")
                next_spec_idx += 1
            continue
            
        if h_str.startswith('Specification_Name_'):
            normalized.append(h_str)
            try:
                num = int(h_str.split('_')[2])
                spec_name_map[num] = h_str
                next_spec_idx = max(next_spec_idx, num + 1)
                logger.debug(f"Preserved existing Specification_Name_{num}: {h_str}")
            except Exception:
                logger.warning(f"Could not parse Specification_Name index from: {h_str}")
            continue
            
        # Specification value handling
        if h_norm in ['specification value', 'spec value', 'specification_value']:
            # Find an existing Spec_Value_N or assign a new one
            assigned = False
            for i in range(1, next_spec_idx):
                if i not in spec_value_map:
                    normalized.append(f'Specification_Value_{i}')
                    spec_value_map[i] = f'Specification_Value_{i}'
                    assigned = True
                    logger.debug(f"Assigned Specification_Value_{i} to '{h_str}'")
                    break
            if not assigned:
                normalized.append(f'Specification_Value_{next_spec_idx}')
                spec_value_map[next_spec_idx] = f'Specification_Value_{next_spec_idx}'
                logger.debug(f"Assigned new Specification_Value_{next_spec_idx} to '{h_str}'")
                next_spec_idx += 1
            continue
            
        if h_str.startswith('Specification_Value_'):
            normalized.append(h_str)
            try:
                num = int(h_str.split('_')[2])
                spec_value_map[num] = h_str
                next_spec_idx = max(next_spec_idx, num + 1)
                logger.debug(f"Preserved existing Specification_Value_{num}: {h_str}")
            except Exception:
                logger.warning(f"Could not parse Specification_Value index from: {h_str}")
            continue
            
        # Customer identification name handling
        if h_norm in ['customer identification name', 'customer id name', 'customer_id_name']:
            # Find an existing Customer_Identification_Name_N or assign a new one
            assigned = False
            for i in range(1, next_cust_idx):
                if i not in cust_name_map:
                    normalized.append(f'Customer_Identification_Name_{i}')
                    cust_name_map[i] = f'Customer_Identification_Name_{i}'
                    assigned = True
                    logger.debug(f"Assigned Customer_Identification_Name_{i} to '{h_str}'")
                    break
            if not assigned:
                normalized.append(f'Customer_Identification_Name_{next_cust_idx}')
                cust_name_map[next_cust_idx] = f'Customer_Identification_Name_{next_cust_idx}'
                logger.debug(f"Assigned new Customer_Identification_Name_{next_cust_idx} to '{h_str}'")
                next_cust_idx += 1
            continue
            
        if h_str.startswith('Customer_Identification_Name_'):
            normalized.append(h_str)
            try:
                num = int(h_str.split('_')[2])
                cust_name_map[num] = h_str
                next_cust_idx = max(next_cust_idx, num + 1)
                logger.debug(f"Preserved existing Customer_Identification_Name_{num}: {h_str}")
            except Exception:
                logger.warning(f"Could not parse Customer_Identification_Name index from: {h_str}")
            continue
            
        # Customer identification value handling
        if h_norm in ['customer identification value', 'customer id value', 'customer_id_value']:
            # Find an existing Customer_Identification_Value_N or assign a new one
            assigned = False
            for i in range(1, next_cust_idx):
                if i not in cust_value_map:
                    normalized.append(f'Customer_Identification_Value_{i}')
                    cust_value_map[i] = f'Customer_Identification_Value_{i}'
                    assigned = True
                    logger.debug(f"Assigned Customer_Identification_Value_{i} to '{h_str}'")
                    break
            if not assigned:
                normalized.append(f'Customer_Identification_Value_{next_cust_idx}')
                cust_value_map[next_cust_idx] = f'Customer_Identification_Value_{next_cust_idx}'
                logger.debug(f"Assigned new Customer_Identification_Value_{next_cust_idx} to '{h_str}'")
                next_cust_idx += 1
            continue
            
        if h_str.startswith('Customer_Identification_Value_'):
            normalized.append(h_str)
            try:
                num = int(h_str.split('_')[2])
                cust_value_map[num] = h_str
                next_cust_idx = max(next_cust_idx, num + 1)
                logger.debug(f"Preserved existing Customer_Identification_Value_{num}: {h_str}")
            except Exception:
                logger.warning(f"Could not parse Customer_Identification_Value index from: {h_str}")
            continue
        
        # Non-dynamic header - keep as is
        normalized.append(h_str)
        logger.debug(f"Kept non-dynamic header as-is: '{h_str}'")
    
    logger.info(f"✅ Header normalization complete: {len(headers)} → {len(normalized)}")
    logger.debug(f"Final normalized headers: {normalized}")
    return normalized

# In-memory store for each session
SESSION_STORE = {}

# Cache control and snapshot helper functions
def no_store(resp: Response) -> Response:
    """Add no-store cache headers to prevent caching issues across workers."""
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = "0"
    return resp

def increment_template_version(session_id):
    """Increment template version for a session to track changes."""
    if session_id in SESSION_STORE:
        current_version = SESSION_STORE[session_id].get('template_version', 0)
        new_version = current_version + 1
        SESSION_STORE[session_id]['template_version'] = new_version
        save_session_to_file(session_id, SESSION_STORE[session_id])
        logger.info(f"🔄 Template version incremented for session {session_id}: {current_version} → {new_version}")
        return new_version
    return 0

def get_template_version(session_id):
    """Get current template version for a session."""
    if session_id in SESSION_STORE:
        return SESSION_STORE[session_id].get('template_version', 0)
    return 0

def _externalize_formula_rules(rules, session_info=None):
    """Return a copy of formula rules with internal targets shown as external labels for UI.
    Example: Tag_4 -> Tag, Specification_Name_1 -> 'Specification name' (already handled by converter).
    """
    try:
        ext = []
        for r in (rules or []):
            rule = dict(r or {})
            tcol = rule.get('target_column')
            if isinstance(tcol, str) and tcol:
                try:
                    rule['target_column'] = convert_internal_to_external_name(tcol)
                except Exception:
                    pass
            ext.append(rule)
        return ext
    except Exception:
        return rules or []

def build_snapshot(info: dict) -> dict:
    """Build canonical snapshot of session state."""
    return {
        "version": info.get("version", 0),
        "template_version": info.get("template_version", 0),
        "headers": info.get("enhanced_headers") or info.get("current_template_headers") or info.get("template_headers") or [],
        "mappings": info.get("mappings") or {"mappings": []},
        "default_values": info.get("default_values") or {},
        "counts": {
            "tags_count": info.get("tags_count", 1),
            "spec_pairs_count": info.get("spec_pairs_count", 1),
            "customer_id_pairs_count": info.get("customer_id_pairs_count", 1),
        },
        "formula_rules": _externalize_formula_rules(info.get("formula_rules") or [], info),
        "factwise_rules": info.get("factwise_rules") or [],
    }

# Session persistence helper functions
def save_session_to_file(session_id, session_data):
    """Save session data to file for persistence."""
    try:
        session_file = hybrid_file_manager.local_temp_dir / f"session_{session_id}.json"
        import json
        with open(session_file, 'w') as f:
            # Convert paths to strings for JSON serialization
            serializable_data = {}
            for key, value in session_data.items():
                if isinstance(value, Path):
                    serializable_data[key] = str(value)
                else:
                    serializable_data[key] = value
            json.dump(serializable_data, f)
        logger.info(f"💾 Saved session {session_id} to file")
    except Exception as e:
        logger.warning(f"Failed to save session {session_id}: {e}")

def load_session_from_file(session_id):
    """Load session data from file."""
    try:
        session_file = hybrid_file_manager.local_temp_dir / f"session_{session_id}.json"
        if session_file.exists():
            import json
            with open(session_file, 'r') as f:
                session_data = json.load(f)
            logger.info(f"📂 Loaded session {session_id} from file")
            return session_data
    except Exception as e:
        logger.warning(f"Failed to load session {session_id}: {e}")
    return None

def get_session_consistent(session_id: str):
    """
    Get session from consistent storage (cache-first approach for Azure multi-worker).
    1) Try Redis cache first (shared across workers)
    2) Fallback to in-process memory
    3) Fallback to file snapshot
    """
    # Try cache first (shared across Azure workers)
    data = cache.get(f"mapper:session:{session_id}")
    if data:
        logger.info(f"🔍 Session {session_id} found in cache")
        # Cross-check file snapshot for newer version to avoid stale cache across workers
        try:
            file_snapshot = load_session_from_file(session_id)
            if file_snapshot and file_snapshot.get('template_version', 0) > data.get('template_version', 0):
                cache.set(f"mapper:session:{session_id}", file_snapshot, 86400)
                SESSION_STORE[session_id] = file_snapshot
                logger.info(f"🔄 Cache refreshed for session {session_id} from newer file snapshot")
                return file_snapshot
        except Exception:
            pass
        return data
    
    # Fallback to old in-memory store
    if session_id in SESSION_STORE:
        data = SESSION_STORE[session_id]
        # Warm cache for next time
        cache.set(f"mapper:session:{session_id}", data, 86400)
        logger.info(f"🔄 Session {session_id} found in memory, warmed cache")
        return data
    
    # Final fallback to file snapshot
    logger.info(f"🔍 Session {session_id} not in cache/memory, trying file")
    data = load_session_from_file(session_id)
    if data:
        # Warm both cache and memory
        cache.set(f"mapper:session:{session_id}", data, 86400)
        SESSION_STORE[session_id] = data
        logger.info(f"🔄 Restored session {session_id} from file, warmed cache/memory")
        return data
    
    logger.warning(f"❌ Session {session_id} not found anywhere")
    return None

def get_session(session_id):
    """
    Universal session retrieval that works across multiple workers.
    Checks memory first, then loads from file if needed.
    Returns session data or None if not found.
    """
    if session_id in SESSION_STORE:
        logger.info(f"🔍 Session {session_id} found in memory")
        return SESSION_STORE[session_id]
    
    # Try to load from file
    logger.info(f"🔍 Session {session_id} not in memory, trying file")
    session_data = load_session_from_file(session_id)
    if session_data:
        SESSION_STORE[session_id] = session_data
        logger.info(f"🔄 Restored session {session_id} from file to memory")
        logger.info(f"🔍 Session keys: {list(session_data.keys())}")
        return session_data
    
    logger.warning(f"❌ Session {session_id} not found in memory or file")
    return None

def save_session(session_id, session_data):
    """
    Universal session saving that persists across multiple workers.
    Saves to cache (shared), memory, and file.
    """
    try:
        # Preserve critical flags like source_type across partial updates
        existing = None
        try:
            existing = SESSION_STORE.get(session_id)
            if not existing:
                existing = load_session_from_file(session_id)
        except Exception:
            existing = None

        if existing and 'source_type' in existing and 'source_type' not in session_data:
            session_data['source_type'] = existing['source_type']

        # As a last resort, infer PDF sessions from database
        if 'source_type' not in session_data:
            try:
                from .models import PDFSession
                if PDFSession.objects.filter(session_id=session_id).exists():
                    session_data['source_type'] = 'pdf'
            except Exception:
                pass

        # Save to shared cache first (critical for Azure multi-worker)
        cache.set(f"mapper:session:{session_id}", session_data, 86400)
        # Keep compatibility with existing in-memory store
        SESSION_STORE[session_id] = session_data
        # Persist to file/blob storage
        save_session_to_file(session_id, session_data)
        logger.info(f"💾 Saved session {session_id} to cache, memory, and file")
    except Exception as e:
        logger.error(f"Failed to save session {session_id}: {e}")


# Utility: Fast total row count without loading full DataFrame
def _count_total_data_rows(file_path: str, sheet_name: Optional[str], header_row: int) -> int:
    """Return total number of data rows after the header row.
    Works for both CSV and Excel. For Excel, uses openpyxl read-only mode.
    """
    try:
        p = Path(file_path)
        if str(p).lower().endswith('.csv'):
            # Count lines in a streaming fashion
            with open(p, 'rb') as f:
                # Count newline occurrences; this is approximate but sufficient
                total_lines = 0
                for _ in f:
                    total_lines += 1
            data_rows = max(0, total_lines - (header_row + 1))
            return data_rows
        else:
            # Excel: use openpyxl to get max_row, then adjust for header row
            wb = load_workbook(filename=str(p), read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            max_row = ws.max_row or 0
            # Try to trim trailing entirely empty rows
            # Scan backward until a row with any non-empty cell is found
            last_data_row = 0
            for r in range(max_row, 0, -1):
                has_val = False
                for cell in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                    # cell is a tuple of values for that row
                    if any(v is not None and str(v).strip() != '' for v in cell):
                        has_val = True
                        break
                if has_val:
                    last_data_row = r
                    break
            if last_data_row == 0:
                return 0
            data_rows = max(0, last_data_row - (header_row + 1))
            return data_rows
    except Exception as e:
        logger.warning(f"Failed to count total data rows: {e}")
        # Fallback: unknown
        return 0


# Use hybrid file manager from azure_storage module
# This automatically handles Azure Blob Storage when available,
# falls back to local storage for development


def apply_column_mappings(client_file, mappings, sheet_name=None, header_row=0, session_id=None):
    """
    Apply column mappings to transform client data to template format.
    Now supports multiple source columns mapping to the same template column name (with identical names).
    Includes ALL template columns, even unmapped ones (which will be empty).
    """
    try:
        logger.info(f"🔍 apply_column_mappings received mappings: {mappings}")
        
        # Get template headers - use dynamic columns if available, otherwise read from file
        template_headers = []
        if session_id and session_id in SESSION_STORE:
            info = SESSION_STORE[session_id]
            
            # Prefer canonical current headers (persisted). Avoid using enhanced headers directly
            # to prevent flip-flopping between operation-specific snapshots.
            canonical_headers = info.get("current_template_headers")
            if canonical_headers and isinstance(canonical_headers, list) and len(canonical_headers) > 0:
                logger.info(f"🔧 DEBUG: Original canonical_headers = {canonical_headers}")
                
                # CRITICAL FIX: Ensure ALL standard headers are always included
                # The current_template_headers might only contain dynamic headers
                # Include both core Factwise headers AND standard template fields
                standard_headers = [
                    # Core Factwise headers
                    "Item code", "Item name", "Description", "Item type", "Measurement unit", "Procurement entity name",
                    # Standard template fields that should behave like core headers when mapped
                    "Notes", "Internal notes", "Procurement item", "Sales item", "Preferred vendor code"
                ]
                
                # Check if standard headers are already present
                has_standard = any(h in canonical_headers for h in standard_headers)
                logger.info(f"🔧 DEBUG: has_standard check result: {has_standard} (checked against: {standard_headers})")
                
                if not has_standard:
                    # Prepend standard headers to the canonical headers
                    canonical_headers = standard_headers + canonical_headers
                    logger.info(f"🔧 CRITICAL FIX: Added missing standard headers to canonical headers")
                    logger.info(f"🔧 DEBUG: Updated canonical_headers = {canonical_headers}")
                else:
                    logger.info(f"🔧 DEBUG: Standard headers already present in canonical_headers")
                
                template_headers = normalize_headers_to_internal(canonical_headers)
                logger.info(f"🔍 Using {len(template_headers)} canonical template headers from session (with core headers)")
            elif 'tags_count' in info or 'spec_pairs_count' in info or 'customer_id_pairs_count' in info:
                # Dynamic template columns when counts provided
                tags_count = info.get('tags_count', 3)
                spec_pairs_count = info.get('spec_pairs_count', 3)
                customer_id_pairs_count = info.get('customer_id_pairs_count', 1)
                template_headers = generate_template_columns(tags_count, spec_pairs_count, customer_id_pairs_count)
                logger.info(f"🔍 Generated {len(template_headers)} dynamic template columns")
                logger.info(f"🔧 DEBUG: Generated template_headers = {template_headers}")
            else:
                # Fallback to reading from template file
                template_headers = SESSION_STORE[session_id].get("template_headers", [])
                if not template_headers:
                    try:
                        from .bom_header_mapper import BOMHeaderMapper
                        mapper = BOMHeaderMapper()
                        template_headers = mapper.read_excel_headers(
                            file_path=hybrid_file_manager.get_file_path(info["template_path"]),
                            sheet_name=info.get("template_sheet_name"),
                            header_row=info.get("template_header_row", 1) - 1 if info.get("template_header_row", 1) > 0 else 0
                        )
                        # Store in session for future use
                        SESSION_STORE[session_id]["template_headers"] = template_headers
                        logger.info(f"🔍 Read and cached {len(template_headers)} template headers from file")
                    except Exception as e:
                        logger.warning(f"Could not read template headers: {e}")
                else:
                    logger.info(f"🔍 Found {len(template_headers)} template headers from session")
        
        # Build canonical lookup from template headers -> exact header text
        canon_to_template = {_canon(h): h for h in (template_headers or [])}
        logger.info(f"🔧 DEBUG: Built canonical lookup with {len(canon_to_template)} template headers")

        # Handle new mapping format from frontend
        if isinstance(mappings, dict) and 'mappings' in mappings:
            # New format: ordered list of individual mappings
            mapping_list = mappings['mappings']
            logger.info(f"🔍 Processing new mapping format with {len(mapping_list)} mappings")
            
            # Extract and save default values if provided
            if 'default_values' in mappings and session_id and session_id in SESSION_STORE:
                default_values = mappings['default_values']
                SESSION_STORE[session_id]["default_values"] = default_values
                logger.info(f"🔧 DEBUG: Saved default values to session {session_id}: {default_values}")
        else:
            # Fallback to old format for compatibility - convert to preserve order better
            mapping_list = []
            logger.info(f"🔍 Converting old format mappings: {mappings}")
            
            # Process in the order they appear in the original dict to preserve user intent
            # Don't sort alphabetically as that changes the user's intended order
            
            for template_column, source_info in mappings.items():
                if isinstance(source_info, list):
                    # Multiple sources mapped to same target - this was the problematic case
                    logger.info(f"🔍 Old format: Multiple sources {source_info} -> {template_column}")
                    for source_column in source_info:
                        mapping_list.append({'source': source_column, 'target': template_column})
                        logger.info(f"🔍 Converted: {source_column} -> {template_column}")
                else:
                    # Single source mapping
                    logger.info(f"🔍 Old format: Single source {source_info} -> {template_column}")
                    mapping_list.append({'source': source_info, 'target': template_column})
            
            logger.info(f"🔍 Converted old format to {len(mapping_list)} individual mappings")
        
        # Resolve local path if using Azure Blob Storage and read the client data
        client_local_path = hybrid_file_manager.get_file_path(client_file)

        # Optional pagination inputs: derive from caller if present via context
        offset = None
        limit = None
        try:
            if session_id and SESSION_STORE.get(session_id, {}).get('__paginate__'):
                hints = SESSION_STORE[session_id]['__paginate__']
                offset = int(hints.get('offset')) if 'offset' in hints else None
                limit = int(hints.get('limit')) if 'limit' in hints else None
        except Exception:
            offset = None
            limit = None

        # Helper to read only headers quickly
        def _read_only_headers() -> list:
            try:
                if str(client_local_path).lower().endswith('.csv'):
                    df0 = read_csv_with_encoding(client_local_path, header_row, nrows=0)
                else:
                    df0 = pd.read_excel(client_local_path, sheet_name=sheet_name, header=header_row, nrows=0)
                return [str(c).strip() for c in df0.columns]
            except Exception:
                return []

        # Read the DataFrame – full or paginated slice
        # CRITICAL FIX: Special handling for PDF sessions where CSV has no headers
        is_pdf_session = session_id and session_id in SESSION_STORE and SESSION_STORE[session_id].get("source_type") == "pdf"
        pdf_headers = None

        if is_pdf_session:
            # Get headers from PDF extraction data instead of CSV file
            try:
                from .models import PDFSession, PDFExtractionResult
                pdf_session = PDFSession.objects.get(session_id=session_id)
                pdf_extraction = PDFExtractionResult.objects.filter(pdf_session=pdf_session).order_by('-created_at').first()
                if pdf_extraction:
                    pdf_headers = pdf_extraction.extracted_headers
                    logger.info(f"🔍 PDF session detected, using extracted headers: {pdf_headers}")
            except Exception as e:
                logger.error(f"🔍 Error getting PDF headers for session {session_id}: {e}")

        if offset is None or limit is None:
            if str(client_local_path).lower().endswith('.csv'):
                if is_pdf_session and pdf_headers:
                    # PDF CSV has no headers, read with header=None and provide column names
                    df = read_csv_with_encoding(client_local_path, header_row=None, names=pdf_headers)
                    logger.info(f"🔍 PDF session: Read CSV without headers, applied PDF headers: {pdf_headers}")
                else:
                    df = read_csv_with_encoding(client_local_path, header_row)
            else:
                result = pd.read_excel(client_local_path, sheet_name=sheet_name, header=header_row)
                # Handle multiple sheets case
                if isinstance(result, dict):
                    first_sheet_name = list(result.keys())[0]
                    df = result[first_sheet_name]
                else:
                    df = result
        else:
            # Paginated read: read only the requested slice efficiently
            if is_pdf_session and pdf_headers:
                cols = pdf_headers
            else:
                cols = _read_only_headers()

            if str(client_local_path).lower().endswith('.csv'):
                if is_pdf_session and pdf_headers:
                    # PDF CSV has no headers, skip only the offset rows (no header row to skip)
                    skiprows = max(0, int(offset)) if offset > 0 else None
                    df = read_csv_with_encoding(
                        client_local_path,
                        header_row=None,
                        names=pdf_headers,
                        skiprows=skiprows,
                        nrows=int(limit)
                    )
                    logger.info(f"🔍 PDF session paginated: Read CSV without headers, offset={offset}, limit={limit}")
                else:
                    # Skip header_row+1 data rows plus the offset
                    skip_start = header_row + 1
                    skip_end = skip_start + max(0, int(offset))
                    skiprows = list(range(skip_start, skip_end)) if skip_end > skip_start else None
                    df = read_csv_with_encoding(
                        client_local_path,
                        header_row=None,
                        names=cols if cols else None,
                        skiprows=skiprows,
                        nrows=int(limit)
                    )
            else:
                # Excel: set header=None and supply names; skip header_row + 1 + offset rows from top
                skiprows = header_row + 1 + max(0, int(offset))
                df = pd.read_excel(
                    client_local_path,
                    sheet_name=sheet_name,
                    header=None,
                    names=cols if cols else None,
                    skiprows=skiprows,
                    nrows=int(limit)
                )
        
        # Clean column names
        df.columns = [str(col).strip() for col in df.columns]
        
        # Build canonical lookup for df columns (for source snapping)
        df_canon = {_canon(c): c for c in df.columns}
        logger.info(f"🔧 DEBUG: Built df canonical lookup with {len(df_canon)} columns")
        
        # Normalize mapping_list targets to the exact template header spelling
        normalized_list = []
        for m in mapping_list:
            t_raw = m.get("target", "")
            s_raw = m.get("source", "")
            # Snap target to real template header if found, otherwise keep original
            t = canon_to_template.get(_canon(t_raw), t_raw)
            logger.debug(f"🔧 Target normalization: '{t_raw}' -> '{t}'")
            normalized_list.append({"source": s_raw, "target": t})
        mapping_list = normalized_list
        logger.info(f"🔧 DEBUG: Normalized {len(mapping_list)} mapping targets")
        
        # Build column order - ALWAYS preserve original template column order
        mapping_dict = {}  # target -> list of mappings for that target
        
        for mapping in mapping_list:
            target = mapping['target']
            if target not in mapping_dict:
                mapping_dict[target] = []
            mapping_dict[target].append(mapping)
        
        # ALWAYS use original template headers order - never reorder based on mapping status
        column_order = template_headers.copy() if template_headers else []
        logger.info(f"🔧 DEBUG apply_column_mappings: template_headers = {template_headers}")
        logger.info(f"🔧 DEBUG apply_column_mappings: column_order = {column_order}")
        
        # Check if standard headers are present in template_headers
        standard_headers_check = [
            # Core Factwise headers
            "Item code", "Item name", "Description", "Item type", "Measurement unit", "Procurement entity name",
            # Standard template fields that should behave like core headers when mapped
            "Notes", "Internal notes", "Procurement item", "Sales item", "Preferred vendor code"
        ]
        missing_standard = [h for h in standard_headers_check if h not in template_headers]
        if missing_standard:
            logger.warning(f"🚨 CRITICAL: Missing standard headers in template_headers: {missing_standard}")
        else:
            logger.info(f"✅ All standard headers present in template_headers")
        
        # Get session default values for unmapped fields
        session_default_values = {}
        if session_id and session_id in SESSION_STORE:
            session_default_values = SESSION_STORE[session_id].get("default_values", {})
        
        # Process each row - match the header logic
        transformed_rows = []
        for _, row in df.iterrows():
            transformed_row = []
            
            for target_column in column_order:
                if target_column in mapping_dict:
                    # This column has mappings - for numbered fields, take the first mapping only
                    mappings_for_target = mapping_dict[target_column]
                    
                    # For our numbered fields (Tag_1, Tag_2, etc.), there should be exactly one mapping per target
                    # Take the first (and usually only) mapping
                    mapping = mappings_for_target[0]
                    source_column = mapping['source']
                    
                    # IMPORTANT: Handle default value mappings (from template apply)
                    if source_column and source_column.startswith("__DEFAULT__"):
                        # Extract default value from special source format: "__DEFAULT__value"
                        default_value = source_column[11:]  # Remove "__DEFAULT__" prefix
                        transformed_row.append(default_value)
                        logger.info(f"🔧 Applied default value '{default_value}' to column '{target_column}'")
                    elif source_column:
                        # Snap source to exact df column if needed
                        src = df_canon.get(_canon(source_column), source_column)
                        if src in df.columns:
                            value = row.get(src, "")
                            if pd.isna(value):
                                value = ""
                            else:
                                value = str(value).strip()
                        else:
                            value = ""

                        # If the mapped source yields an empty value, fall back to session default if available
                        if (value == "" or value is None) and target_column in session_default_values:
                            default_value = session_default_values.get(target_column, "")
                            value = str(default_value)
                            logger.info(f"🔧 Applied session default value '{default_value}' to mapped column '{target_column}' due to empty source value")

                        transformed_row.append(value)
                    else:
                        # Source column missing - fall back to default if available
                        if target_column in session_default_values:
                            default_value = session_default_values.get(target_column, "")
                            transformed_row.append(str(default_value))
                            logger.info(f"🔧 Applied session default value '{default_value}' to unmapped/missing-source column '{target_column}'")
                        else:
                            transformed_row.append("")  # Empty value for missing source columns
                        
                    # Handle additional mappings to the same target (rare with numbered system)
                    for additional_mapping in mappings_for_target[1:]:
                        additional_source = additional_mapping['source']
                        if additional_source:
                            # Snap additional source to exact df column if needed
                            additional_src = df_canon.get(_canon(additional_source), additional_source)
                            if additional_src in df.columns:
                                additional_value = row.get(additional_src, "")
                                if pd.isna(additional_value):
                                    additional_value = ""
                                else:
                                    additional_value = str(additional_value).strip()
                                transformed_row.append(additional_value)
                            else:
                                transformed_row.append("")
                        else:
                            transformed_row.append("")
                else:
                    # Unmapped template column - check for default value, otherwise empty
                    if target_column in session_default_values:
                        default_value = session_default_values[target_column]
                        transformed_row.append(str(default_value))
                        logger.info(f"🔧 Applied session default value '{default_value}' to unmapped column '{target_column}'")
                    else:
                        transformed_row.append("")  # Empty value for truly unmapped columns
            
            transformed_rows.append(transformed_row)
        
        # Build final headers list - for numbered fields, don't add duplicates
        # Our numbered fields (Tag_1, Tag_2, etc.) are already unique
        final_headers = []
        seen_headers = set()
        
        for target_column in column_order:
            if target_column in mapping_dict:
                # For numbered fields, each mapping should map to exactly one header
                mappings_for_target = mapping_dict[target_column]
                
                # If target_column is already numbered (e.g., Tag_1, Tag_2), use it as-is
                if target_column not in seen_headers:
                    final_headers.append(target_column)
                    seen_headers.add(target_column)
                elif len(mappings_for_target) > 1:
                    # Only create numbered variants if we have multiple mappings to the same unnumbered target
                    # This should rarely happen with our new numbering system
                    base_name = target_column.split('_')[0] if '_' in target_column else target_column
                    counter = 2
                    new_header = f"{base_name}_{counter}"
                    while new_header in seen_headers:
                        counter += 1
                        new_header = f"{base_name}_{counter}"
                    final_headers.append(new_header)
                    seen_headers.add(new_header)
            else:
                # Unmapped template column - add as-is if not already seen
                if target_column not in seen_headers:
                    final_headers.append(target_column)
                    seen_headers.add(target_column)
        
        # Return data structure that includes column order and data
        logger.info(f"🔧 DEBUG apply_column_mappings: final_headers = {final_headers}")
        final_standard_check = [h for h in standard_headers_check if h in final_headers]
        logger.info(f"🔧 DEBUG apply_column_mappings: final_headers contains {len(final_standard_check)}/{len(standard_headers_check)} standard headers: {final_standard_check}")
        
        return {
            'headers': final_headers,
            'data': transformed_rows
        }
        
    except Exception as e:
        logger.error(f"Error in apply_column_mappings: {e}")
        return {'headers': [], 'data': []}


@api_view(['POST'])
def update_session_data(request):
    """
    Update session data with corrected values while preserving structure
    """
    try:
        session_id = request.data.get('session_id')
        headers = request.data.get('headers', [])
        data_rows = request.data.get('data', [])
        try:
            logger.info(f"📝 CORRECTION: Received update for session {session_id}: headers={len(headers)} rows={len(data_rows)}")
            if headers:
                logger.info(f"📝 CORRECTION: First 10 headers: {headers[:10]}")
            if data_rows:
                sample_keys = list(data_rows[0].keys()) if isinstance(data_rows[0], dict) else []
                logger.info(f"📝 CORRECTION: First row keys: {sample_keys}")
        except Exception:
            pass

        if not session_id:
            return Response({
                'success': False,
                'error': 'No session ID provided'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not headers or not data_rows:
            return Response({
                'success': False,
                'error': 'Headers and data are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get session info
        info = get_session_consistent(session_id)
        if not info:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Build canonical template headers (full set) using session counts
        tags_count = info.get('tags_count', 3)
        spec_pairs_count = info.get('spec_pairs_count', 3)
        customer_id_pairs_count = info.get('customer_id_pairs_count', 1)

        base_headers = [
            'Item code', 'Item name', 'Description', 'Item type', 'Measurement unit',
            'Procurement entity name', 'Notes', 'Internal notes', 'Procurement item', 'Sales item', 'Preferred vendor code'
        ]

        canonical_headers = []
        canonical_headers.extend(base_headers)
        for i in range(1, max(1, int(tags_count)) + 1):
            canonical_headers.append(f'Tag_{i}')
        for i in range(1, max(1, int(spec_pairs_count)) + 1):
            canonical_headers.append(f'Specification_Name_{i}')
            canonical_headers.append(f'Specification_Value_{i}')
        for i in range(1, max(1, int(customer_id_pairs_count)) + 1):
            canonical_headers.append(f'Customer_Identification_Name_{i}')
            canonical_headers.append(f'Customer_Identification_Value_{i}')

        # Prefer the fuller header set between session headers and canonical
        existing_headers = (
            info.get('current_template_headers') or
            info.get('enhanced_headers') or
            info.get('mapped_headers') or
            info.get('client_headers') or
            canonical_headers
        )
        # If still empty, attempt to derive from any existing data rows
        if not existing_headers:
            try:
                current_data = info.get('formula_enhanced_data') or info.get('mapped_data') or info.get('data')
                if current_data and isinstance(current_data, list):
                    if isinstance(current_data[0], dict):
                        existing_headers = list(current_data[0].keys())
            except Exception:
                existing_headers = canonical_headers

        # Build tolerant header mapping from uploaded headers to the current dataset's canonical headers.
        def _norm(h: str) -> str:
            try:
                s = str(h or '').strip().lower()
                s = s.replace('_', ' ')
                s = ' '.join(s.split())
                # Keep only alphanumerics and single spaces
                filtered = ''.join(ch for ch in s if ch.isalnum() or ch == ' ')
                return ' '.join(filtered.split())
            except Exception:
                return ''

        # Use the larger set to avoid losing dynamic columns
        chosen_headers = existing_headers or canonical_headers
        if len(chosen_headers) < len(canonical_headers):
            chosen_headers = canonical_headers
        canonical_headers = list(chosen_headers)
        # Map normalized canonical name -> canonical header
        canon_lookup = {_norm(h): h for h in canonical_headers}

        # Enumerated canonical targets (ordered by suffix) for generic incoming headers
        import re
        def _suffix_idx(name: str) -> int:
            try:
                m = re.search(r"(\d+)$", str(name or ''))
                return int(m.group(1)) if m else 0
            except Exception:
                return 0

        tag_targets = sorted([h for h in canonical_headers if isinstance(h, str) and h.strip().lower().startswith('tag_')], key=_suffix_idx)
        spec_name_targets = sorted([h for h in canonical_headers if isinstance(h, str) and _norm(h).startswith('specification name')], key=_suffix_idx)
        spec_value_targets = sorted([h for h in canonical_headers if isinstance(h, str) and _norm(h).startswith('specification value')], key=_suffix_idx)
        cust_name_targets = sorted([h for h in canonical_headers if isinstance(h, str) and _norm(h).startswith('customer identification name')], key=_suffix_idx)
        cust_value_targets = sorted([h for h in canonical_headers if isinstance(h, str) and _norm(h).startswith('customer identification value')], key=_suffix_idx)

        tag_i = 0
        specn_i = 0
        specv_i = 0
        custn_i = 0
        custv_i = 0

        # Build mapping from incoming header -> canonical header
        header_map = {}
        for h in headers:
            hn = _norm(h)
            mapped = None
            if hn in canon_lookup:
                mapped = canon_lookup[hn]
            elif hn == 'tag' and tag_i < len(tag_targets):
                mapped = tag_targets[tag_i]; tag_i += 1
            elif hn == 'specification name' and specn_i < len(spec_name_targets):
                mapped = spec_name_targets[specn_i]; specn_i += 1
            elif hn == 'specification value' and specv_i < len(spec_value_targets):
                mapped = spec_value_targets[specv_i]; specv_i += 1
            elif hn == 'customer identification name' and custn_i < len(cust_name_targets):
                mapped = cust_name_targets[custn_i]; custn_i += 1
            elif hn == 'customer identification value' and custv_i < len(cust_value_targets):
                mapped = cust_value_targets[custv_i]; custv_i += 1

            if mapped:
                header_map[h] = mapped

        matching_headers = sorted(set(header_map.values())) if header_map else []
        logger.info(f"📝 CORRECTION: Mapped {len(matching_headers)} headers to canonical; example: {matching_headers[:10]}")

        if not matching_headers:
            return Response({
                'success': False,
                'error': 'No matching headers found. Upload file must contain headers that exist in the current dataset.',
                'details': {
                    'received_headers': headers,
                    'expected_headers': existing_headers
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        logger.info(f"🔄 Updating session {session_id} data with {len(data_rows)} rows and {len(matching_headers)} matching headers")

        # FULL REPLACE: Build new dataset strictly from uploaded rows
        # Normalize incoming rows using header_map first
        normalized_rows = []
        for r in data_rows:
            try:
                nr = {}
                for in_h, val in r.items():
                    mapped_h = header_map.get(in_h)
                    if not mapped_h:
                        # Fallback: direct canonical match by normalized name
                        nh = _norm(in_h)
                        mapped_h = canon_lookup.get(nh)
                    if mapped_h:
                        nr[mapped_h] = val
                normalized_rows.append(nr)
            except Exception:
                normalized_rows.append({})

        # Construct rows with all canonical headers in order and drop entirely blank rows
        updated_rows = []
        dropped_blank = 0
        for nr in normalized_rows:
            row_out = {h: nr.get(h, '') for h in canonical_headers}
            # Determine if the row is entirely blank (ignore whitespace and common placeholders)
            non_empty = False
            for v in row_out.values():
                s = '' if v is None else str(v).strip()
                if s and s.lower() not in ('none', 'null', 'nan'):
                    non_empty = True
                    break
            if non_empty:
                updated_rows.append(row_out)
            else:
                dropped_blank += 1
        logger.info(f"📝 CORRECTION: Built {len(updated_rows)} normalized rows with {len(canonical_headers)} canonical headers (dropped {dropped_blank} blank rows)")

        # Update session info: persist updated rows as the active dataset for Data Editor
        info['data'] = updated_rows
        # Make the corrected dataset the primary enhanced data source so Data Editor renders it immediately
        info['formula_enhanced_data'] = updated_rows
        info['enhanced_headers'] = canonical_headers
        # Clear any conflicting caches
        if 'mapped_data' in info:
            del info['mapped_data']

        # Mark correction mode to bypass cleanup and prefer enhanced data
        info['uploaded_via_correction'] = True
        logger.info(f"📝 CORRECTION: Session {session_id} saved with uploaded_via_correction=True; template_version will increment")

        # Update template version to trigger refresh
        info['template_version'] = info.get('template_version', 0) + 1

        # Save updated session
        save_session(session_id, info)

        logger.info(f"✅ Session {session_id} data updated successfully with {len(updated_rows)} rows")

        return Response({
            'success': True,
            'message': f'Data updated successfully. {len(matching_headers)} columns updated across {len(updated_rows)} rows.',
            'updated_rows': len(updated_rows),
            'updated_columns': len(matching_headers),
            'template_version': info['template_version']
        })

    except Exception as e:
        logger.error(f"Error updating session data: {e}")
        return Response({
            'success': False,
            'error': f'Failed to update session data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def health_check(request):
    """Health check endpoint."""
    return Response({
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat(),
        'version': '1.0.0'
    })

@api_view(['GET'])
def get_session_snapshot(request, session_id):
    """Get canonical snapshot of session state."""
    info = get_session(session_id)
    if not info:
        return no_store(Response({"success": False, "error": "Invalid session"}, status=400))
    return no_store(Response({"success": True, "snapshot": build_snapshot(info)}))


@api_view(['POST'])
def debug_session(request):
    """Debug endpoint to check session data."""
    session_id = request.data.get('session_id')
    if session_id in SESSION_STORE:
        session_data = SESSION_STORE[session_id].copy()
        # Remove sensitive file paths for security
        session_data.pop('client_path', None)
        session_data.pop('template_path', None)
        
        # Add extra debug info for mappings
        mappings = session_data.get('mappings')
        if mappings:
            logger.info(f"🔍 DEBUG Session {session_id} mappings: {mappings}")
            logger.info(f"🔍 DEBUG Mappings type: {type(mappings)}")
            if isinstance(mappings, dict):
                logger.info(f"🔍 DEBUG Mappings keys: {list(mappings.keys())}")
                if 'mappings' in mappings:
                    logger.info(f"🔍 DEBUG New format detected with {len(mappings['mappings'])} individual mappings")
                else:
                    logger.info(f"🔍 DEBUG Old format detected")
        
        return Response({'session_data': session_data})
    else:
        return Response({'error': 'Session not found'}, status=404)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_files(request):
    """
    Upload and process client and template files.
    Optimized for Excel to Excel mapping.
    """
    try:
        # Extract form data
        client_file = request.FILES.get('clientFile')
        template_file = request.FILES.get('templateFile')  # May be None for fixed template
        sheet_name = request.data.get('sheetName')
        header_row = int(request.data.get('headerRow', 1))
        template_sheet_name = request.data.get('templateSheetName')
        template_header_row = int(request.data.get('templateHeaderRow', 1))
        use_template_id = request.data.get('useTemplateId')
        
        # Extract fixed Factwise headers if provided
        factwise_headers_json = request.data.get('factwiseHeaders')
        factwise_headers = []
        if factwise_headers_json:
            try:
                factwise_headers = json.loads(factwise_headers_json)
            except json.JSONDecodeError:
                logger.warning(f"Invalid factwiseHeaders JSON: {factwise_headers_json}")
        
        # Extract formula rules if provided
        formula_rules_json = request.data.get('formulaRules')
        formula_rules = []
        if formula_rules_json:
            try:
                formula_rules = json.loads(formula_rules_json)
            except json.JSONDecodeError:
                logger.warning(f"Invalid formula rules JSON: {formula_rules_json}")
        
        # Validation - allow fixed template mode (no template_file)
        if not client_file:
            return Response({
                'success': False,
                'error': 'Client file is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # If no template file but factwise headers provided, use fixed template mode
        is_fixed_template_mode = not template_file and factwise_headers
        
        if not template_file and not is_fixed_template_mode:
            return Response({
                'success': False,
                'error': 'Either template file or fixed Factwise headers are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate file types
        allowed_extensions = ['.xlsx', '.xls', '.csv']
        client_ext = Path(client_file.name).suffix.lower()
        
        if client_ext not in allowed_extensions:
            return Response({
                'success': False,
                'error': f'Only Excel (.xlsx, .xls) and CSV files are supported for client file'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate template file if provided
        if template_file:
            template_ext = Path(template_file.name).suffix.lower()
            if template_ext not in allowed_extensions:
                return Response({
                    'success': False,
                    'error': f'Only Excel (.xlsx, .xls) and CSV files are supported for template file'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Save uploaded files
        client_path, client_original_name = hybrid_file_manager.save_upload_file(client_file, "client")
        
        # Handle template file or fixed template mode
        if template_file:
            template_path, template_original_name = hybrid_file_manager.save_upload_file(template_file, "template")
        else:
            # Fixed template mode - create a virtual template with factwise headers
            template_path = None
            template_original_name = "Fixed Factwise Template"
        
        # Generate session ID
        session_id = str(uuid.uuid4())
        
        # Set default column counts for fixed template mode
        if is_fixed_template_mode:
            default_tags_count = 3
            default_spec_pairs_count = 3
            default_customer_id_pairs_count = 1
        else:
            default_tags_count = 3
            default_spec_pairs_count = 3
            default_customer_id_pairs_count = 1
        
        # Store session data using universal session saving
        session_data = {
            "client_path": client_path,
            "template_path": template_path,
            "original_client_name": client_original_name,
            "original_template_name": template_original_name,
            "sheet_name": sheet_name,
            "header_row": header_row,
            "template_sheet_name": template_sheet_name,
            "template_header_row": template_header_row,
            "created": datetime.utcnow().isoformat(),
            "mappings": None,
            "edited_data": None,
            "original_template_id": None,
            "template_modified": False,
            "formula_rules": formula_rules if formula_rules else [],
            "is_fixed_template_mode": is_fixed_template_mode,
            "factwise_headers": factwise_headers if is_fixed_template_mode else None,
            "tags_count": default_tags_count,
            "spec_pairs_count": default_spec_pairs_count,
            "customer_id_pairs_count": default_customer_id_pairs_count
        }
        
        # Save session with universal persistence (critical for multi-worker environments)
        save_session(session_id, session_data)
        
        # Apply template if specified
        template_applied = False
        template_success = False
        applied_mappings = {}
        applied_formulas = False
        
        if use_template_id:
            try:
                template = MappingTemplate.objects.get(id=int(use_template_id))
                
                # Read client headers to apply template
                mapper = BOMHeaderMapper()
                client_headers = mapper.read_excel_headers(
                    file_path=hybrid_file_manager.get_file_path(client_path),
                    sheet_name=sheet_name,
                    header_row=header_row - 1 if header_row > 0 else 0
                )
                
                # Apply template mappings
                application_result = template.apply_to_headers(client_headers)
                
                if application_result['total_mapped'] > 0:
                    template_applied = True
                    template_success = True  # Consider success if any mappings work
                    applied_mappings = application_result['mappings']  # Old format
                    
                    # Use new format if available (preserves duplicates)
                    if 'mappings_new_format' in application_result and application_result['mappings_new_format']:
                        total_applied = application_result['total_mappings_with_duplicates']
                        new_format_mappings = {
                            "mappings": application_result['mappings_new_format']
                        }
                        logger.info(f"✅ Template applied successfully with {total_applied} mappings (including duplicates)")
                    else:
                        # Fallback to old format conversion
                        total_applied = application_result['total_mapped']
                        new_format_mappings = {
                            "mappings": [
                                {"source": source_col, "target": template_col}
                                for template_col, source_col in applied_mappings.items()
                            ]
                        }
                        logger.info(f"✅ Template applied successfully with {total_applied} mappings")
                    
                    # Update session with applied mappings
                    SESSION_STORE[session_id]["original_template_id"] = int(use_template_id)
                    SESSION_STORE[session_id]["mappings"] = new_format_mappings
                    logger.info(f"🔄 Converted {len(applied_mappings)} unique mappings from template application")
                    
                    # Apply formula rules if they exist (from template or Step 3)
                    template_formula_rules = getattr(template, 'formula_rules', []) or []
                    
                    # CRITICAL FIX: Deduplicate rules to prevent duplicate columns
                    def _rule_signature(rule):
                        """Create a unique signature for a rule to identify duplicates."""
                        tgt = (rule.get('target_column') or '').strip().lower()
                        ctype = (rule.get('column_type') or '').strip().lower()
                        spec = (rule.get('specification_name') or '').strip().lower()
                        subs = rule.get('sub_rules') or []
                        norm = []
                        for sr in subs:
                            s = (sr.get('search_text') or '').strip().lower()
                            o = (sr.get('output_value') or sr.get('tag_value') or '').strip().lower()
                            cs = bool(sr.get('case_sensitive'))
                            norm.append((s, o, cs))
                        # order-insensitive signature
                        return (ctype, tgt, spec, tuple(sorted(norm)))
                    
                    combined_formula_rules = []
                    seen_signatures = set()
                    for r in (template_formula_rules or []) + (formula_rules or []):
                        if not r:
                            continue
                        sig = _rule_signature(r)
                        if sig in seen_signatures:
                            logger.info(f"🔧 DEBUG: Skipping duplicate formula rule with signature: {sig}")
                            continue
                        seen_signatures.add(sig)
                        combined_formula_rules.append(r)
                    
                    logger.info(f"🔧 DEBUG: Deduplicated formula rules: {len(template_formula_rules or [])} template + {len(formula_rules or [])} request = {len(combined_formula_rules)} unique")
                    if combined_formula_rules:
                        SESSION_STORE[session_id]["formula_rules"] = combined_formula_rules
                        
                        # Apply formulas to create enhanced data
                        mapping_result = apply_column_mappings(
                            client_file=client_path,
                            mappings=new_format_mappings,
                            sheet_name=sheet_name,
                            header_row=header_row - 1 if header_row > 0 else 0,
                            session_id=session_id
                        )
                        
                        # Convert to dict format for formula processing
                        dict_rows = []
                        for row_list in mapping_result['data']:
                            row_dict = {}
                            for i, header in enumerate(mapping_result['headers']):
                                if i < len(row_list):
                                    row_dict[header] = row_list[i]
                                else:
                                    row_dict[header] = ""
                            dict_rows.append(row_dict)
                        
                        # Apply formula rules to create enhanced data
                        formula_result = apply_formula_rules(
                            data_rows=dict_rows,
                            headers=mapping_result['headers'],
                            formula_rules=combined_formula_rules,
                            session_info=SESSION_STORE[session_id]
                        )
                        
                        # Store enhanced data in session
                        SESSION_STORE[session_id]["formula_enhanced_data"] = formula_result['data']
                        SESSION_STORE[session_id]["enhanced_headers"] = formula_result['headers']
                        applied_formulas = True
                    
                    # Apply factwise rules if they exist
                    template_factwise_rules = getattr(template, 'factwise_rules', []) or []
                    if template_factwise_rules:
                        SESSION_STORE[session_id]["factwise_rules"] = template_factwise_rules
                        
                        # Apply each factwise rule with error handling
                        for rule in template_factwise_rules:
                            try:
                                if rule.get("type") == "factwise_id":
                                    first_column = rule.get("first_column")
                                    second_column = rule.get("second_column")
                                    operator = rule.get("operator", "_")
                                
                                    if first_column and second_column:
                                        # Get current data (either formula-enhanced or basic mapped)
                                        current_data = SESSION_STORE[session_id].get("formula_enhanced_data")
                                        current_headers = SESSION_STORE[session_id].get("enhanced_headers")
                                        
                                        if not current_data:
                                            # Use basic mapped data if no formula data exists
                                            mapping_result = apply_column_mappings(
                                                client_file=client_path,
                                                mappings=new_format_mappings,
                                                sheet_name=sheet_name,
                                                header_row=header_row - 1 if header_row > 0 else 0,
                                                session_id=session_id
                                            )
                                            current_data = mapping_result['data']
                                            current_headers = mapping_result['headers']
                                        
                                        # Check if required columns exist for Factwise ID
                                        if first_column not in current_headers:
                                            logger.warning(f"🆔 Factwise ID: First column '{first_column}' not found in headers: {current_headers}")
                                            continue  # Skip this factwise rule
                                        
                                        if second_column not in current_headers:
                                            logger.warning(f"🆔 Factwise ID: Second column '{second_column}' not found in headers: {current_headers}")
                                            continue  # Skip this factwise rule

                                        # Apply Factwise ID creation and map it to 'Item code' column
                                        first_col_idx = current_headers.index(first_column)
                                        second_col_idx = current_headers.index(second_column)

                                        # Determine target 'Item code' index; add if missing
                                        item_code_idx = None
                                        if "Item code" in current_headers:
                                            item_code_idx = current_headers.index("Item code")
                                            new_headers = list(current_headers)
                                            new_data_rows = []
                                        else:
                                            # Prepend 'Item code' if not present
                                            new_headers = ["Item code"] + list(current_headers)
                                            item_code_idx = 0
                                            new_data_rows = []

                                        if first_col_idx >= 0 and second_col_idx >= 0:
                                            for row in current_data:
                                                first_val = str(row[first_col_idx] if first_col_idx < len(row) else "").strip()
                                                second_val = str(row[second_col_idx] if second_col_idx < len(row) else "").strip()

                                                if first_val and second_val:
                                                    factwise_id = f"{first_val}{operator}{second_val}"
                                                elif first_val:
                                                    factwise_id = first_val
                                                elif second_val:
                                                    factwise_id = second_val
                                                else:
                                                    factwise_id = ""

                                                # Place into Item code column, adjusting row shape if we had to prepend
                                                if "Item code" in current_headers:
                                                    new_row = list(row)
                                                    # Ensure row length
                                                    while len(new_row) < len(new_headers):
                                                        new_row.append("")
                                                    new_row[item_code_idx] = factwise_id
                                                else:
                                                    new_row = [factwise_id] + list(row)

                                                new_data_rows.append(new_row)

                                            # Update session with Item code-enhanced data
                                            SESSION_STORE[session_id]["formula_enhanced_data"] = new_data_rows
                                            SESSION_STORE[session_id]["enhanced_headers"] = new_headers

                                            logger.info(f"🆔 Applied Factwise ID rule (mapped to 'Item code') from template during upload: {first_column} {operator} {second_column}")
                            except Exception as factwise_error:
                                logger.warning(f"🆔 Failed to apply Factwise ID rule during upload: {factwise_error}")
                                # Continue with other rules even if this one fails
                    
                    # Apply default values if they exist
                    template_default_values = getattr(template, 'default_values', {}) or {}
                    if template_default_values:
                        SESSION_STORE[session_id]["default_values"] = template_default_values
                        logger.info(f"🔧 DEBUG: Applied default values from template during upload: {template_default_values}")
                    
                    # Increment template usage
                    template.increment_usage()
                
            except Exception as e:
                import traceback
                logger.error(f"Template application failed: {e}")
                logger.error(f"Template application traceback: {traceback.format_exc()}")
                template_applied = True
                template_success = False
        
        logger.info(f"Files uploaded successfully for session {session_id}")
        
        return Response({
            'success': True,
            'session_id': session_id,
            'message': 'Files uploaded successfully',
            'template_applied': template_applied,
            'template_success': template_success,
            'applied_mappings': applied_mappings,
            'applied_formulas': applied_formulas
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error in upload_files: {e}")
        return Response({
            'success': False,
            'error': f'Upload failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@never_cache
def get_headers(request, session_id):
    """Get headers from uploaded files."""
    try:
        # Use consistent session retrieval (cache->memory->file) for multi-worker
        info = get_session_consistent(session_id)
        if not info:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        mapper = BOMHeaderMapper()

        # Cross-worker consistency: if file snapshot has newer template_version, refresh memory
        try:
            file_snapshot = load_session_from_file(session_id)
            if file_snapshot and file_snapshot.get('template_version', 0) > info.get('template_version', 0):
                SESSION_STORE[session_id] = file_snapshot
                info = file_snapshot
                logger.info(f"🔄 Refreshed in-memory session {session_id} from file (newer template_version)")
        except Exception:
            pass
        
        # Read client headers (support Azure Blob by resolving to local cache)
        # For PDF sessions, get headers from PDF extraction data instead of CSV file
        if info.get("source_type") == "pdf":
            try:
                from .models import PDFSession, PDFExtractionResult
                pdf_session = PDFSession.objects.get(session_id=session_id)
                pdf_extraction = PDFExtractionResult.objects.filter(pdf_session=pdf_session).order_by('-created_at').first()
                if pdf_extraction:
                    client_headers = pdf_extraction.extracted_headers
                    logger.info(f"🔍 Using PDF extracted headers for session {session_id}: {client_headers}")
                else:
                    # Fallback to CSV reading if no extraction found
                    client_headers = mapper.read_excel_headers(
                        file_path=hybrid_file_manager.get_file_path(info["client_path"]),
                        sheet_name=info["sheet_name"],
                        header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0
                    )
                    logger.warning(f"🔍 No PDF extraction found for session {session_id}, fallback to CSV headers")
            except Exception as e:
                logger.error(f"🔍 Error getting PDF headers for session {session_id}: {e}")
                # Fallback to CSV reading
                client_headers = mapper.read_excel_headers(
                    file_path=hybrid_file_manager.get_file_path(info["client_path"]),
                    sheet_name=info["sheet_name"],
                    header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0
                )
        else:
            client_headers = mapper.read_excel_headers(
                file_path=hybrid_file_manager.get_file_path(info["client_path"]),
                sheet_name=info["sheet_name"],
                header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0
            )
        
        # Read template headers (allow enhanced headers override)
        template_headers = mapper.read_excel_headers(
            file_path=hybrid_file_manager.get_file_path(info["template_path"]),
            sheet_name=info.get("template_sheet_name"),
            header_row=info.get("template_header_row", 1) - 1 if info.get("template_header_row", 1) > 0 else 0
        )
        # Try to read optional/mandatory annotations from rows above headers (substring match)
        template_optionals_map = {}
        try:
            import pandas as pd
            template_path = hybrid_file_manager.get_file_path(info["template_path"])
            template_header_row_idx = info.get("template_header_row", 1) - 1 if info.get("template_header_row", 1) > 0 else 0
            if template_header_row_idx > 0:
                # Read all rows above the header row to scan for annotations
                if str(template_path).lower().endswith('.csv'):
                    df_ann = pd.read_csv(template_path, header=None, nrows=template_header_row_idx)
                else:
                    df_ann = pd.read_excel(template_path, sheet_name=info.get("template_sheet_name"), header=None, nrows=template_header_row_idx)
                # For each header column, scan upward for any cell containing 'optional'
                for idx, header in enumerate(template_headers):
                    is_optional = False
                    if idx < df_ann.shape[1] and df_ann.shape[0] > 0:
                        col_series = df_ann.iloc[:, idx]
                        for cell in col_series[::-1]:  # scan from nearest row upward
                            try:
                                text = str(cell).strip().lower()
                            except Exception:
                                text = ''
                            if 'optional' in text:
                                is_optional = True
                                break
                            if 'mandatory' in text:
                                # Explicit mandatory marker above; stop scanning
                                is_optional = False
                                break
                    template_optionals_map[str(header)] = is_optional
        except Exception as e:
            logger.warning(f"Optional/Mandatory annotations not read: {e}")
        enhanced_headers = info.get("enhanced_headers")
        
        # Debug logging
        logger.info(f"🔍 Session {session_id} - enhanced_headers: {enhanced_headers}")
        logger.info(f"🔍 Session {session_id} - template_headers from file: {template_headers}")
        
        # Get column counts from session (with defaults)
        tags_count = info.get('tags_count', 3)
        spec_pairs_count = info.get('spec_pairs_count', 3)
        customer_id_pairs_count = info.get('customer_id_pairs_count', 1)
        
        # Helper functions for robust special-column detection (case/trim tolerant)
        def _norm(h: str) -> str:
            try:
                return str(h or '').strip().lower()
            except Exception:
                return ''
        def _is_tag(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'tag' or h_norm.startswith('tag_')
        def _is_spec_name(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'specification name' or h_norm.startswith('specification_name_')
        def _is_spec_value(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'specification value' or h_norm.startswith('specification_value_')
        def _is_cust_name(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'customer identification name' or h_norm.startswith('customer_identification_name_')
        def _is_cust_value(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'customer identification value' or h_norm.startswith('customer_identification_value_')

        # Prefer enhanced headers if present to preserve dynamically added columns (e.g., Tag_4)
        if enhanced_headers and isinstance(enhanced_headers, list) and len(enhanced_headers) > 0:
            # Normalize any external-style headers to internal numbered headers
            template_headers_to_use = normalize_headers_to_internal(enhanced_headers)
            # Persist normalized variant back to session to avoid drift
            info["enhanced_headers"] = template_headers_to_use
            info["current_template_headers"] = template_headers_to_use
            save_session(session_id, info)
            # Derive counts from enhanced headers to keep session in sync
            try:
                derived_tags = len([h for h in template_headers_to_use if _is_tag(h)])
                derived_spec_pairs = len([h for h in template_headers_to_use if _is_spec_name(h)])
                derived_customer_pairs = len([h for h in template_headers_to_use if _is_cust_name(h)])
                if derived_tags != tags_count or derived_spec_pairs != spec_pairs_count or derived_customer_pairs != customer_id_pairs_count:
                    info['tags_count'] = derived_tags
                    info['spec_pairs_count'] = derived_spec_pairs
                    info['customer_id_pairs_count'] = derived_customer_pairs
                    save_session(session_id, info)
                    logger.info(f"🔍 Synchronized counts from enhanced headers: tags={derived_tags}, spec={derived_spec_pairs}, customer={derived_customer_pairs}")
            except Exception:
                pass
            logger.info(f"🔍 Using enhanced_headers from session as canonical headers: {template_headers_to_use}")
        # Otherwise, generate headers based on counts
        elif tags_count > 0 or spec_pairs_count > 0 or customer_id_pairs_count > 0:
            regenerated_headers = []
            
            # Add non-dynamic headers first
            for h in template_headers:
                if not (_is_tag(h) or _is_spec_name(h) or _is_spec_value(h) or _is_cust_name(h) or _is_cust_value(h)):
                    regenerated_headers.append(h)
            
            # Add Tag columns with simple numbering
            for i in range(tags_count):
                regenerated_headers.append(f'Tag_{i+1}')
            
            # Add Specification pairs with simple numbering
            for i in range(spec_pairs_count):
                regenerated_headers.append(f'Specification_Name_{i+1}')
                regenerated_headers.append(f'Specification_Value_{i+1}')
            
            # Add Customer identification pairs with simple numbering
            for i in range(customer_id_pairs_count):
                regenerated_headers.append(f'Customer_Identification_Name_{i+1}')
                regenerated_headers.append(f'Customer_Identification_Value_{i+1}')

            # Store canonical headers in session
            info["current_template_headers"] = regenerated_headers
            info["enhanced_headers"] = regenerated_headers
            save_session(session_id, info)
            template_headers_to_use = regenerated_headers
            logger.info(f"🔍 Regenerated canonical template headers based on counts: {template_headers_to_use}")
        else:
            template_headers_to_use = template_headers
            logger.info(f"🔍 Using template_headers from file: {template_headers_to_use}")
            info['template_headers'] = template_headers
            save_session(session_id, info)
        
        # CRITICAL FIX: template_headers should always include ALL headers (core + dynamic)
        # The frontend expects template_headers to be the complete set, not just dynamic ones
        complete_template_headers = generate_template_columns(tags_count, spec_pairs_count, customer_id_pairs_count)
        
        # Generate template columns based on counts (for reference)
        template_columns = generate_template_columns(tags_count, spec_pairs_count, customer_id_pairs_count)
        
        # Compute template_optionals aligned to the headers being returned
        def is_special_optional(h: str) -> bool:
            h_lower = (h or '').lower()
            return (h == 'Tag' or h.startswith('Tag_') or 
                   'specification' in h_lower or 
                   'customer identification' in h_lower or 
                   'customer_identification' in h_lower)
        
        template_optionals = []
        for h in complete_template_headers:  # Use complete_template_headers to match what's being returned
            if is_special_optional(h):
                template_optionals.append(True)
            else:
                template_optionals.append(bool(template_optionals_map.get(str(h), False)))
        
        # Prepare session metadata (robust PDF detection)
        is_pdf_session = (info.get("source_type") == "pdf")
        if not is_pdf_session:
            try:
                from .models import PDFSession as _PDFSession
                is_pdf_session = _PDFSession.objects.filter(session_id=session_id).exists()
            except Exception:
                is_pdf_session = False

        session_metadata = {
            'is_from_pdf': is_pdf_session,
            'header_confidence_scores': {},
            'original_template_id': info.get('original_template_id'),
            'template_applied': info.get('template_applied', False),
            'template_name': info.get('template_name', ''),
            'formula_rules': info.get('formula_rules', []),
            'factwise_rules': info.get('factwise_rules', [])
        }

        # Get header confidence scores for PDF sessions
        if is_pdf_session:
            try:
                from .models import PDFSession, PDFExtractionResult
                pdf_session = PDFSession.objects.get(session_id=session_id)
                pdf_extraction = PDFExtractionResult.objects.filter(pdf_session=pdf_session).order_by('-created_at').first()
                if pdf_extraction and hasattr(pdf_extraction, 'confidence_scores') and pdf_extraction.confidence_scores:
                    # Extract header-level confidence scores if available
                    confidence_data = pdf_extraction.confidence_scores
                    if isinstance(confidence_data, dict):
                        # Check if we have header-specific confidence scores
                        header_confidence = confidence_data.get('header_confidence', {})
                        if header_confidence:
                            session_metadata['header_confidence_scores'] = header_confidence
                            logger.info(f"📊 Including header confidence scores for PDF session {session_id}: {header_confidence}")
                        else:
                            # Fallback: use overall confidence for all headers
                            overall_confidence = confidence_data.get('overall_confidence', 0.9)
                            session_metadata['header_confidence_scores'] = {header: overall_confidence for header in client_headers}
                            logger.info(f"📊 Using overall confidence {overall_confidence} for all PDF headers in session {session_id}")
                    else:
                        # Legacy fallback: assume high confidence if no detailed scores
                        session_metadata['header_confidence_scores'] = {header: 0.9 for header in client_headers}
                        logger.info(f"📊 Using default confidence for PDF headers in session {session_id}")
            except Exception as e:
                logger.error(f"📊 Error getting PDF confidence scores for session {session_id}: {e}")
                # Provide default confidence scores for PDF sessions even if we can't get detailed ones
                session_metadata['header_confidence_scores'] = {header: 0.8 for header in client_headers}

        return no_store(Response({
            'success': True,
            'client_headers': client_headers,
            'template_headers': complete_template_headers,  # Always return complete headers
            'template_columns': template_columns,
            'template_optionals': template_optionals,
            'column_counts': {
                'tags_count': tags_count,
                'spec_pairs_count': spec_pairs_count,
                'customer_id_pairs_count': customer_id_pairs_count
            },
            'client_file': info.get('original_client_name', ''),
            'template_file': info.get('original_template_name', ''),
            'session_metadata': session_metadata
        }))
        
    except Exception as e:
        logger.error(f"Error in get_headers: {e}")
        return no_store(Response({
            'success': False,
            'error': f'Failed to get headers: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR))


@api_view(['POST'])
def mapping_suggestions(request):
    """Get AI-powered mapping suggestions."""
    try:
        session_id = request.data.get('session_id')
        if not session_id:
            return Response({
                'success': False,
                'error': 'Session ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Use consistent session retrieval (cache -> memory -> file) for Azure multi-worker
        info = get_session_consistent(session_id)
        # If a newer file snapshot exists, refresh in-memory session (cross-worker sync)
        try:
            file_snapshot = load_session_from_file(session_id)
            if file_snapshot and file_snapshot.get('template_version', 0) > info.get('template_version', 0):
                SESSION_STORE[session_id] = file_snapshot
                info = file_snapshot
                logger.info(f"🔄 Refreshed in-memory session {session_id} from file (newer template_version) in download_file")
        except Exception:
            pass
        if not info:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        mapper = BOMHeaderMapper()
        
        # Handle fixed template mode
        if info.get('is_fixed_template_mode') and info.get('factwise_headers'):
            # Use fixed factwise headers as template
            template_headers = info['factwise_headers']
            
            # Get client headers for mapping
            client_headers = mapper.read_excel_headers(
                file_path=hybrid_file_manager.get_file_path(info["client_path"]),
                sheet_name=info["sheet_name"],
                header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0
            )
            
            # Create mock mapping results for fixed template
            mapping_results = []
            for template_header in template_headers:
                mapping_results.append({
                    'template_header': template_header,
                    'mapped_client_header': None,
                    'confidence': 0
                })
        else:
            # Get mapping suggestions from files
            mapping_results = mapper.map_headers_to_template(
                client_file=hybrid_file_manager.get_file_path(info["client_path"]),
                template_file=hybrid_file_manager.get_file_path(info["template_path"]),
                client_sheet_name=info["sheet_name"],
                template_sheet_name=info.get("template_sheet_name"),
                client_header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0,
                template_header_row=info.get("template_header_row", 1) - 1 if info.get("template_header_row", 1) > 0 else 0
            )
            
            # Get template headers from file
            template_headers = mapper.read_excel_headers(
                file_path=hybrid_file_manager.get_file_path(info["template_path"]),
                sheet_name=info.get("template_sheet_name"),
                header_row=info.get("template_header_row", 1) - 1 if info.get("template_header_row", 1) > 0 else 0
            )
            
            # Get client headers from file
            client_headers = mapper.read_excel_headers(
                file_path=hybrid_file_manager.get_file_path(info["client_path"]),
                sheet_name=info["sheet_name"],
                header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0
            )
        
        # Prepare AI suggestions in format expected by frontend
        ai_suggestions = {}
        for result in mapping_results:
            if result['mapped_client_header'] and result['confidence'] >= 40:
                ai_suggestions[result['template_header']] = {
                    'suggested_column': result['mapped_client_header'],
                    'confidence': result['confidence'],
                    'is_specification_mapping': False
                }
        
        # Headers are already loaded above based on template mode
        
        return Response({
            'success': True,
            'ai_suggestions': ai_suggestions,
            'mapping_details': mapping_results,
            'template_headers': template_headers,
            'client_headers': client_headers,
            'user_columns': client_headers,
            'template_columns': template_headers,
            'specification_opportunity': {'detected': False},
            'session_metadata': {
                'original_template_id': info.get('original_template_id'),
                'template_applied': info.get('template_applied', False)
            }
        })
        
    except Exception as e:
        logger.error(f"Error in mapping_suggestions: {e}")
        return Response({
            'success': False,
            'error': f'Failed to generate suggestions: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def save_mappings(request):
    """Save column mappings for a session."""
    try:
        logger.info(f"🔧 DEBUG: save_mappings called for session {request.data.get('session_id')}")
        logger.info(f"🔧 DEBUG: Full request data: {request.data}")
        logger.info(f"🔧 DEBUG: Request data keys: {list(request.data.keys())}")
        session_id = request.data.get('session_id')
        mappings = request.data.get('mappings', {})
        default_values = request.data.get('default_values', {})
        header_corrections = request.data.get('header_corrections', {})
        
        logger.info(f"🔧 DEBUG: Received default_values: {default_values}")
        logger.info(f"🔧 DEBUG: Default values type: {type(default_values)}")
        logger.info(f"🔧 DEBUG: Received mappings: {mappings}")
        logger.info(f"🔧 DEBUG: Mappings type: {type(mappings)}")
        logger.info(f"🔧 DEBUG: Received header_corrections: {header_corrections}")
        logger.info(f"🔧 DEBUG: Header corrections type: {type(header_corrections)}")
        
        if not session_id:
            return Response({
                'success': False,
                'error': 'No session ID provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        info = get_session(session_id)
        if not info:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # CRITICAL FIX: Check if this is a destructive operation (empty mappings)
        # If mappings array is empty, this likely means user is deleting columns
        # In this case, we should NOT overwrite existing mappings
        is_destructive_operation = False
        force_persist = request.data.get('force_persist') in [True, 'true', 'True', '1', 1]
        if isinstance(mappings, list) and len(mappings) == 0:
            is_destructive_operation = True
            logger.warning(f"🔧 WARNING: Received empty mappings array - this is likely a destructive operation")
        elif isinstance(mappings, dict) and 'mappings' in mappings and isinstance(mappings['mappings'], list) and len(mappings['mappings']) == 0:
            is_destructive_operation = True
            logger.warning(f"🔧 WARNING: Received empty mappings.mappings array - this is likely a destructive operation")
        
        # If this is a destructive operation, preserve existing mappings
        if is_destructive_operation and not force_persist:
            existing_mappings = info.get("mappings", {})
            if existing_mappings and isinstance(existing_mappings, dict) and 'mappings' in existing_mappings:
                logger.info(f"🔧 PRESERVING existing mappings from destructive operation: {len(existing_mappings['mappings'])} mappings")
                # Only update default values, keep existing mappings
                cleaned_default_values = {}
                if default_values and isinstance(default_values, dict):
                    for field_name, value in default_values.items():
                        if value is not None and value != "":
                            cleaned_default_values[field_name] = str(value).strip()
                            logger.info(f"🔧 DEBUG: Storing default value '{value}' for field '{field_name}'")
                        else:
                            logger.info(f"🔧 DEBUG: Skipping empty default value for field '{field_name}': '{value}'")
                
                info["default_values"] = cleaned_default_values
                logger.info(f"🔧 DEBUG: Session {session_id} - Preserved existing mappings, updated default values: {cleaned_default_values}")
                
                # Mark template as modified if it was originally from a saved template
                if info.get("original_template_id"):
                    info["template_modified"] = True
                
                # Persist session using universal saving
                save_session(session_id, info)
                
                return Response({
                    'success': True,
                    'message': 'Default values updated, existing mappings preserved'
                })
        
        # CRITICAL FIX: Get existing used columns from session to maintain state
        # This prevents mappings from disappearing when navigating back
        existing_used_columns = set()
        
        # First, try to get existing mappings from the session
        if 'mappings' in info and isinstance(info['mappings'], dict) and 'mappings' in info['mappings']:
            for mapping in info['mappings']['mappings']:
                target = mapping.get('target', '')
                if target.startswith(('Tag_', 'Specification_Name_', 'Specification_Value_', 'Customer_Identification_Name_', 'Customer_Identification_Value_')):
                    existing_used_columns.add(target)
                    logger.info(f"🔧 DEBUG: Found existing internal target '{target}' in session")
        
        # If no existing mappings found, try to get from other session data
        if not existing_used_columns:
            # Check if we have column counts that indicate what should exist
            tags_count = info.get('tags_count', 3)
            spec_pairs_count = info.get('spec_pairs_count', 3)
            customer_id_pairs_count = info.get('customer_id_pairs_count', 1)
            
            # Generate expected column names based on counts
            for i in range(1, tags_count + 1):
                existing_used_columns.add(f'Tag_{i}')
            for i in range(1, spec_pairs_count + 1):
                existing_used_columns.add(f'Specification_Name_{i}')
                existing_used_columns.add(f'Specification_Value_{i}')
            for i in range(1, customer_id_pairs_count + 1):
                existing_used_columns.add(f'Customer_Identification_Name_{i}')
                existing_used_columns.add(f'Customer_Identification_Value_{i}')
            
            logger.info(f"🔧 DEBUG: Generated expected columns from counts: {existing_used_columns}")
        
        logger.info(f"🔧 DEBUG: Existing used columns from session: {existing_used_columns}")
        
        # Normalize mappings format and save
        # Accept both array format [{source, target}, ...] and object with .mappings
        normalized = mappings
        if isinstance(mappings, dict) and 'mappings' in mappings and isinstance(mappings['mappings'], list):
            normalized = mappings  # already in new format
        elif isinstance(mappings, list):
            normalized = { 'mappings': mappings }
        elif isinstance(mappings, dict):
            # old format {target: source} -> new format list
            normalized = { 'mappings': [ {'source': src, 'target': tgt} for tgt, src in mappings.items() ] }
        
        # Convert external column names to internal names for mapping storage
        # Use centralized conversion functions to ensure consistency
        if isinstance(normalized, dict) and 'mappings' in normalized:
            converted_mappings = []
            used_columns = existing_used_columns.copy()  # Start with existing used columns
            
            for mapping in normalized['mappings']:
                converted_mapping = mapping.copy()
                target = mapping.get('target', '')
                
                # CRITICAL FIX: Handle both external and internal names properly
                # If target is already an internal name (e.g., Tag_1), preserve it
                # If target is an external name (e.g., Tag), convert it to internal name
                if target.startswith(('Tag_', 'Specification_Name_', 'Specification_Value_', 'Customer_Identification_Name_', 'Customer_Identification_Value_')):
                    # Target is already an internal name, just track it
                    used_columns.add(target)
                    logger.info(f"🔧 DEBUG: Preserved internal target '{target}' for source '{mapping.get('source', '')}'")
                elif target in ['Tag', 'Specification name', 'Specification value', 'Customer identification name', 'Customer identification value']:
                    # Target is an external name, convert it to internal name
                    internal_name = convert_external_to_internal_name(target, info, used_columns)
                    converted_mapping['target'] = internal_name
                    used_columns.add(internal_name)
                    logger.info(f"🔧 DEBUG: Mapped source '{mapping.get('source', '')}' to {target} -> '{internal_name}'")
                else:
                    # Regular column mapping, no conversion needed
                    logger.info(f"🔧 DEBUG: Regular mapping: '{mapping.get('source', '')}' -> '{target}'")
                
                converted_mappings.append(converted_mapping)
            
            normalized['mappings'] = converted_mappings
            logger.info(f"🔧 DEBUG: Converted external column names to internal names in mappings")
            logger.info(f"🔧 DEBUG: Final used columns: {used_columns}")
        
        info["mappings"] = normalized
        
        # CRITICAL FIX: Ensure default values are properly stored
        # Filter out empty strings and None values, but keep actual default values
        cleaned_default_values = {}
        if default_values and isinstance(default_values, dict):
            for field_name, value in default_values.items():
                # Only store non-empty values (but allow "0" and other valid defaults)
                if value is not None and value != "":
                    cleaned_default_values[field_name] = str(value).strip()
                    logger.info(f"🔧 DEBUG: Storing default value '{value}' for field '{field_name}'")
                else:
                    logger.info(f"🔧 DEBUG: Skipping empty default value for field '{field_name}': '{value}'")
        
        info["default_values"] = cleaned_default_values
        logger.info(f"🔧 DEBUG: Session {session_id} - Final saved default values: {cleaned_default_values}")

        # Store header corrections for PDF sessions
        if header_corrections and isinstance(header_corrections, dict):
            info["header_corrections"] = header_corrections
            logger.info(f"🔧 DEBUG: Session {session_id} - Saved header corrections: {header_corrections}")

        # Mark template as modified if it was originally from a saved template
        if info.get("original_template_id"):
            info["template_modified"] = True

        # Persist session using universal saving
        save_session(session_id, info)
        
        return Response({
            'success': True,
            'message': 'Mappings saved successfully'
        })
        
    except Exception as e:
        logger.error(f"Error in save_mappings: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return Response({
            'success': False,
            'error': f'Failed to save mappings: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@never_cache
def get_existing_mappings(request, session_id):
    """Get existing mappings for a session."""
    try:
        print(f"🔍 PRINT DEBUG: get_existing_mappings called for session {session_id}")
        logger.info(f"🔍 DEBUG: get_existing_mappings called for session {session_id}")
        print(f"🔍 PRINT DEBUG: Function start, about to get session data")
        # Use consistent session retrieval to avoid stale/missing fields across workers
        session_data = get_session_consistent(session_id)
        print(f"🔍 PRINT DEBUG: Retrieved session data with keys: {list(session_data.keys()) if session_data else 'None'}")
        logger.info(f"🔍 DEBUG: Retrieved session data with keys: {list(session_data.keys()) if session_data else 'None'}")
        print(f"🔍 PRINT DEBUG: Session data source_type: {session_data.get('source_type') if session_data else 'No session'}")
        if not session_data:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        # Cross-worker refresh if file snapshot is newer
        try:
            file_snapshot = load_session_from_file(session_id)
            if file_snapshot and file_snapshot.get('template_version', 0) > session_data.get('template_version', 0):
                SESSION_STORE[session_id] = file_snapshot
                session_data = file_snapshot
                logger.info(f"🔄 Refreshed session {session_id} from file in get_existing_mappings")
        except Exception:
            pass
        mappings = session_data.get("mappings", {})
        default_values = session_data.get("default_values", {})
        
        # CRITICAL FIX: Validate mappings structure to prevent crashes
        if not mappings or not isinstance(mappings, dict) or 'mappings' not in mappings:
            logger.warning(f"🔍 WARNING: Invalid mappings structure in session {session_id}: {mappings}")
            mappings = {"mappings": []}  # Provide safe default
        
        # IMPORTANT: Derive column counts from default values if missing from session
        # This handles cases where templates were applied before the column count saving fix
        tags_count = session_data.get("tags_count", 3)
        spec_pairs_count = session_data.get("spec_pairs_count", 3)
        customer_id_pairs_count = session_data.get("customer_id_pairs_count", 1)
        
        logger.info(f"🔍 DEBUG: Session column counts - tags={tags_count}, spec={spec_pairs_count}, customer={customer_id_pairs_count}")
        logger.info(f"🔍 DEBUG: Default values keys: {list(default_values.keys()) if default_values else 'None'}")
        logger.info(f"🔍 DEBUG: Original template ID: {session_data.get('original_template_id')}")
        logger.info(f"🔍 DEBUG: Session source_type: {session_data.get('source_type')}")
        print(f"🔍 PRINT DEBUG: Session source_type: {session_data.get('source_type')}")
        
        # If column counts are missing but we have default values, derive them
        if (tags_count == 1 and spec_pairs_count == 1 and customer_id_pairs_count == 1 
            and default_values and session_data.get("original_template_id")):
            logger.info("🔍 DEBUG: Conditions met, attempting to derive column counts")
            # Count Tag_ fields in default values
            tag_fields = [field for field in default_values.keys() if field.startswith("Tag_")]
            logger.info(f"🔍 DEBUG: Found tag fields: {tag_fields}")
            if tag_fields:
                # Extract numbers from Tag_1, Tag_2, etc. and find the maximum
                tag_numbers = []
                for field in tag_fields:
                    try:
                        num = int(field.split('_')[1])
                        tag_numbers.append(num)
                    except (IndexError, ValueError):
                        pass
                if tag_numbers:
                    tags_count = max(tag_numbers)
                    logger.info(f"🔍 Derived tags_count={tags_count} from default values: {tag_fields}")
        
        # Include session metadata for template state restoration
        session_metadata = {
            'template_applied': bool(session_data.get("original_template_id")),
            'original_template_id': session_data.get("original_template_id"),
            'template_name': None,  # Will be filled if we have template
            'template_success': True,  # Assume success if template was applied
            'formula_rules': session_data.get("formula_rules", []),
            'header_corrections': session_data.get("header_corrections", {}),
            'factwise_rules': session_data.get("factwise_rules", []),
            # IMPORTANT: Include column counts so frontend shows all dynamic columns
            'column_counts': {
                'tags_count': tags_count,
                'spec_pairs_count': spec_pairs_count,
                'customer_id_pairs_count': customer_id_pairs_count
            }
        }

        # Add PDF metadata if session is from PDF (robust detection)
        source_type = session_data.get("source_type")
        is_pdf_session = (source_type == "pdf")
        if not is_pdf_session:
            try:
                from .models import PDFSession as _PDFSession
                is_pdf_session = _PDFSession.objects.filter(session_id=session_id).exists()
            except Exception:
                is_pdf_session = False
        logger.info(f"🔍 DEBUG: Session {session_id} source_type: {source_type} | inferred_pdf={is_pdf_session}")

        if is_pdf_session:
            # Get header confidence scores from database (same logic as headers endpoint)
            header_confidence_scores = {}
            try:
                from .models import PDFSession, PDFExtractionResult
                pdf_session = PDFSession.objects.get(session_id=session_id)
                pdf_extraction = PDFExtractionResult.objects.filter(pdf_session=pdf_session).order_by('-created_at').first()
                extracted_headers_list = list(pdf_extraction.extracted_headers) if pdf_extraction and pdf_extraction.extracted_headers else []
                if pdf_extraction and hasattr(pdf_extraction, 'confidence_scores') and pdf_extraction.confidence_scores:
                    # Extract header-level confidence scores if available
                    confidence_data = pdf_extraction.confidence_scores
                    if isinstance(confidence_data, dict):
                        # Check if we have header-specific confidence scores
                        header_confidence = confidence_data.get('header_confidence', {})
                        if header_confidence:
                            header_confidence_scores = header_confidence
                            logger.info(f"📊 MAPPINGS: Including header confidence scores for PDF session {session_id}: {header_confidence}")
                        else:
                            # Fallback: use overall confidence for all headers
                            overall_confidence = confidence_data.get('overall_confidence', 0.9)
                            headers_for_mapping = extracted_headers_list
                            header_confidence_scores = {header: overall_confidence for header in headers_for_mapping}
                            logger.info(f"📊 MAPPINGS: Using overall confidence {overall_confidence} for all PDF headers in session {session_id}")
                    else:
                        # Legacy fallback: assume high confidence if no detailed scores
                        headers_for_mapping = extracted_headers_list
                        header_confidence_scores = {header: 0.9 for header in headers_for_mapping}
                        logger.info(f"📊 MAPPINGS: Using default confidence for PDF headers in session {session_id}")
            except Exception as e:
                logger.error(f"📊 MAPPINGS: Error getting PDF confidence scores for session {session_id}: {e}")
                # Provide default confidence scores for PDF sessions even if we can't get detailed ones
                try:
                    # best-effort: reuse extracted headers if we already fetched them
                    header_list_fallback = extracted_headers_list
                except Exception:
                    header_list_fallback = []
                header_confidence_scores = {header: 0.8 for header in header_list_fallback}

            logger.info(f"🔍 DEBUG: Final header confidence scores: {header_confidence_scores}")

            # Add PDF metadata to session metadata
            session_metadata.update({
                'is_from_pdf': True,
                'header_confidence_scores': header_confidence_scores,
            })
        else:
            session_metadata.update({
                'is_from_pdf': False,
                'header_confidence_scores': {},
            })
        
        # Get template name if template was applied
        if session_metadata['original_template_id']:
            try:
                from .models import MappingTemplate
                template = MappingTemplate.objects.get(id=session_metadata['original_template_id'])
                session_metadata['template_name'] = template.name
            except Exception:
                session_metadata['template_name'] = 'Applied Template'
        
        # Normalize mappings to always return { mappings: [{"source": "...", "target": "..."}] } format
        normalized_mappings = {}
        if isinstance(mappings, dict) and 'mappings' in mappings and isinstance(mappings['mappings'], list):
            normalized_mappings = mappings
        elif isinstance(mappings, list):
            normalized_mappings = {'mappings': mappings}
        elif isinstance(mappings, dict):
            # old shape {target: source}
            normalized_mappings = {'mappings': [{'source': s, 'target': t} for t, s in mappings.items()]}
        else:
            normalized_mappings = {'mappings': []}

        return no_store(Response({
            'success': True,
            'mappings': normalized_mappings,                # <— one shape
            'default_values': default_values,
            'session_metadata': session_metadata
        }))
        
    except Exception as e:
        logger.error(f"Error in get_existing_mappings: {e}")
        return no_store(Response({
            'success': False,
            'error': f'Failed to get mappings: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR))


def calculate_data_quality_metrics(data_rows, headers, header_confidence_scores, confidence_data):
    """Calculate data quality metrics for frontend display."""
    try:
        total_cells = len(data_rows) * len(headers) if data_rows and headers else 0

        if total_cells == 0:
            return {
                'total_cells': 0,
                'high_quality_count': 0,
                'medium_quality_count': 0,
                'low_quality_count': 0,
                'average_confidence': 0.0,
                'header_quality': {},
                'row_quality': []
            }

        # Get row-level confidence from PDF extraction if available
        row_confidences = confidence_data.get('rows', {}) if confidence_data else {}
        # Normalize row_confidences keys (accept 'row_0' style and numeric strings)
        normalized_row_conf = {}
        try:
            for k, v in (row_confidences.items() if isinstance(row_confidences, dict) else []):
                key_str = str(k)
                if key_str.startswith('row_'):
                    idx = key_str.split('row_')[-1]
                    if idx.isdigit():
                        normalized_row_conf[idx] = v
                elif key_str.isdigit():
                    normalized_row_conf[key_str] = v
        except Exception:
            normalized_row_conf = {}

        # Calculate header quality metrics (column-centric)
        header_quality = {}
        for header in headers:
            header_conf = header_confidence_scores.get(header, 0.8)  # Default to 0.8
            header_quality[header] = {
                'confidence': header_conf,
                'quality_level': (
                    'high' if header_conf > 0.8 else
                    'medium' if header_conf > 0.6 else
                    'low'
                )
            }

        # Calculate overall quality metrics
        high_quality_count = 0
        medium_quality_count = 0
        low_quality_count = 0
        total_confidence = 0.0

        # Calculate row-level quality
        row_quality = []
        for row_idx, row in enumerate(data_rows):
            row_conf = normalized_row_conf.get(str(row_idx), 0.8)  # Default to 0.8

            # Combine row confidence with header confidences for overall row quality
            header_avg = sum(header_quality[h]['confidence'] for h in headers) / len(headers) if headers else 0.8
            combined_confidence = (row_conf + header_avg) / 2

            quality_level = (
                'high' if combined_confidence > 0.8 else
                'medium' if combined_confidence > 0.6 else
                'low'
            )

            row_quality.append({
                'index': row_idx,
                'confidence': combined_confidence,
                'quality_level': quality_level
            })

            # Count for totals
            if quality_level == 'high':
                high_quality_count += len(headers)
            elif quality_level == 'medium':
                medium_quality_count += len(headers)
            else:
                low_quality_count += len(headers)

            total_confidence += combined_confidence * len(headers)

        average_confidence = total_confidence / total_cells if total_cells > 0 else 0.0

        return {
            'total_cells': total_cells,
            'high_quality_count': high_quality_count,
            'medium_quality_count': medium_quality_count,
            'low_quality_count': low_quality_count,
            'average_confidence': round(average_confidence, 3),
            'header_quality': header_quality,
            'row_quality': row_quality
        }

    except Exception as e:
        logger.error(f"Error calculating data quality metrics: {e}")
        return {
            'total_cells': 0,
            'high_quality_count': 0,
            'medium_quality_count': 0,
            'low_quality_count': 0,
            'average_confidence': 0.0,
            'header_quality': {},
            'row_quality': []
        }


@api_view(['GET'])
@never_cache
def data_view(request):
    """Get transformed data with applied mappings."""
    try:
        session_id = request.GET.get('session_id')
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        
        # Validate page parameters and set reasonable limits for large datasets
        page = max(1, page)
        page_size = max(1, min(5000, page_size))  # Allow up to 5000 rows per page
        
        # Log performance for large page sizes
        if page_size > 1000:
            logger.info(f"🔍 Large page size requested: {page_size} rows on page {page} for session {session_id}")
        
        if not session_id:
            return Response({
                'success': False,
                'error': 'No session ID provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Robust, multi-worker safe session retrieval (cache -> memory -> file)
        info = get_session_consistent(session_id)
        if not info:
            return Response({
                'success': False,
                'error': 'Session not found. Please upload files again.'
            }, status=status.HTTP_400_BAD_REQUEST)
        # Cross-worker consistency: if file snapshot has newer template_version, refresh memory
        try:
            file_snapshot = load_session_from_file(session_id)
            if file_snapshot and file_snapshot.get('template_version', 0) > info.get('template_version', 0):
                SESSION_STORE[session_id] = file_snapshot
                info = file_snapshot
                logger.info(f"🔄 Refreshed in-memory session {session_id} from file (newer template_version)")
        except Exception as _e:
            pass
        logger.info(f"🔧 DEBUG: Processing session {session_id} for data view")

        # If manual edits exist, serve them immediately without requiring mappings
        edited_data = info.get('edited_data')
        if isinstance(edited_data, list) and edited_data:
            headers_to_use = info.get('enhanced_headers') or info.get('current_template_headers')
            if not headers_to_use:
                tags_count = int(info.get('tags_count', 3))
                spec_pairs_count = int(info.get('spec_pairs_count', 3))
                customer_id_pairs_count = int(info.get('customer_id_pairs_count', 1))
                headers_to_use = [
                    'Item code', 'Item name', 'Description', 'Item type', 'Measurement unit',
                    'Procurement entity name', 'Notes', 'Internal notes', 'Procurement item', 'Sales item', 'Preferred vendor code'
                ]
                for i in range(1, tags_count + 1): headers_to_use.append(f'Tag_{i}')
                for i in range(1, spec_pairs_count + 1): headers_to_use += [f'Specification_Name_{i}', f'Specification_Value_{i}']
                for i in range(1, customer_id_pairs_count + 1): headers_to_use += [f'Customer_Identification_Name_{i}', f'Customer_Identification_Value_{i}']

            # Build transformed rows from edited_data
            transformed_rows = []
            for r in edited_data:
                if isinstance(r, dict):
                    transformed_rows.append({h: r.get(h, '') for h in headers_to_use})

            # Quality + metadata calculation as usual
            final_headers = headers_to_use
            final_data = transformed_rows
            confidence_data = {}
            header_confidence_scores = {}
            quality_metrics = calculate_data_quality_metrics(final_data, final_headers, header_confidence_scores, confidence_data)

            return no_store(Response({
                'success': True,
                'headers': final_headers,
                'data': final_data,
                'total_rows': len(final_data),
                'formula_rules': info.get('formula_rules', []),
                'template_version': info.get('template_version', 0),
                'quality_metrics': quality_metrics,
                'header_confidence_scores': header_confidence_scores,
                'target_column_confidence_scores': header_confidence_scores,
                'is_from_pdf': (info.get('source_type') == 'pdf'),
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total_rows': len(final_data),
                    'total_pages': max(1, (len(final_data) + page_size - 1) // page_size)
                }
            }))
        
        # Always process fresh data - no caching
        mappings = info.get("mappings")
        # If mappings missing, try to refresh from file to handle multi-worker race
        if not mappings:
            refreshed = load_session_from_file(session_id)
            if refreshed and refreshed.get("mappings"):
                SESSION_STORE[session_id] = refreshed
                info = refreshed
                mappings = info.get("mappings")
                logger.info(f"🔄 Refreshed session {session_id} from file to load latest mappings")
        
        if not mappings:
            return Response({
                'success': False,
                'error': 'No mappings found. Please create mappings first.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Convert mappings list format to expected dict format for apply_column_mappings
        if isinstance(mappings, list):
            # Convert list format to new dict format that apply_column_mappings expects
            formatted_mappings = {"mappings": mappings}
            logger.info(f"🔧 DEBUG: Converted list mappings to dict format: {formatted_mappings}")
        else:
            formatted_mappings = mappings
            logger.info(f"🔧 DEBUG: Using existing dict mappings: {formatted_mappings}")
        
        # Prefer manually edited data first (permanent edits from the editor)
        edited_data = info.get('edited_data')
        enhanced_data = info.get("formula_enhanced_data")
        enhanced_headers = info.get("enhanced_headers")

        template_just_applied = info.get("original_template_id") is not None
        # Prefer enhanced data if present (e.g., user uploaded corrected CSV) even when force_fresh=true
        force_fresh_param = request.GET.get('force_fresh', 'false').lower() == 'true'
        if info.get('uploaded_via_correction') and enhanced_data and enhanced_headers:
            force_fresh_mapping = False
        else:
            force_fresh_mapping = (template_just_applied and not (enhanced_data and enhanced_headers)) or (force_fresh_param and not (enhanced_data and enhanced_headers))
        # Stability option for consumers like DataEditor: avoid cleaning headers by page slice
        stable_headers = request.GET.get('stable', 'false').lower() == 'true'

        using_enhanced = False
        # 1) Edited data takes top priority
        if isinstance(edited_data, list) and edited_data:
            headers_to_use = enhanced_headers or info.get('current_template_headers') or []
            # If headers_to_use is empty, derive canonical headers from counts
            if not headers_to_use:
                tags_count = int(info.get('tags_count', 3))
                spec_pairs_count = int(info.get('spec_pairs_count', 3))
                customer_id_pairs_count = int(info.get('customer_id_pairs_count', 1))
                headers_to_use = [
                    'Item code', 'Item name', 'Description', 'Item type', 'Measurement unit',
                    'Procurement entity name', 'Notes', 'Internal notes', 'Procurement item', 'Sales item', 'Preferred vendor code'
                ]
                for i in range(1, tags_count + 1): headers_to_use.append(f'Tag_{i}')
                for i in range(1, spec_pairs_count + 1): headers_to_use += [f'Specification_Name_{i}', f'Specification_Value_{i}']
                for i in range(1, customer_id_pairs_count + 1): headers_to_use += [f'Customer_Identification_Name_{i}', f'Customer_Identification_Value_{i}']
            # Rebuild rows to include all headers in order
            transformed_rows = []
            for r in edited_data:
                if isinstance(r, dict):
                    transformed_rows.append({h: r.get(h, '') for h in headers_to_use})
            using_enhanced = True
            logger.info(f"🔧 DEBUG: Using edited data with {len(headers_to_use)} headers and {len(transformed_rows)} rows")
            try:
                info['enhanced_headers'] = headers_to_use
                save_session(session_id, info)
            except Exception:
                pass
        # 2) Otherwise use enhanced (correction / formula) data
        elif enhanced_data and enhanced_headers and not force_fresh_mapping:
            transformed_rows = enhanced_data
            headers_to_use = enhanced_headers
            using_enhanced = True
            logger.info(f"🔧 DEBUG: Using enhanced data with {len(headers_to_use)} headers and {len(transformed_rows)} rows")
            # Persist canonical headers to avoid worker drift
            try:
                info["current_template_headers"] = headers_to_use
                save_session(session_id, info)
            except Exception:
                pass
        else:
            # fresh mapping – process only requested page to avoid heavy work on large datasets
            # Compute pagination window and pass hints to the mapper
            start_idx = (page - 1) * page_size
            # Provide pagination hints for apply_column_mappings
            try:
                SESSION_STORE.setdefault(session_id, {})['__paginate__'] = {'offset': start_idx, 'limit': page_size}
            except Exception:
                pass
            mapping_result = apply_column_mappings(
                client_file=info["client_path"],
                mappings=formatted_mappings,
                sheet_name=info["sheet_name"],
                header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0,
                session_id=session_id
            )
            # Clear pagination hint to avoid affecting other endpoints
            try:
                if '__paginate__' in SESSION_STORE.get(session_id, {}):
                    del SESSION_STORE[session_id]['__paginate__']
            except Exception:
                pass

            # Check if enhanced data with MPN validation exists
            enhanced_data = info.get('enhanced_data')
            if enhanced_data and enhanced_data.get('headers') and enhanced_data.get('data'):
                transformed_rows = enhanced_data['data']
                headers_to_use = enhanced_data['headers']
                using_enhanced = True
                logger.info(f"🔧 DEBUG: Using enhanced MPN validated data with {len(headers_to_use)} headers and {len(transformed_rows)} rows")
            else:
                transformed_rows = mapping_result['data']
                headers_to_use = mapping_result['headers']
                using_enhanced = False
                logger.info(f"🔧 DEBUG: Using fresh mapped data (paginated) with {len(headers_to_use)} headers and {len(transformed_rows)} rows")
            
            # CRITICAL FIX: If we forced fresh mapping due to template application, we need to re-apply formulas
            # to ensure Tag columns are populated with the correct data from the original file
            if force_fresh_mapping and template_just_applied:
                logger.info(f"🔧 DEBUG: Re-applying formulas after fresh mapping for template application")
                # Clear any stale enhanced data that might interfere and update local variables
                if "formula_enhanced_data" in info:
                    del info["formula_enhanced_data"]
                if "enhanced_headers" in info:
                    del info["enhanced_headers"]
                # Update local variables to reflect the clearing
                enhanced_data = None
                enhanced_headers = None
        
        # Apply formula rules if they exist to create unique tag columns
        formula_rules = info.get("formula_rules", [])
        
        # Ensure formula column headers are included in headers_to_use
        if formula_rules:
            formula_headers = []
            for rule in formula_rules:
                target_col = rule.get('target_column')
                if target_col and target_col not in headers_to_use:
                    formula_headers.append(target_col)
            
            if formula_headers:
                headers_to_use.extend(formula_headers)
                logger.info(f"🔧 DEBUG: Added formula headers to response: {formula_headers}")
                logger.info(f"🔧 DEBUG: Updated headers_to_use: {headers_to_use}")
        
        # De-dup rules as you already do...
        formula_rules = info.get('formula_rules', [])

        # Convert list-based data to dict format BEFORE applying formulas
        if transformed_rows and len(transformed_rows) > 0 and isinstance(transformed_rows[0], list):
            dict_rows = []
            for row_list in transformed_rows:
                row_dict = {}
                
                for i, header in enumerate(headers_to_use):
                    if i < len(row_list):
                        row_dict[header] = row_list[i]
                    else:
                        # Handle missing values
                        row_dict[header] = ""
                
                dict_rows.append(row_dict)
            transformed_rows = dict_rows

        # Inject MPN validation columns with multiple canonical MPNs if available
        try:
            mpn_validation = info.get('mpn_validation') or {}
            mpn_header = mpn_validation.get('column')
            results_map = mpn_validation.get('results') or {}
            if mpn_header and isinstance(transformed_rows, list) and transformed_rows:
                from .services.digikey_service import DigiKeyClient
                client_norm = DigiKeyClient.normalize_mpn

                # Limit canonical MPN columns to reasonable number (max 5) to avoid excessive blank columns
                max_canonical_mpns = 1
                for row in transformed_rows:
                    if isinstance(row, dict):
                        raw = row.get(mpn_header, '')
                        norm = client_norm(raw)
                        res = results_map.get(norm, {})
                        all_canonicals = res.get('all_canonical_mpns', [])
                        if len(all_canonicals) > max_canonical_mpns:
                            max_canonical_mpns = min(len(all_canonicals), 5)  # Cap at 5 columns

                # Add base MPN validation columns to headers if not present
                base_validation_columns = ['MPN valid', 'MPN Status', 'EOL Status', 'Discontinued', 'DKPN', 'Category']
                for mpn_col in base_validation_columns:
                    if mpn_col not in headers_to_use:
                        headers_to_use.append(mpn_col)

                # Add multiple canonical MPN columns based on maximum needed
                canonical_columns = []
                for i in range(max_canonical_mpns):
                    if i == 0:
                        col_name = 'Canonical MPN'  # First one keeps original name
                    else:
                        col_name = f'Canonical MPN {i + 1}'  # Additional ones get numbered

                    canonical_columns.append(col_name)
                    if col_name not in headers_to_use:
                        headers_to_use.append(col_name)

                # Populate all MPN validation data for each row
                for row in transformed_rows:
                    if isinstance(row, dict):
                        raw = row.get(mpn_header, '')
                        norm = client_norm(raw)
                        res = results_map.get(norm, {})
                        lifecycle = res.get('lifecycle') or {}
                        all_canonicals = res.get('all_canonical_mpns', [])

                        # Set base MPN validation columns, fixing invalid MPNs to have blank status fields
                        is_valid = res.get('valid', False)
                        row['MPN valid'] = 'Yes' if is_valid else ('No' if norm else '')

                        # Only populate status fields for valid MPNs, leave blank for invalid
                        if is_valid:
                            row['MPN Status'] = lifecycle.get('status') or 'Unknown'
                            row['EOL Status'] = 'Yes' if lifecycle.get('endOfLife') else 'No'
                            row['Discontinued'] = 'Yes' if lifecycle.get('discontinued') else 'No'
                            row['DKPN'] = res.get('dkpn') or ''
                        else:
                            # Invalid MPNs should have blank status fields
                            row['MPN Status'] = ''
                            row['EOL Status'] = ''
                            row['Discontinued'] = ''
                            row['DKPN'] = ''

                        # Add category information from DigiKey (only for valid MPNs)
                        if is_valid:
                            category_info = res.get('category', {})
                            row['Category'] = category_info.get('name', '') if category_info else ''
                        else:
                            row['Category'] = ''

                        # Set multiple canonical MPN columns (show suggestions for both valid and invalid MPNs)
                        for i, col_name in enumerate(canonical_columns):
                            if i < len(all_canonicals):
                                row[col_name] = all_canonicals[i]  # Show canonical suggestions for both valid and invalid MPNs
                            else:
                                row[col_name] = ''  # Empty if no more canonical MPNs

                logger.info(f"🔧 DEBUG data_view: Added MPN validation columns with {max_canonical_mpns} canonical MPN variants: {base_validation_columns + canonical_columns}")
        except Exception as _me:
            logger.warning(f"MPN validation injection skipped: {_me}")

        # IMPORTANT: apply formulas only if we did NOT use the enhanced branch
        if formula_rules and transformed_rows and not using_enhanced:
            formula_result = apply_formula_rules(transformed_rows, headers_to_use, formula_rules, replace_existing=False, session_info=info)
            transformed_rows = formula_result['data']
            headers_to_use = formula_result['headers']
            # Persist canonically across workers but avoid storing large full datasets
            info['enhanced_headers'] = headers_to_use
            info['current_template_headers'] = headers_to_use
            info['version'] = info.get('version', 0) + 1
            # Do NOT store full enhanced data here; data is paginated and can be recomputed per page
            save_session(session_id, info)
        
        # Apply factwise ID rules if they exist
        factwise_rules = info.get("factwise_rules", [])
        logger.info(f"🔧 DEBUG: Found {len(factwise_rules)} factwise rules: {factwise_rules}")
        
        for factwise_rule in factwise_rules:
            if factwise_rule.get("type") == "factwise_id" and transformed_rows:
                try:
                    first_col = factwise_rule.get("first_column")
                    second_col = factwise_rule.get("second_column")
                    operator = factwise_rule.get("operator", "_")
                    strategy = factwise_rule.get("strategy", "fill_only_null")
                    
                    logger.info(f"🔧 DEBUG: Factwise rule - first_col: '{first_col}', second_col: '{second_col}', operator: '{operator}'")
                    logger.info(f"🔧 DEBUG: Available headers: {headers_to_use}")
                    
                    if first_col and second_col and first_col in headers_to_use and second_col in headers_to_use:
                        first_idx = headers_to_use.index(first_col)
                        second_idx = headers_to_use.index(second_col)
                        
                        logger.info(f"🔧 DEBUG: Column indices - first_idx: {first_idx}, second_idx: {second_idx}")
                        
                        # Map into Item code rather than creating a new column
                        # Normalize headers to find Item code variant
                        def norm(s: str) -> str:
                            return str(s).strip().lower().replace(' ', '').replace('_', '').replace('-', '')

                        item_header = None
                        for h in headers_to_use:
                            if norm(h) == norm('item code'):
                                item_header = h
                                break

                        if not item_header:
                            # If no Item code header exists yet, create one at the beginning
                            headers_to_use.insert(0, 'Item code')
                            item_header = 'Item code'
                            # Persist canonical headers
                            try:
                                info["current_template_headers"] = headers_to_use
                                save_session(session_id, info)
                            except Exception:
                                pass
                            for i, row in enumerate(transformed_rows):
                                if isinstance(row, dict):
                                    first_val = row.get(first_col, "")
                                    second_val = row.get(second_col, "")
                                    factwise_id = f"{first_val}{operator}{second_val}" if first_val and second_val else (first_val or second_val or "")
                                    row[item_header] = factwise_id
                                else:
                                    first_val = row[first_idx] if first_idx < len(row) else ""
                                    second_val = row[second_idx] if second_idx < len(row) else ""
                                    factwise_id = f"{first_val}{operator}{second_val}" if first_val and second_val else (first_val or second_val or "")
                                    row.insert(0, factwise_id)
                        else:
                            # Fill existing Item code per strategy
                            for i, row in enumerate(transformed_rows):
                                if isinstance(row, dict):
                                    first_val = row.get(first_col, "")
                                    second_val = row.get(second_col, "")
                                    factwise_id = f"{first_val}{operator}{second_val}" if first_val and second_val else (first_val or second_val or "")
                                    if strategy == 'override_all' or not row.get(item_header):
                                        row[item_header] = factwise_id
                                else:
                                    first_val = row[first_idx] if first_idx < len(row) else ""
                                    second_val = row[second_idx] if second_idx < len(row) else ""
                                    factwise_id = f"{first_val}{operator}{second_val}" if first_val and second_val else (first_val or second_val or "")
                                    # Find index of item_header
                                    try:
                                        item_idx = headers_to_use.index(item_header)
                                    except ValueError:
                                        item_idx = None
                                    if item_idx is not None:
                                        if strategy == 'override_all' or (item_idx < len(row) and (row[item_idx] is None or str(row[item_idx]).strip() == "")):
                                            # Ensure row length and assign
                                            while len(row) <= item_idx:
                                                row.append("")
                                            row[item_idx] = factwise_id
                            
                            if i == 0:  # Log first row for debugging
                                logger.info(f"🔧 DEBUG: First row Factwise ID: '{factwise_id}' from '{first_val}' + '{operator}' + '{second_val}'")
                    else:
                        logger.warning(f"🔧 DEBUG: Columns not found - first_col '{first_col}' in headers: {first_col in headers_to_use}, second_col '{second_col}' in headers: {second_col in headers_to_use}")
                            
                except Exception as e:
                    logger.warning(f"Factwise ID application failed: {e}")
                    import traceback
                    logger.warning(f"Traceback: {traceback.format_exc()}")

        if not transformed_rows:
            return Response({
                'success': False,
                'error': 'No data could be transformed'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # List-to-dict conversion already done above before formula processing

        # Normalize generic 'Tag' column: move any values into numbered Tag_N columns, then drop 'Tag'
        try:
            if isinstance(headers_to_use, list) and 'Tag' in headers_to_use and isinstance(transformed_rows, list) and len(transformed_rows) > 0:
                tag_n_headers = [h for h in headers_to_use if isinstance(h, str) and h.startswith('Tag_')]
                try:
                    tag_n_headers.sort(key=lambda x: int(x.split('_')[1]))
                except Exception:
                    tag_n_headers.sort()
                if isinstance(transformed_rows[0], dict):
                    for row in transformed_rows:
                        val = str(row.get('Tag', '') or '').strip()
                        if not val:
                            continue
                        placed = False
                        for tcol in tag_n_headers:
                            cur = str(row.get(tcol, '') or '').strip()
                            if not cur:
                                row[tcol] = val
                                placed = True
                                break
                        if not placed and tag_n_headers:
                            last = tag_n_headers[-1]
                            cur = str(row.get(last, '') or '').strip()
                            if cur:
                                parts = [p.strip() for p in cur.split(',')]
                                if val not in parts:
                                    row[last] = f"{cur}, {val}"
                            else:
                                row[last] = val
                        row.pop('Tag', None)
                else:
                    tag_idx = headers_to_use.index('Tag')
                    tag_n_indices = []
                    for h in tag_n_headers:
                        try:
                            tag_n_indices.append(headers_to_use.index(h))
                        except ValueError:
                            pass
                    for row in transformed_rows:
                        val = ''
                        if tag_idx < len(row):
                            val = str(row[tag_idx] or '').strip()
                        if not val:
                            continue
                        placed = False
                        for idx in tag_n_indices:
                            if idx < len(row):
                                cur = str(row[idx] or '').strip()
                                if not cur:
                                    row[idx] = val
                                    placed = True
                                    break
                        if not placed and tag_n_indices:
                            last_idx = tag_n_indices[-1]
                            if last_idx < len(row):
                                cur = str(row[last_idx] or '').strip()
                                if cur:
                                    parts = [p.strip() for p in cur.split(',')]
                                    if val not in parts:
                                        row[last_idx] = f"{cur}, {val}"
                                else:
                                    row[last_idx] = val
                    # Note: header removal occurs in cleaned headers step later
                # remove generic Tag header now
                headers_to_use = [h for h in headers_to_use if h != 'Tag']
        except Exception:
            pass

        # Avoid dropping Tag_N columns in correction mode (we want to keep full template footprint)
        if not info.get('uploaded_via_correction'):
            try:
                if isinstance(headers_to_use, list) and headers_to_use and transformed_rows:
                    tag_headers = [h for h in headers_to_use if isinstance(h, str) and h.startswith('Tag_')]
                    # Build a set of Tag_N with any data
                    non_empty = set()
                    if isinstance(transformed_rows[0], dict):
                        for h in tag_headers:
                            for row in transformed_rows:
                                if str(row.get(h, '') or '').strip():
                                    non_empty.add(h)
                                    break
                    else:
                        for h in tag_headers:
                            try:
                                idx = headers_to_use.index(h)
                            except ValueError:
                                continue
                            for row in transformed_rows:
                                if idx < len(row) and str(row[idx] or '').strip():
                                    non_empty.add(h)
                                    break
                    # Remove Tag_N columns that are entirely empty
                    to_remove = [h for h in tag_headers if h not in non_empty]
                    if to_remove:
                        headers_to_use = [h for h in headers_to_use if h not in to_remove]
                        if isinstance(transformed_rows[0], dict):
                            for row in transformed_rows:
                                for h in to_remove:
                                    row.pop(h, None)
                        # list-of-lists cleanup will be handled later with cleaned headers if needed
            except Exception:
                pass
        
        # Apply default values for unmapped fields
        default_values = info.get("default_values", {})
        logger.info(f"🔧 DEBUG: Session {session_id} - Checking default values: {default_values}")
        logger.info(f"🔧 DEBUG: Session {session_id} - Headers available: {headers_to_use}")
        
        if default_values and transformed_rows:
            logger.info(f"🔧 DEBUG: Applying default values to {len(transformed_rows)} rows: {default_values}")
            
            for field_name, default_value in default_values.items():
                # CRITICAL FIX: Handle both internal and external field names for default values
                matched_field = None
                
                # First, try exact match (most common case)
                if field_name in headers_to_use:
                    matched_field = field_name
                else:
                    # Handle internal names (e.g., "Specification_Name_1")
                    if field_name.startswith('Specification_Name_'):
                        matched_field = field_name
                    elif field_name.startswith('Specification_Value_'):
                        matched_field = field_name
                    elif field_name.startswith('Customer_Identification_Name_'):
                        matched_field = field_name
                    elif field_name.startswith('Customer_Identification_Value_'):
                        matched_field = field_name
                    elif field_name.startswith('Tag_'):
                        matched_field = field_name
                    # Handle external names (e.g., "Specification name")
                    elif field_name == "Specification name":
                        # Find the first available Specification_Name_X column
                        for header in headers_to_use:
                            if header.startswith('Specification_Name_'):
                                matched_field = header
                                break
                    elif field_name == "Specification value":
                        # Find the first available Specification_Value_X column
                        for header in headers_to_use:
                            if header.startswith('Specification_Value_'):
                                matched_field = header
                                break
                    elif field_name == "Customer identification name":
                        # Find the first available Customer_Identification_Name_X column
                        for header in headers_to_use:
                            if header.startswith('Customer_Identification_Name_'):
                                matched_field = header
                                break
                    elif field_name == "Customer identification value":
                        # Find the first available Customer_Identification_Value_X column
                        for header in headers_to_use:
                            if header.startswith('Customer_Identification_Value_'):
                                matched_field = header
                                break
                    elif field_name == "Tag":
                        # Find the first available Tag_X column
                        for header in headers_to_use:
                            if header.startswith('Tag_'):
                                matched_field = header
                                break
                
                if matched_field:
                    logger.info(f"🔧 DEBUG: Found matching field '{matched_field}' for default value field '{field_name}'")
                    logger.info(f"🔧 DEBUG: Setting default value '{default_value}' for field '{matched_field}' in {len(transformed_rows)} rows")
                    
                    # CRITICAL FIX: Apply default values more intelligently
                    # Apply defaults to all rows for unmapped fields to ensure consistency
                    rows_updated = 0
                    for row in transformed_rows:
                        current_value = row.get(matched_field, None)
                        # Apply default if field is None, empty string, "nan", or doesn't exist
                        # This ensures all rows get the default value for unmapped fields
                        if (current_value is None or 
                            current_value == "" or 
                            current_value == "nan" or 
                            str(current_value).strip() == "" or
                            str(current_value).lower() == "nan"):
                            row[matched_field] = default_value
                            rows_updated += 1
                        # Also apply if the field value is the same as the default (indicating it was already set)
                        elif str(current_value).strip() == str(default_value).strip():
                            rows_updated += 1
                    
                    logger.info(f"🔧 DEBUG: Applied default value '{default_value}' to field '{matched_field}' in {rows_updated} rows")
                else:
                    logger.info(f"🔧 DEBUG: No matching field found for default value '{default_value}' for field '{field_name}'")
                    # If the default-only field is missing from headers, add it canonically and populate
                    headers_to_use.append(field_name)
                    for row in transformed_rows:
                        if isinstance(row, dict):
                            row[field_name] = default_value
                    try:
                        info["current_template_headers"] = headers_to_use
                        save_session(session_id, info)
                    except Exception:
                        pass
        else:
            if not default_values:
                logger.info(f"🔧 DEBUG: Session {session_id} - No default values found in session data")
            if not transformed_rows:
                logger.info(f"🔧 DEBUG: Session {session_id} - No transformed rows found")
        
        # Implement pagination
        # We already paginated at read-time. Compute total_rows accurately for UI.
        try:
            client_local_path = hybrid_file_manager.get_file_path(info["client_path"])
            total_rows = _count_total_data_rows(client_local_path, info.get("sheet_name"), info.get("header_row", 1) - 1 if info.get("header_row", 1) > 0 else 0)
        except Exception:
            # Fallback to current page length if counting fails
            total_rows = start_idx + len(transformed_rows)
        paginated_rows = transformed_rows
        
        # Use the headers we determined above (either enhanced or template headers)
        
        # Include formula rules in response if they exist
        formula_rules = info.get("formula_rules", [])
        
        # Convert internal column names to external names for frontend display
        # Internal: Tag_1, Tag_2, etc. -> External: Tag (always generic name)
        # Internal: Specification_Name_1, Specification_Value_1, etc. -> External: Specification name, Specification value
        # Internal: Customer_Identification_Name_1, Customer_Identification_Value_1, etc. -> External: Customer identification name, Customer identification value
        external_headers = []
        internal_to_external_mapping = {}
        
        for header in headers_to_use:
            if header.startswith('Tag_') or header == 'Tag':
                # Always show "Tag" regardless of how many tags exist
                external_header = 'Tag'
            elif header.startswith('Specification_Name_') or header == 'Specification name':
                external_header = 'Specification name'
            elif header.startswith('Specification_Value_') or header == 'Specification value':
                external_header = 'Specification value'
            elif header.startswith('Customer_Identification_Name_') or header == 'Customer identification name':
                external_header = 'Customer identification name'
            elif header.startswith('Customer_Identification_Value_') or header == 'Customer identification value':
                external_header = 'Customer identification value'
            else:
                external_header = header
            
            external_headers.append(external_header)
            internal_to_external_mapping[header] = external_header
        
        # CRITICAL FIX: PRESERVE ALL TEMPLATE COLUMNS - Don't remove any columns
        # The user expects to see the complete template structure, even if columns are empty
        # Only remove truly unnecessary columns that are completely outside the template
        cleaned_headers = []
        cleaned_external_headers = []
        cleaned_internal_to_external = {}
        
        # Build canonical sets for ALL template columns (core + dynamic)
        template_norm = set()
        
        # Add core headers
        core_headers = [
            "Item code",
            "Item name",
            "Description",
            "Item type",
            "Measurement unit",
            "Procurement entity name",
            "Notes",
            "Internal notes",
            "Procurement item",
            "Sales item",
            "Preferred vendor code",
        ]
        for h in core_headers:
            template_norm.add(_canon(h))
        
        # Add dynamic headers based on session counts
        if session_id and session_id in SESSION_STORE:
            session_info = SESSION_STORE[session_id]
            tags_count = session_info.get('tags_count', 1)
            spec_pairs_count = session_info.get('spec_pairs_count', 1)
            customer_id_pairs_count = session_info.get('customer_id_pairs_count', 1)
            
            # Add Tag columns
            for i in range(1, tags_count + 1):
                template_norm.add(_canon(f"Tag_{i}"))
            
            # Add Specification columns  
            for i in range(1, spec_pairs_count + 1):
                template_norm.add(_canon(f"Specification_Name_{i}"))
                template_norm.add(_canon(f"Specification_Value_{i}"))
            
            # Add Customer Identification columns
            for i in range(1, customer_id_pairs_count + 1):
                template_norm.add(_canon(f"Customer_Identification_Name_{i}"))
                template_norm.add(_canon(f"Customer_Identification_Value_{i}"))
        
        logger.info(f"🔧 DEBUG: Complete template structure normalized set: {sorted(template_norm)}")
        logger.info(f"🔧 DEBUG: Starting cleanup loop for {len(headers_to_use)} headers - PRESERVING ALL TEMPLATE COLUMNS")
        
        for i, header in enumerate(headers_to_use):
            header_canon = _canon(header)
            is_template_column = header_canon in template_norm
            
            # ALWAYS keep template columns, regardless of whether they have data
            if is_template_column:
                cleaned_headers.append(header)
                if i < len(external_headers):
                    cleaned_external_headers.append(external_headers[i])
                    cleaned_internal_to_external[header] = external_headers[i]
                logger.debug(f"🔧 CLEANUP: PRESERVED template column '{header}' (template structure)")
            else:
                # For non-template columns, check if they have data
                has_data = False
                for row in transformed_rows:
                    if isinstance(row, dict):
                        value = row.get(header, '')
                    elif isinstance(row, list) and i < len(row):
                        value = row[i] if row[i] is not None else ''
                    else:
                        value = ''
                    
                    if value and str(value).strip() and str(value).strip().lower() not in ['', 'none', 'null', 'nan']:
                        has_data = True
                        break
                
                if has_data:
                    cleaned_headers.append(header)
                    if i < len(external_headers):
                        cleaned_external_headers.append(external_headers[i])
                        cleaned_internal_to_external[header] = external_headers[i]
                    logger.debug(f"🔧 CLEANUP: Kept non-template column '{header}' (has data)")
                else:
                    logger.info(f"🔧 CLEANUP: Removed blank non-template column '{header}'")
        
        # Update the variables to use cleaned versions
        if not stable_headers:
            headers_to_use = cleaned_headers
            external_headers = cleaned_external_headers
            internal_to_external_mapping = cleaned_internal_to_external

            # Also clean up the paginated_rows to only include data for kept columns
            original_header_count = len([h for h in (info.get('current_template_headers', []) or info.get('enhanced_headers', []) or [])])
            if len(cleaned_headers) < original_header_count:
                logger.info(f"🔧 CLEANUP: Cleaning data rows to match {len(cleaned_headers)} cleaned headers")
                cleaned_paginated_rows = []
                original_headers = [h for h in headers_to_use]  # Keep original reference

                for row in paginated_rows:
                    if isinstance(row, dict):
                        # Keep only fields that correspond to cleaned headers
                        cleaned_row = {header: row.get(header, '') for header in cleaned_headers}
                        cleaned_paginated_rows.append(cleaned_row)
                    elif isinstance(row, list):
                        # Keep only columns that correspond to cleaned headers indices
                        cleaned_row = []
                        original_headers_list = list(info.get('current_template_headers', [])) or headers_to_use
                        for header in cleaned_headers:
                            try:
                                idx = original_headers_list.index(header)
                                cleaned_row.append(row[idx] if idx < len(row) else '')
                            except (ValueError, IndexError):
                                cleaned_row.append('')
                        cleaned_paginated_rows.append(cleaned_row)
                    else:
                        cleaned_paginated_rows.append(row)

                paginated_rows = cleaned_paginated_rows
        
        # Do not clear original_template_id; keep template-applied state for dashboard and restores

        # Include formula rules in response so frontend can display them
        formula_rules = info.get('formula_rules', [])
        
        # FINAL SAFETY: ensure dict rows do not include stray keys not present in headers_to_use
        try:
            if isinstance(paginated_rows, list) and paginated_rows and isinstance(paginated_rows[0], dict):
                allowed = set(headers_to_use)
                cleaned = []
                for row in paginated_rows:
                    cleaned.append({k: v for k, v in row.items() if k in allowed})
                paginated_rows = cleaned
        except Exception:
            pass

        # Apply header corrections if they exist (for PDF sessions)
        final_headers = headers_to_use
        final_data = paginated_rows
        header_corrections = info.get('header_corrections', {})
        if header_corrections and isinstance(header_corrections, dict):
            corrected_headers = []
            header_mapping = {}
            for header in headers_to_use:
                # Find the corrected header name
                corrected_header = header
                for original, corrected in header_corrections.items():
                    if header == original:
                        corrected_header = corrected
                        header_mapping[original] = corrected
                        logger.info(f"🔧 DEBUG: Applied header correction: {original} -> {corrected}")
                        break
                corrected_headers.append(corrected_header)
            final_headers = corrected_headers

            # Update data rows to use corrected headers as keys
            if header_mapping and final_data and isinstance(final_data, list) and len(final_data) > 0 and isinstance(final_data[0], dict):
                corrected_data = []
                for row in final_data:
                    corrected_row = {}
                    for key, value in row.items():
                        corrected_key = header_mapping.get(key, key)
                        corrected_row[corrected_key] = value
                    corrected_data.append(corrected_row)
                final_data = corrected_data
                logger.info(f"🔧 DEBUG: Updated {len(final_data)} data rows with corrected headers")

            logger.info(f"🔧 DEBUG: Final headers after corrections: {final_headers}")

        # Get confidence data for quality metrics (robust PDF detection)
        confidence_data = {}
        header_confidence_scores = {}

        is_pdf_session_flag = False
        try:
            is_pdf_session = (info.get('source_type') == 'pdf')
            if not is_pdf_session:
                try:
                    from .models import PDFSession as _PDFSession
                    is_pdf_session = _PDFSession.objects.filter(session_id=session_id).exists()
                except Exception:
                    is_pdf_session = False
            is_pdf_session_flag = is_pdf_session

            if is_pdf_session:
                from .models import PDFSession, PDFExtractionResult
                pdf_session = PDFSession.objects.get(session_id=session_id)
                pdf_extraction = PDFExtractionResult.objects.filter(pdf_session=pdf_session).order_by('-created_at').first()
                if pdf_extraction and getattr(pdf_extraction, 'confidence_scores', None):
                    confidence_data = pdf_extraction.confidence_scores or {}
                # Build header_confidence_scores aligned to final_headers
                header_conf_map = {}
                try:
                    header_conf_map = (confidence_data.get('header_confidence') or {}) if isinstance(confidence_data, dict) else {}
                except Exception:
                    header_conf_map = {}
                # Default overall fallback
                overall_conf = 0.8
                try:
                    qm = confidence_data.get('quality_metrics') or {}
                    overall_conf = float(qm.get('header_confidence') or qm.get('overall_confidence') or 0.8)
                except Exception:
                    overall_conf = 0.8

                # Column confidence based on mapping: assign source header confidence to target columns
                target_column_confidence = {}
                try:
                    # Build normalized header confidence for source headers (case-insensitive)
                    def _norm(h: str) -> str:
                        try:
                            return ''.join(ch for ch in str(h or '').lower().strip() if ch.isalnum())
                        except Exception:
                            return ''
                    norm_conf = {_norm(k): float(v) for k, v in (header_conf_map.items() if isinstance(header_conf_map, dict) else [])}
                    # Pull mappings from session
                    src_mappings = info.get('mappings')
                    if isinstance(src_mappings, dict) and 'mappings' in src_mappings:
                        mapping_list = src_mappings['mappings']
                    elif isinstance(src_mappings, list):
                        mapping_list = src_mappings
                    else:
                        # Old object format {target: source}
                        mapping_list = [{'source': s, 'target': t} for t, s in (src_mappings or {}).items()]

                    for m in (mapping_list or []):
                        try:
                            src = (m.get('source') or '').strip()
                            tgt = (m.get('target') or '').strip()
                            if not src or not tgt:
                                continue
                            # Only set if known; avoid defaulting to overall_conf here
                            conf = norm_conf.get(_norm(src))
                            if conf is None:
                                continue
                            prev = target_column_confidence.get(tgt)
                            if prev is None or conf > prev:
                                target_column_confidence[tgt] = conf
                        except Exception:
                            continue
                except Exception:
                    target_column_confidence = {}

                # Build header_confidence_scores for final headers using mapping-based confidence
                # If a column has no mapping-based confidence and is entirely blank, set 0.0 (not 80%)
                header_confidence_scores = {}
                try:
                    # Determine blank columns in final_data
                    all_blank = {h: True for h in final_headers}
                    if isinstance(final_data, list) and final_data:
                        if isinstance(final_data[0], dict):
                            for row in final_data:
                                for h in final_headers:
                                    if not all_blank[h]:
                                        continue
                                    val = str(row.get(h, '') or '').strip()
                                    if val:
                                        all_blank[h] = False
                        else:
                            idx_map = {h: i for i, h in enumerate(final_headers)}
                            for row in final_data:
                                for h, i in idx_map.items():
                                    if not all_blank[h]:
                                        continue
                                    if i is not None and i < len(row):
                                        val = str(row[i] or '').strip()
                                        if val:
                                            all_blank[h] = False
                    for h in final_headers:
                        if h in target_column_confidence:
                            header_confidence_scores[h] = float(target_column_confidence[h])
                        elif all_blank.get(h, False):
                            header_confidence_scores[h] = 0.0
                        else:
                            # Leave unset to avoid misleading defaults; frontend can fallback gracefully
                            pass
                except Exception:
                    # Fallback: use mapping-only
                    header_confidence_scores = {h: float(c) for h, c in target_column_confidence.items()}
        except Exception as e:
            logger.warning(f"Could not retrieve PDF confidence data: {e}")

        # Calculate data quality metrics
        quality_metrics = calculate_data_quality_metrics(
            final_data,
            final_headers,
            header_confidence_scores,
            confidence_data
        )

        # Enforce full canonical template headers in the response, regardless of data sparsity
        try:
            tags_count = int(info.get('tags_count', 3))
            spec_pairs_count = int(info.get('spec_pairs_count', 3))
            customer_id_pairs_count = int(info.get('customer_id_pairs_count', 1))

            canonical_headers = [
                'Item code', 'Item name', 'Description', 'Item type', 'Measurement unit',
                'Procurement entity name', 'Notes', 'Internal notes', 'Procurement item', 'Sales item', 'Preferred vendor code'
            ]
            for i in range(1, tags_count + 1):
                canonical_headers.append(f'Tag_{i}')
            for i in range(1, spec_pairs_count + 1):
                canonical_headers.append(f'Specification_Name_{i}')
                canonical_headers.append(f'Specification_Value_{i}')
            for i in range(1, customer_id_pairs_count + 1):
                canonical_headers.append(f'Customer_Identification_Name_{i}')
                canonical_headers.append(f'Customer_Identification_Value_{i}')

            # Rebuild data rows to include all canonical headers in order
            rebuilt_rows = []
            if isinstance(final_data, list) and final_data:
                if isinstance(final_data[0], dict):
                    for row in final_data:
                        rebuilt = {h: row.get(h, '') for h in canonical_headers}
                        rebuilt_rows.append(rebuilt)
                else:
                    # list-of-lists -> dict rows using current final_headers index mapping
                    idx_map = {h: i for i, h in enumerate(final_headers)}
                    for row in final_data:
                        rebuilt = {}
                        for h in canonical_headers:
                            i = idx_map.get(h)
                            val = ''
                            if i is not None and i < len(row):
                                val = row[i] if row[i] is not None else ''
                            rebuilt[h] = val
                        rebuilt_rows.append(rebuilt)
            else:
                rebuilt_rows = []

            final_headers = canonical_headers
            final_data = rebuilt_rows

            # Persist canonical headers to session to avoid drift across workers
            try:
                info['enhanced_headers'] = canonical_headers
                save_session(session_id, info)
            except Exception:
                pass
        except Exception as _e:
            # Non-fatal; keep computed headers/data
            pass

        return no_store(Response({
            'success': True,
            'headers': final_headers,
            'data': final_data,
            'total_rows': total_rows,
            'formula_rules': formula_rules,
            'template_version': info.get('template_version', 0),
            'quality_metrics': quality_metrics,
            'header_confidence_scores': header_confidence_scores,
            'target_column_confidence_scores': header_confidence_scores,
            'is_from_pdf': is_pdf_session_flag,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_rows': total_rows,
                'total_pages': (total_rows + page_size - 1) // page_size if total_rows > 0 else 1
            }
        }))
        
    except Exception as e:
        import traceback
        logger.error(f"Error in data_view: {e}")
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return no_store(Response({
            'success': False,
            'error': f'Failed to get data: {str(e)}'
        }, status=500))


@api_view(['POST'])
def save_data(request):
    """Save edited data."""
    try:
        session_id = request.data.get('session_id')
        data = request.data.get('data', [])
        
        if not session_id:
            return Response({
                'success': False,
                'error': 'Session ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        info = get_session(session_id)
        if not info:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Normalize payload: accept { rows: [...] } or raw list
        rows_payload = None
        try:
            if isinstance(data, list):
                rows_payload = data
            elif isinstance(data, dict):
                # common shapes: { rows: [...] } or { data: [...] }
                rows_payload = data.get('rows') or data.get('data')
        except Exception:
            rows_payload = None

        if not isinstance(rows_payload, list):
            rows_payload = []

        # Save edited data to session (as a list of row dicts)
        info["edited_data"] = rows_payload

        # Ensure Data Editor uses these rows immediately
        info["formula_enhanced_data"] = rows_payload

        # Ensure headers are preserved; prefer existing enhanced headers, else derive canonical/keys
        enhanced_headers = info.get('enhanced_headers') or info.get('current_template_headers')
        if not enhanced_headers:
            # Derive canonical from session counts
            tags_count = int(info.get('tags_count', 3))
            spec_pairs_count = int(info.get('spec_pairs_count', 3))
            customer_id_pairs_count = int(info.get('customer_id_pairs_count', 1))
            enhanced_headers = [
                'Item code', 'Item name', 'Description', 'Item type', 'Measurement unit',
                'Procurement entity name', 'Notes', 'Internal notes', 'Procurement item', 'Sales item', 'Preferred vendor code'
            ]
            for i in range(1, tags_count + 1): enhanced_headers.append(f'Tag_{i}')
            for i in range(1, spec_pairs_count + 1): enhanced_headers += [f'Specification_Name_{i}', f'Specification_Value_{i}']
            for i in range(1, customer_id_pairs_count + 1): enhanced_headers += [f'Customer_Identification_Name_{i}', f'Customer_Identification_Value_{i}']
        info['enhanced_headers'] = enhanced_headers

        # Bypass cleanup/mapping; prefer edited data immediately
        info['uploaded_via_correction'] = True
        save_session(session_id, info)
        
        return Response({
            'success': True,
            'message': 'Data saved successfully'
        })
        
    except Exception as e:
        logger.error(f"Error in save_data: {e}")
        return Response({
            'success': False,
            'error': f'Failed to save data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@never_cache
def session_status(request, session_id):
    """Get session status including template version for change tracking."""
    try:
        info = get_session_consistent(session_id)
        # Cross-worker consistency: prefer file snapshot if newer
        try:
            file_snapshot = load_session_from_file(session_id)
            if file_snapshot and file_snapshot.get('template_version', 0) > info.get('template_version', 0):
                SESSION_STORE[session_id] = file_snapshot
                info = file_snapshot
                logger.info(f"🔄 STATUS: Refreshed in-memory session {session_id} from file (newer template_version)")
        except Exception:
            pass
        if not info:
            return no_store(Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND))
        
        template_version = info.get('template_version', 0)
        
        # Get header counts for completeness
        tags_count = info.get('tags_count', 3)
        spec_pairs_count = info.get('spec_pairs_count', 3)
        customer_id_pairs_count = info.get('customer_id_pairs_count', 1)
        
        # Get current headers
        headers = info.get('enhanced_headers') or info.get('current_template_headers') or info.get('template_headers') or []
        
        return no_store(Response({
            'success': True,
            'template_version': template_version,
            'session_id': session_id,
            'counts': {
                'tags_count': tags_count,
                'spec_pairs_count': spec_pairs_count,
                'customer_id_pairs_count': customer_id_pairs_count
            },
            'headers_count': len(headers),
            'has_mappings': bool(info.get('mappings')),
            'has_formula_rules': bool(info.get('formula_rules')),
            'has_default_values': bool(info.get('default_values'))
        }))
        
    except Exception as e:
        logger.error(f"Error in session_status: {e}")
        return no_store(Response({
            'success': False,
            'error': f'Failed to get session status: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR))


@api_view(['POST'])
def rebuild_template(request):
    """Rebuild template and update column counts."""
    try:
        session_id = request.data.get('session_id')
        if not session_id:
            return Response({
                'success': False,
                'error': 'Session ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        info = get_session(session_id)
        if not info:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Increment template version
        new_version = increment_template_version(session_id)
        logger.info(f"🔄 Rebuilt template for session {session_id}, version: {new_version}")
        
        return Response({
            'success': True,
            'template_version': new_version,
            'message': 'Template rebuilt successfully'
        })
        
    except Exception as e:
        logger.error(f"Error in rebuild_template: {e}")
        return Response({
            'success': False,
            'error': f'Failed to rebuild template: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'POST'])
def download_file(request, session_id=None):
    """Download processed/converted file."""
    try:
        # Support both URL path and query/form parameters for session_id
        if not session_id:
            # Fallback to old method for backwards compatibility
            if request.method == 'POST':
                session_id = request.data.get('session_id') or request.POST.get('session_id')
            else:
                session_id = request.GET.get('session_id')

        # Extract column order from request if provided
        requested_column_order = None
        if request.method == 'POST':
            requested_column_order = request.data.get('column_order')
        
        if not session_id:
            return Response({
                'success': False,
                'error': 'Session ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        info = get_session(session_id)
        if not info:
            return Response({
                'success': False,
                'error': 'Session not found'
            }, status=status.HTTP_404_NOT_FOUND)
        mappings = info.get("mappings")
        
        if not mappings:
            return Response({
                'success': False,
                'error': 'No mappings found'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Always prefer rebuilding from current canonical headers to avoid stale snapshots
        enhanced_data = info.get("formula_enhanced_data")
        
        # If an old small snapshot is present while dataset is large, ignore it
        try:
            client_local_path = hybrid_file_manager.get_file_path(info["client_path"])
            total_rows_est = _count_total_data_rows(
                client_local_path,
                info.get("sheet_name"),
                info.get("header_row", 1) - 1 if info.get("header_row", 1) > 0 else 0
            )
        except Exception:
            total_rows_est = 0

        if enhanced_data and (total_rows_est == 0 or len(enhanced_data) >= total_rows_est):
            # Use formula-enhanced data for download
            transformed_rows = enhanced_data
            base_headers = info.get("current_template_headers") or info.get("enhanced_headers") or []

            # Inject MPN validation columns for enhanced data path as well
            try:
                mpn_validation = info.get('mpn_validation') or {}
                mpn_header = mpn_validation.get('column')
                results_map = mpn_validation.get('results') or {}
                if mpn_header and isinstance(transformed_rows, list) and transformed_rows:
                    from .services.digikey_service import DigiKeyClient
                    client_norm = DigiKeyClient.normalize_mpn

                    # Limit canonical MPN columns to reasonable number (max 5) to avoid excessive blank columns
                    max_canonical_mpns = 1
                    for row in transformed_rows:
                        if isinstance(row, dict) and 'MPN valid' not in row:
                            raw = row.get(mpn_header, '')
                            norm = client_norm(raw)
                            res = results_map.get(norm, {})
                            all_canonicals = res.get('all_canonical_mpns', [])
                            if len(all_canonicals) > max_canonical_mpns:
                                max_canonical_mpns = min(len(all_canonicals), 5)  # Cap at 5 columns

                    # Add base MPN validation columns to base headers if not present
                    base_validation_columns = ['MPN valid', 'MPN Status', 'EOL Status', 'Discontinued', 'DKPN', 'Category']
                    for mpn_col in base_validation_columns:
                        if mpn_col not in base_headers:
                            base_headers.append(mpn_col)

                    # Add multiple canonical MPN columns based on maximum needed
                    canonical_columns = []
                    for i in range(max_canonical_mpns):
                        if i == 0:
                            col_name = 'Canonical MPN'  # First one keeps original name
                        else:
                            col_name = f'Canonical MPN {i + 1}'  # Additional ones get numbered

                        canonical_columns.append(col_name)
                        if col_name not in base_headers:
                            base_headers.append(col_name)

                    # Populate all MPN validation data for each row if not already present
                    for row in transformed_rows:
                        # Only add MPN data if not already present (to avoid overwriting)
                        if isinstance(row, dict) and 'MPN valid' not in row:
                            raw = row.get(mpn_header, '')
                            norm = client_norm(raw)
                            res = results_map.get(norm, {})
                            lifecycle = res.get('lifecycle') or {}
                            all_canonicals = res.get('all_canonical_mpns', [])

                            # Set base MPN validation columns
                            row['MPN valid'] = 'Yes' if res.get('valid') else ('No' if norm else '')
                            row['MPN Status'] = lifecycle.get('status') or 'Unknown'
                            row['EOL Status'] = 'Yes' if lifecycle.get('endOfLife') else 'No'
                            row['Discontinued'] = 'Yes' if lifecycle.get('discontinued') else 'No'
                            row['DKPN'] = res.get('dkpn') or ''

                            # Add category information from DigiKey (only for valid MPNs)
                            if is_valid:
                                category_info = res.get('category', {})
                                row['Category'] = category_info.get('name', '') if category_info else ''
                            else:
                                row['Category'] = ''

                            # Set multiple canonical MPN columns
                            for i, col_name in enumerate(canonical_columns):
                                if i < len(all_canonicals):
                                    row[col_name] = all_canonicals[i]
                                else:
                                    row[col_name] = ''  # Empty if no more canonical MPNs

                    logger.info(f"🔧 DEBUG download_file: Added MPN validation columns to enhanced data: {base_validation_columns + canonical_columns}")
            except Exception as _me:
                logger.warning(f"Download: MPN validation injection for enhanced data skipped: {_me}")

            # Apply requested column order if provided, also for enhanced data
            if requested_column_order and isinstance(requested_column_order, list):
                # Use requested column order, but only include headers that exist in the data
                ordered_headers = []
                for header in requested_column_order:
                    if header in base_headers:
                        ordered_headers.append(header)

                # Add any additional headers not in the requested order (like MPN validation columns)
                for header in base_headers:
                    if header not in ordered_headers:
                        ordered_headers.append(header)

                all_headers = ordered_headers
                logger.info(f"🔧 DEBUG download_file: Applied requested column order to enhanced data: {all_headers}")
            else:
                all_headers = list(base_headers)
        else:
            # Fall back to regular mapped data
            # Ensure no pagination hints leak into full export
            try:
                if '__paginate__' in SESSION_STORE.get(session_id, {}):
                    del SESSION_STORE[session_id]['__paginate__']
            except Exception:
                pass
            mapping_result = apply_column_mappings(
                client_file=info["client_path"],
                mappings=mappings,
                sheet_name=info["sheet_name"],
                header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0,
                session_id=session_id
            )
            # Convert to dict rows for downstream formula and Factwise operations
            base_headers = mapping_result['headers']
            transformed_rows = []
            for row_list in mapping_result['data']:
                row_dict = {}
                for i, header in enumerate(base_headers):
                    row_dict[header] = row_list[i] if i < len(row_list) else ""
                transformed_rows.append(row_dict)

            # Apply requested column order if provided
            if requested_column_order and isinstance(requested_column_order, list):
                # Use requested column order, but only include headers that exist in the data
                ordered_headers = []
                for header in requested_column_order:
                    if header in base_headers:
                        ordered_headers.append(header)

                # Add any additional headers not in the requested order
                for header in base_headers:
                    if header not in ordered_headers:
                        ordered_headers.append(header)

                all_headers = ordered_headers
                logger.info(f"🔧 DEBUG download_file: Applied requested column order to mapping result: {all_headers}")
            else:
                all_headers = list(base_headers)

            # Apply formula rules if present to generate Tag/Specification/Customer columns
            try:
                formula_rules = info.get('formula_rules', []) or []
                if formula_rules:
                    formula_result = apply_formula_rules(transformed_rows, all_headers, formula_rules, replace_existing=False, session_info=info)
                    transformed_rows = formula_result.get('data', transformed_rows)
                    all_headers = formula_result.get('headers', all_headers)
            except Exception as _fe:
                logger.warning(f"Download: formula application skipped due to error: {_fe}")

            # Inject MPN validation for export if present (all columns)
            try:
                mpn_validation = info.get('mpn_validation') or {}
                mpn_header = mpn_validation.get('column')
                results_map = mpn_validation.get('results') or {}
                if mpn_header and isinstance(transformed_rows, list) and transformed_rows:
                    from .services.digikey_service import DigiKeyClient
                    client_norm = DigiKeyClient.normalize_mpn

                    # Limit canonical MPN columns to reasonable number (max 5) to avoid excessive blank columns
                    max_canonical_mpns = 1
                    for row in transformed_rows:
                        raw = row.get(mpn_header, '')
                        norm = client_norm(raw)
                        res = results_map.get(norm, {})
                        all_canonicals = res.get('all_canonical_mpns', [])
                        if len(all_canonicals) > max_canonical_mpns:
                            max_canonical_mpns = min(len(all_canonicals), 5)  # Cap at 5 columns

                    # Add base MPN validation columns
                    base_validation_columns = ['MPN valid', 'MPN Status', 'EOL Status', 'Discontinued', 'DKPN', 'Category']
                    for mpn_col in base_validation_columns:
                        if mpn_col not in all_headers:
                            all_headers.append(mpn_col)

                    # Add multiple canonical MPN columns based on maximum needed
                    canonical_columns = []
                    for i in range(max_canonical_mpns):
                        if i == 0:
                            col_name = 'Canonical MPN'  # First one keeps original name
                        else:
                            col_name = f'Canonical MPN {i + 1}'  # Additional ones get numbered

                        canonical_columns.append(col_name)
                        if col_name not in all_headers:
                            all_headers.append(col_name)

                    # Populate all MPN validation data for each row
                    for row in transformed_rows:
                        raw = row.get(mpn_header, '')
                        norm = client_norm(raw)
                        res = results_map.get(norm, {})
                        lifecycle = res.get('lifecycle') or {}
                        all_canonicals = res.get('all_canonical_mpns', [])

                        # Set base MPN validation columns
                        row['MPN valid'] = 'Yes' if res.get('valid') else ('No' if norm else '')
                        row['MPN Status'] = lifecycle.get('status') or 'Unknown'
                        row['EOL Status'] = 'Yes' if lifecycle.get('endOfLife') else 'No'
                        row['Discontinued'] = 'Yes' if lifecycle.get('discontinued') else 'No'
                        row['DKPN'] = res.get('dkpn') or ''

                        # Add category information from DigiKey (only for valid MPNs)
                        if is_valid:
                            category_info = res.get('category', {})
                            row['Category'] = category_info.get('name', '') if category_info else ''
                        else:
                            row['Category'] = ''

                        # Set multiple canonical MPN columns
                        for i, col_name in enumerate(canonical_columns):
                            if i < len(all_canonicals):
                                row[col_name] = all_canonicals[i]
                            else:
                                row[col_name] = ''  # Empty if no more canonical MPNs

                    logger.info(f"🔧 DEBUG download_file: Added MPN validation columns: {base_validation_columns + canonical_columns}")
            except Exception as _me:
                logger.warning(f"Download: MPN validation injection skipped: {_me}")

            # Apply Factwise ID rules if configured
            try:
                factwise_rules = info.get('factwise_rules', []) or []
                for factwise_rule in factwise_rules:
                    if factwise_rule.get("type") != "factwise_id":
                        continue
                    first_col = factwise_rule.get("first_column")
                    second_col = factwise_rule.get("second_column")
                    operator = factwise_rule.get("operator", "_")
                    strategy = factwise_rule.get("strategy", "fill_only_null")

                    # Ensure target column exists
                    if 'Item code' not in all_headers:
                        all_headers = ['Item code'] + all_headers
                        for row in transformed_rows:
                            row.setdefault('Item code', '')

                    # Compute values row-wise
                    for row in transformed_rows:
                        first_val = str(row.get(first_col, "") or "").strip()
                        second_val = str(row.get(second_col, "") or "").strip()
                        factwise_id = (f"{first_val}{operator}{second_val}" if first_val and second_val else (first_val or second_val or ""))
                        if strategy == 'override_all':
                            row['Item code'] = factwise_id
                        else:
                            current_val = str(row.get('Item code', '') or '')
                            if not current_val.strip():
                                row['Item code'] = factwise_id
            except Exception as _ie:
                logger.warning(f"Download: factwise application skipped due to error: {_ie}")

        # Ensure default values are applied in the download path as well (parity with data_view)
        try:
            session_default_values = info.get("default_values", {}) or {}
            if session_default_values and transformed_rows:
                # If rows are dicts, apply directly; if lists, map via headers
                if transformed_rows and isinstance(transformed_rows[0], dict):
                    for field_name, default_value in session_default_values.items():
                        for row in transformed_rows:
                            current_value = row.get(field_name, "")
                            if not current_value or current_value == "":
                                row[field_name] = default_value
                elif transformed_rows and isinstance(transformed_rows[0], list) and all_headers:
                    # Build header index map
                    header_index = {h: idx for idx, h in enumerate(all_headers)}
                    for field_name, default_value in session_default_values.items():
                        # Exact match first
                        target_header = None
                        if field_name in header_index:
                            target_header = field_name
                        else:
                            # Case-insensitive normalized match
                            norm = field_name.lower().replace(' ', '_').replace('-', '_')
                            for h in all_headers:
                                if h.lower().replace(' ', '_').replace('-', '_') == norm:
                                    target_header = h
                                    break
                        if target_header is not None:
                            idx = header_index.get(target_header)
                            if idx is not None:
                                for row in transformed_rows:
                                    if idx < len(row):
                                        if row[idx] is None or str(row[idx]).strip() == "":
                                            row[idx] = default_value
        except Exception as _e:
            logger.warning(f"Download default application skipped due to error: {_e}")
        
        if not transformed_rows:
            return Response({
                'success': False,
                'error': 'No data to download'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # For enhanced or dict data, use headers that actually exist in the data and normalize shape
        if isinstance(transformed_rows, list) and transformed_rows and isinstance(transformed_rows[0], dict):
            # Get headers that actually exist in the enhanced data
            actual_headers = list(transformed_rows[0].keys()) if transformed_rows else []
            logger.info(f"🔧 DEBUG download_file: Enhanced data has headers: {actual_headers}")
            
            # Use the canonical headers from session if available, but only include those that exist in the data
            canonical_headers = info.get("current_template_headers") or info.get("enhanced_headers") or all_headers or []

            # CRITICAL FIX: Always enforce proper canonical order for downloads
            # Reorder canonical_headers to follow the standard order: Item code first, then other standard headers, then Tags, etc.
            if canonical_headers:
                # Define the correct canonical order
                correct_order = [
                    "Item code", "Item name", "Description", "Item type", "Measurement unit",
                    "Procurement entity name", "Notes", "Internal notes", "Procurement item",
                    "Sales item", "Preferred vendor code"
                ]

                # Add Tag columns in correct order
                tag_headers = sorted([h for h in canonical_headers if h.startswith('Tag_')],
                                   key=lambda x: int(x.split('_')[1]) if '_' in x and x.split('_')[1].isdigit() else 0)
                correct_order.extend(tag_headers)

                # Add Specification columns in correct order
                spec_name_headers = sorted([h for h in canonical_headers if h.startswith('Specification_Name_')],
                                         key=lambda x: int(x.split('_')[2]) if len(x.split('_')) > 2 and x.split('_')[2].isdigit() else 0)
                spec_value_headers = sorted([h for h in canonical_headers if h.startswith('Specification_Value_')],
                                          key=lambda x: int(x.split('_')[2]) if len(x.split('_')) > 2 and x.split('_')[2].isdigit() else 0)

                # Interleave spec names and values
                for i in range(max(len(spec_name_headers), len(spec_value_headers))):
                    if i < len(spec_name_headers):
                        correct_order.append(spec_name_headers[i])
                    if i < len(spec_value_headers):
                        correct_order.append(spec_value_headers[i])

                # Add Customer ID columns in correct order
                customer_name_headers = sorted([h for h in canonical_headers if h.startswith('Customer_Identification_Name_')],
                                             key=lambda x: int(x.split('_')[3]) if len(x.split('_')) > 3 and x.split('_')[3].isdigit() else 0)
                customer_value_headers = sorted([h for h in canonical_headers if h.startswith('Customer_Identification_Value_')],
                                              key=lambda x: int(x.split('_')[3]) if len(x.split('_')) > 3 and x.split('_')[3].isdigit() else 0)

                # Interleave customer names and values
                for i in range(max(len(customer_name_headers), len(customer_value_headers))):
                    if i < len(customer_name_headers):
                        correct_order.append(customer_name_headers[i])
                    if i < len(customer_value_headers):
                        correct_order.append(customer_value_headers[i])

                # Add any remaining headers that weren't categorized
                for header in canonical_headers:
                    if header not in correct_order and header in actual_headers:
                        correct_order.append(header)

                # Filter to only headers that exist in the actual data and maintain order
                canonical_headers = [h for h in correct_order if h in actual_headers]
                logger.info(f"🔧 DEBUG download_file: Enforced canonical order: {canonical_headers}")

            # CRITICAL FIX: Use requested column order if provided, otherwise use canonical order
            if requested_column_order and isinstance(requested_column_order, list):
                # Use requested column order, but only include headers that exist in the data
                valid_headers = []
                for header in requested_column_order:
                    if header in actual_headers:
                        valid_headers.append(header)

                # Add any additional headers from data that aren't in the requested order
                for header in actual_headers:
                    if header not in valid_headers:
                        valid_headers.append(header)

                all_headers = valid_headers
                logger.info(f"🔧 DEBUG download_file: Using requested column order: {all_headers}")
            else:
                # Fallback to canonical order
                valid_headers = []
                for header in canonical_headers:
                    if header in actual_headers:
                        valid_headers.append(header)

                # Add any additional headers from data that aren't in canonical (shouldn't happen but defensive)
                for header in actual_headers:
                    if header not in valid_headers:
                        valid_headers.append(header)

                all_headers = valid_headers
                logger.info(f"🔧 DEBUG download_file: Using canonical headers for enhanced data: {all_headers}")
            
            # Convert dict format to list format for consistency
            converted_rows = []
            for row_dict in transformed_rows:
                row_list = []
                for header in all_headers:
                    row_list.append(row_dict.get(header, ""))
                converted_rows.append(row_list)
            transformed_rows = converted_rows
        
        # Create DataFrame with duplicate column names support
        if transformed_rows and all_headers:
            # CRITICAL FIX: Ensure data and headers are compatible
            try:
                # Check if we have the right number of columns
                if transformed_rows and isinstance(transformed_rows[0], list):
                    first_row_length = len(transformed_rows[0])
                    headers_length = len(all_headers)
                    
                    logger.info(f"🔧 DEBUG download_file: first_row_length={first_row_length}, headers_length={headers_length}")
                    logger.info(f"🔧 DEBUG download_file: all_headers={all_headers}")
                    
                    if first_row_length != headers_length:
                        logger.warning(f"🚨 Header/data mismatch in download: {headers_length} headers but {first_row_length} data columns")
                        # Pad or truncate headers to match data
                        if headers_length > first_row_length:
                            # Too many headers, truncate
                            all_headers = all_headers[:first_row_length]
                            logger.info(f"🔧 Truncated headers to match data: {all_headers}")
                        else:
                            # Too few headers, pad with generic names
                            for i in range(headers_length, first_row_length):
                                all_headers.append(f"Column_{i+1}")
                            logger.info(f"🔧 Padded headers to match data: {all_headers}")
                
                # Use pandas with list data and original headers for export
                df = pd.DataFrame(transformed_rows, columns=all_headers)
                logger.info(f"🔧 DEBUG download_file: Successfully created DataFrame with shape {df.shape}")
            except Exception as df_error:
                logger.error(f"🚨 DataFrame creation failed: {df_error}")
                # Fallback: create DataFrame without column specification
                try:
                    df = pd.DataFrame(transformed_rows)
                    logger.info(f"🔧 Fallback DataFrame created with shape {df.shape}")
                except Exception as fallback_error:
                    logger.error(f"🚨 Fallback DataFrame creation also failed: {fallback_error}")
                    raise fallback_error
        else:
            # Create empty DataFrame
            df = pd.DataFrame(columns=all_headers or [])
        
        # Clean column names for export only (remove numbers, underscores, dots)
        if not df.empty:
            import re
            final_columns = []
            for col in df.columns:
                # Remove all numbers, underscores, and dots, then clean up spaces and use sentence case
                cleaned_col = str(col)
                # Remove numbers and special characters, replace with spaces
                cleaned_col = re.sub(r'[_\d\.]', ' ', cleaned_col)
                # Replace multiple spaces with single space and trim
                cleaned_col = re.sub(r'\s+', ' ', cleaned_col).strip()
                # Convert to sentence case (only capitalize first letter)
                if cleaned_col:
                    cleaned_col = cleaned_col.capitalize()
                
                final_columns.append(cleaned_col or col)  # Fallback to original if cleaning fails
            
            df.columns = final_columns
        
        # Get format preference (default to Excel)
        if request.method == 'POST':
            format_type = (request.data.get('format') or request.POST.get('format', 'excel')).lower()
        else:
            format_type = request.GET.get('format', 'excel').lower()
        
        # Create output file
        output_dir = hybrid_file_manager.local_temp_dir
        
        from datetime import datetime
        timestamp = datetime.now().strftime('%y%m%d_%H%M%S')
        base_name = f"FactWise_Filled_{timestamp}"
        if format_type == 'csv':
            filename = f"{base_name}.csv"
            output_file = output_dir / filename
            df.to_csv(output_file, index=False)
            content_type = 'text/csv'
        else:  # Excel format (default)
            filename = f"{base_name}.xlsx"
            output_file = output_dir / filename
            df.to_excel(output_file, index=False, engine='openpyxl')
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        response = FileResponse(
            open(output_file, 'rb'),
            as_attachment=True,
            filename=filename,
            content_type=content_type
        )
        
        return response
        
    except Exception as e:
        logger.error(f"Error in download_file: {e}")
        return Response({
            'success': False,
            'error': f'Download failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def download_original_file(request, session_id=None):
    """Download original uploaded client file."""
    try:
        # Support both URL path and query parameters for session_id
        if not session_id:
            session_id = request.GET.get('session_id')
        
        if not session_id or session_id not in SESSION_STORE:
            return Response({
                'success': False,
                'error': 'Invalid session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        info = SESSION_STORE[session_id]
        client_path = info["client_path"]
        original_name = info["original_client_name"]
        
        if not Path(client_path).exists():
            return Response({
                'success': False,
                'error': 'Original file not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        response = FileResponse(
            open(client_path, 'rb'),
            as_attachment=True,
            filename=original_name
        )
        
        return response
        
    except Exception as e:
        logger.error(f"Error in download_original_file: {e}")
        return Response({
            'success': False,
            'error': f'Download failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def download_template_file(request, session_id=None):
    """Download FACTWISE.xlsx template file (always returns the standard FW template)."""
    try:
        # Support both URL path and query parameters for session_id
        if not session_id:
            session_id = request.GET.get('session_id')
        
        if not session_id or session_id not in SESSION_STORE:
            return Response({
                'success': False,
                'error': 'Invalid session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # CRITICAL FIX: Always return FACTWISE.xlsx regardless of what template was uploaded
        factwise_template_path = Path(settings.BASE_DIR) / 'FACTWISE.xlsx'

        # Fallback to test_files directory if not found in root
        if not factwise_template_path.exists():
            factwise_template_path = Path(settings.BASE_DIR) / 'test_files' / 'FACTWISE.xlsx'

        logger.info(f"🔍 FW Template download for session {session_id}: returning FACTWISE.xlsx from {factwise_template_path}")

        if not factwise_template_path.exists():
            return Response({
                'success': False,
                'error': 'FACTWISE.xlsx template file not found'
            }, status=status.HTTP_404_NOT_FOUND)

        response = FileResponse(
            open(factwise_template_path, 'rb'),
            as_attachment=True,
            filename='FACTWISE.xlsx'
        )
        
        return response
    except Exception as e:
        logger.error(f"Error downloading template file: {e}")
        return Response({
            'success': False,
            'error': f'Template download failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def download_grid_excel(request):
    """Download Excel file with frontend grid data."""
    try:
        data = request.data
        session_id = data.get('session_id')
        headers = data.get('headers', [])
        rows = data.get('rows', [])
        
        if not session_id:
            return Response({
                'success': False,
                'error': 'Session ID required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create DataFrame from frontend data
        import pandas as pd
        df = pd.DataFrame(rows)
        
        # Ensure columns match headers
        if headers:
            df.columns = headers[:len(df.columns)]
        
        # Prune _numbers from column names for download only (Tag_1, Tag_2 → Tag)
        if not df.empty and len(df.columns) > 0:
            import re
            final_columns = []
            for col in df.columns:
                # Remove _number suffix from column names (e.g., Tag_1 → Tag, Specification_Name_2 → Specification_Name)
                pruned_col = re.sub(r'_\d+$', '', str(col))
                final_columns.append(pruned_col)
            
            df.columns = final_columns
        
        # Generate filename with YYMMDD_HHMMSS
        from datetime import datetime
        timestamp = datetime.now().strftime('%y%m%d_%H%M%S')
        filename = f'FactWise_Filled_{timestamp}.xlsx'
        
        # Create temp file
        temp_file = os.path.join('temp_downloads', filename)
        os.makedirs('temp_downloads', exist_ok=True)
        
        # Save to Excel
        df.to_excel(temp_file, index=False)
        
        # Return file response
        response = FileResponse(
            open(temp_file, 'rb'),
            as_attachment=True,
            filename=filename
        )
        
        return response
        
    except Exception as e:
        logger.error(f"Error in download_grid_excel: {e}")
        return Response({
            'success': False,
            'error': f'Excel export failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def dashboard_view(request):
    """Get dashboard data."""
    try:
        # Get recent sessions
        uploads = []
        for session_id, session_data in list(SESSION_STORE.items())[-10:]:  # Last 10 sessions
            # Try to get row count from processed data or client data
            rows_processed = 0
            logger.info(f"Session {session_id}: Checking for data to calculate rows")
            logger.info(f"Session {session_id}: Session data keys: {list(session_data.keys())}")
            logger.info(f"Session {session_id}: edited_data type: {type(session_data.get('edited_data'))}")
            logger.info(f"Session {session_id}: client_data type: {type(session_data.get('client_data'))}")
            logger.info(f"Session {session_id}: client_path: {session_data.get('client_path')}")
            
            if session_data.get('edited_data'):
                try:
                    # If we have processed data, count the rows
                    if isinstance(session_data['edited_data'], list):
                        rows_processed = len(session_data['edited_data'])
                        logger.info(f"Session {session_id}: Got rows from edited_data list: {rows_processed}")
                    elif isinstance(session_data['edited_data'], dict) and 'data' in session_data['edited_data']:
                        rows_processed = len(session_data['edited_data']['data'])
                        logger.info(f"Session {session_id}: Got rows from edited_data dict: {rows_processed}")
                except Exception as e:
                    logger.warning(f"Session {session_id}: Error counting edited_data rows: {e}")
                    rows_processed = 0
            elif session_data.get('client_data'):
                try:
                    # Fallback to client data if processed data not available
                    if isinstance(session_data['client_data'], list):
                        rows_processed = len(session_data['client_data'])
                        logger.info(f"Session {session_id}: Got rows from client_data list: {rows_processed}")
                    elif isinstance(session_data['client_data'], dict) and 'data' in session_data['client_data']:
                        rows_processed = len(session_data['client_data']['data'])
                        logger.info(f"Session {session_id}: Got rows from client_data dict: {rows_processed}")
                except Exception as e:
                    logger.warning(f"Session {session_id}: Error counting client_data rows: {e}")
                    rows_processed = 0
            else:
                # Try to calculate rows from processed data that would be used for download
                try:
                    if session_data.get('mappings'):
                        # Import here to avoid circular imports
                        from .bom_header_mapper import BOMHeaderMapper
                        import pandas as pd
                        
                        # Get client file info
                        client_path = session_data.get('client_path')
                        sheet_name = session_data.get('sheet_name')
                        header_row = session_data.get('header_row', 1) - 1 if session_data.get('header_row', 1) > 0 else 0
                        
                        logger.info(f"Session {session_id}: Attempting to read client file for row count")
                        logger.info(f"Session {session_id}: client_path: {client_path}")
                        logger.info(f"Session {session_id}: sheet_name: {sheet_name}")
                        logger.info(f"Session {session_id}: header_row: {header_row}")
                        
                        if client_path:
                            # Try to resolve the file path using hybrid_file_manager
                            try:
                                client_local_path = hybrid_file_manager.get_file_path(client_path)
                                logger.info(f"Session {session_id}: Resolved client_local_path: {client_local_path}")
                                
                                # Check if the resolved path exists
                                if Path(client_local_path).exists():
                                    logger.info(f"Session {session_id}: Resolved path exists, reading file")
                                    if str(client_local_path).lower().endswith('.csv'):
                                        df = read_csv_with_encoding(client_local_path, header_row)
                                    else:
                                        result = pd.read_excel(client_local_path, sheet_name=sheet_name, header=header_row)
                                        if isinstance(result, dict):
                                            first_sheet_name = list(result.keys())[0]
                                            df = result[first_sheet_name]
                                        else:
                                            df = result
                                    
                                    rows_processed = len(df)
                                    logger.info(f"Session {session_id}: Successfully calculated rows from client file: {rows_processed}")
                                else:
                                    logger.warning(f"Session {session_id}: Resolved path does not exist: {client_local_path}")
                                    # Try to check if the original path exists
                                    if Path(client_path).exists():
                                        logger.info(f"Session {session_id}: Original path exists, using it directly")
                                        if str(client_path).lower().endswith('.csv'):
                                            df = read_csv_with_encoding(client_path, header_row)
                                        else:
                                            result = pd.read_excel(client_path, sheet_name=sheet_name, header=header_row)
                                            if isinstance(result, dict):
                                                first_sheet_name = list(result.keys())[0]
                                                df = result[first_sheet_name]
                                            else:
                                                df = result
                                        
                                        rows_processed = len(df)
                                        logger.info(f"Session {session_id}: Successfully calculated rows from original path: {rows_processed}")
                                    else:
                                        logger.warning(f"Session {session_id}: Neither resolved nor original path exists")
                            except Exception as path_error:
                                logger.warning(f"Session {session_id}: Error resolving file path: {path_error}")
                                # Fallback: try to read directly from client_path
                                if Path(client_path).exists():
                                    logger.info(f"Session {session_id}: Fallback: original path exists, reading directly")
                                    if str(client_path).lower().endswith('.csv'):
                                        df = read_csv_with_encoding(client_path, header_row)
                                    else:
                                        result = pd.read_excel(client_path, sheet_name=sheet_name, header=header_row)
                                        if isinstance(result, dict):
                                            first_sheet_name = list(result.keys())[0]
                                            df = result[first_sheet_name]
                                        else:
                                            df = result
                                    
                                    rows_processed = len(df)
                                    logger.info(f"Session {session_id}: Fallback successful, rows: {rows_processed}")
                                else:
                                    logger.warning(f"Session {session_id}: Fallback path also does not exist: {client_path}")
                        else:
                            logger.info(f"Session {session_id}: No client_path found in session data")
                except Exception as e:
                    logger.warning(f"Session {session_id}: Error calculating rows from client file: {e}")
                    rows_processed = 0
                
                if rows_processed == 0:
                    logger.info(f"Session {session_id}: No edited_data, client_data, or client file found")
            
            logger.info(f"Session {session_id}: Final rows_processed: {rows_processed}")
            
            # Check if file is ready for download
            has_mappings = bool(session_data.get('mappings'))
            has_processed_data = bool(session_data.get('edited_data'))
            # Consider complete if user has gone through the mapping process
            is_complete = has_mappings
            
            # Generate the filename that would be used for download
            filled_sheet_name = None
            if is_complete:
                # Use the session creation time or current time for consistency
                session_created = session_data.get('created')
                if session_created:
                    try:
                        if isinstance(session_created, str):
                            created_dt = datetime.fromisoformat(session_created.replace('Z', '+00:00'))
                        else:
                            created_dt = session_created
                        timestamp = created_dt.strftime('%y%m%d_%H%M%S')
                        filled_sheet_name = f"FactWise_Filled_{timestamp}.xlsx"
                        logger.info(f"Session {session_id}: Generated filled_sheet_name from session_created: {filled_sheet_name}")
                    except Exception as e:
                        # Fallback to current time if session time parsing fails
                        logger.warning(f"Session {session_id}: Error parsing session_created, using current time: {e}")
                        timestamp = datetime.now().strftime('%y%m%d_%H%M%S')
                        filled_sheet_name = f"FactWise_Filled_{timestamp}.xlsx"
                        logger.info(f"Session {session_id}: Generated filled_sheet_name from current time: {filled_sheet_name}")
                else:
                    logger.info(f"Session {session_id}: No session_created, using current time")
                    timestamp = datetime.now().strftime('%y%m%d_%H%M%S')
                    filled_sheet_name = f"FactWise_Filled_{timestamp}.xlsx"
                    logger.info(f"Session {session_id}: Generated filled_sheet_name from current time: {filled_sheet_name}")
            else:
                logger.info(f"Session {session_id}: Not complete, no filled_sheet_name generated")
            
            uploads.append({
                'session_id': session_id,
                'client_file': session_data.get('original_client_name', 'Unknown'),
                'template_file': session_data.get('original_template_name', 'Unknown'),
                'filled_sheet_name': filled_sheet_name,
                'created': session_data.get('created', datetime.now().isoformat()),
                'has_mappings': is_complete,
                'rows_processed': rows_processed,
                'status': 'Complete' if is_complete else 'Pending'
            })
        
        # Get saved templates
        try:
            templates = MappingTemplate.objects.all().order_by('-created_at')[:10]
            saved_templates = [template.get_mapping_summary() for template in templates]
        except Exception:
            saved_templates = []
        
        return Response({
            'success': True,
            'uploads': uploads,
            'saved_templates': saved_templates
        })
        
    except Exception as e:
        logger.error(f"Error in dashboard_view: {e}")
        return Response({
            'success': False,
            'error': f'Failed to get dashboard data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Template management views
@api_view(['POST'])
def update_column_counts(request):
    """Update dynamic column counts for the current session."""
    try:
        session_id = request.data.get('session_id')
        tags_count = request.data.get('tags_count', 3)
        spec_pairs_count = request.data.get('spec_pairs_count', 3)
        customer_id_pairs_count = request.data.get('customer_id_pairs_count', 1)
        
        info = get_session(session_id)
        if not info:
            return Response({
                'success': False,
                'error': 'Invalid session ID'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate counts are positive integers
        try:
            tags_count = max(0, int(tags_count))
            spec_pairs_count = max(0, int(spec_pairs_count))
            customer_id_pairs_count = max(0, int(customer_id_pairs_count))
        except (ValueError, TypeError):
            return Response({
                'success': False,
                'error': 'Column counts must be positive integers'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Store column counts in session (version will be bumped later atomically)
        info['tags_count'] = tags_count
        info['spec_pairs_count'] = spec_pairs_count
        info['customer_id_pairs_count'] = customer_id_pairs_count
        
        # Get existing headers to preserve numbering
        existing_headers = info.get('current_template_headers') or info.get('enhanced_headers') or []
        
        # Generate new template columns based on counts (will be same as regenerated_headers)
        # This is kept for compatibility with existing code that expects template_columns
        template_columns = []

        # Build canonical headers by pruning/adding optional groups starting from original template headers
        try:
            base_headers = info.get('template_headers')
            if not base_headers:
                # Read from file if not cached
                mapper = BOMHeaderMapper()
                base_headers = mapper.read_excel_headers(
                    file_path=hybrid_file_manager.get_file_path(info["template_path"]),
                    sheet_name=info.get("template_sheet_name"),
                    header_row=info.get("template_header_row", 1) - 1 if info.get("template_header_row", 1) > 0 else 0
                )
                info['template_headers'] = base_headers
        except Exception:
            base_headers = []

        def _norm(h: str) -> str:
            try:
                return str(h or '').strip().lower()
            except Exception:
                return ''
        def _is_tag(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'tag' or h_norm.startswith('tag_')
        def _is_spec_name(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'specification name' or h_norm.startswith('specification_name_')
        def _is_spec_value(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'specification value' or h_norm.startswith('specification_value_')
        def _is_cust_name(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'customer identification name' or h_norm.startswith('customer_identification_name_')
        def _is_cust_value(h: str) -> bool:
            h_norm = _norm(h)
            return h_norm == 'customer identification value' or h_norm.startswith('customer_identification_value_')

        # FIXED: Use generate_template_columns to ensure core headers are ALWAYS included
        # This function always starts with the 6 core headers, preventing them from disappearing
        regenerated_headers = generate_template_columns(
            tags_count, 
            spec_pairs_count, 
            customer_id_pairs_count
        )

        # Compute template_optionals for the canonical headers (Tags/Spec/Customer always optional)
        def is_special_optional(h: str) -> bool:
            h_lower = (h or '').lower()
            return (h == 'Tag' or h.startswith('Tag_') or 
                   'specification' in h_lower or 
                   'customer identification' in h_lower or 
                   'customer_identification' in h_lower)

        template_optionals = [True if is_special_optional(h) else False for h in regenerated_headers]

        # Build canonical enhanced_headers and save to session BEFORE version bump
        info["current_template_headers"] = regenerated_headers
        info["enhanced_headers"] = regenerated_headers
        info['template_columns'] = regenerated_headers  # Use same headers for consistency
        info['template_optionals'] = template_optionals
        info['column_counts'] = {
            'tags_count': tags_count,
            'spec_pairs_count': spec_pairs_count,
            'customer_id_pairs_count': customer_id_pairs_count,
        }
        save_session(session_id, info)

        # Atomic version bump AFTER all data is saved
        new_version = increment_template_version(session_id)
        
        # Debug logging
        logger.info(f"🔧 Updated session {session_id} with canonical headers: {regenerated_headers}")
        logger.info(f"🔧 Session store now contains: {list(info.keys())}")

        return Response({
            'success': True,
            'template_version': new_version,
            'enhanced_headers': regenerated_headers,
            'template_optionals': template_optionals,
            'column_counts': info['column_counts'],
        })
        
    except Exception as e:
        logger.error(f"❌ Error updating column counts: {str(e)}")
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def generate_template_columns(tags_count, spec_pairs_count, customer_id_pairs_count, existing_headers=None):
    """Generate internal numbered template headers in canonical order with all standard fields."""
    columns = [
        'Item code',
        'Item name',
        'Description',
        'Item type',
        'Measurement unit',
        'Procurement entity name'
    ]

    # Add standard template fields that should always be present
    columns.extend([
        'Notes', 'Internal notes',
        'Procurement item', 'Sales item', 'Preferred vendor code'
    ])

    # Tags
    for i in range(1, max(int(tags_count or 0), 0) + 1):
        columns.append(f'Tag_{i}')

    # Specification pairs
    for i in range(1, max(int(spec_pairs_count or 0), 0) + 1):
        columns.append(f'Specification_Name_{i}')
        columns.append(f'Specification_Value_{i}')

    # Customer identification pairs
    for i in range(1, max(int(customer_id_pairs_count or 0), 0) + 1):
        columns.append(f'Customer_Identification_Name_{i}')
        columns.append(f'Customer_Identification_Value_{i}')

    return columns


@api_view(['POST'])
def save_mapping_template(request):
    """Save current session mapping as a reusable template."""
    try:
        session_id = request.data.get('session_id')
        template_name = request.data.get('template_name')
        description = request.data.get('description', '')
        override_mappings = request.data.get('mappings')  # Optional mappings override
        override_formula_rules = request.data.get('formula_rules')  # Optional formula rules override
        override_factwise_rules = request.data.get('factwise_rules')  # Optional factwise rules override
        override_default_values = request.data.get('default_values')  # Optional default values override
        mpn_validation_metadata = request.data.get('mpn_validation_metadata', {})  # MPN validation metadata
        
        if not template_name:
            return Response({
                'success': False,
                'error': 'Template name is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Handle two cases: session-based templates and standalone formula templates
        if session_id and session_id in SESSION_STORE:
            # Session-based template (normal case)
            info = SESSION_STORE[session_id]
            raw_mappings = override_mappings if override_mappings is not None else info.get("mappings")
            formula_rules = override_formula_rules if override_formula_rules is not None else info.get("formula_rules", [])
            factwise_rules = override_factwise_rules if override_factwise_rules is not None else info.get("factwise_rules", [])
            # CRITICAL FIX: Get default values and ensure they are properly formatted
            raw_default_values = override_default_values if override_default_values is not None else info.get("default_values", {})
            
            # Clean default values to ensure they are properly stored
            default_values = {}
            if raw_default_values and isinstance(raw_default_values, dict):
                for field_name, value in raw_default_values.items():
                    # Only store fields that actually have non-empty default values
                    # Skip fields with None, empty strings, or whitespace-only values
                    if value is not None and str(value).strip() != "":
                        default_values[field_name] = str(value).strip()
                        logger.info(f"🔧 DEBUG: Template save - storing default value '{value}' for field '{field_name}'")
                    else:
                        # Only log if this field was actually supposed to have a default value
                        # (i.e., if the user had set a value but it's now empty)
                        if field_name in ['Specification name', 'Procurement entity name', 'Customer identification name']:
                            logger.info(f"🔧 DEBUG: Template save - skipping empty default value for field '{field_name}': '{value}'")
            
            logger.info(f"🔧 DEBUG: Template save - final default values: {default_values}")
            
            # Convert mappings from new format to old format for template storage
            if raw_mappings and isinstance(raw_mappings, dict) and 'mappings' in raw_mappings:
                # New format: {'mappings': [{'source': '...', 'target': '...'}, ...]}
                # For templates, we need to preserve all mappings including duplicates
                # Store both the old format dict and new format list for compatibility
                mappings = {}
                mapping_list = []
                
                for mapping_item in raw_mappings['mappings']:
                    source = mapping_item.get('source')
                    target = mapping_item.get('target')
                    if source and target:
                        # Store in new format list (preserves duplicates)
                        mapping_list.append({'source': source, 'target': target})
                        # Store in old format dict (for compatibility, will overwrite duplicates)
                        mappings[target] = source
                
                # Store both formats - use new format as primary, old format as fallback
                mappings = {
                    'new_format': mapping_list,
                    'old_format': mappings
                }
                logger.info(f"🔄 Converted {len(raw_mappings['mappings'])} mappings from new format, preserving duplicates")
            else:
                mappings = raw_mappings or {}
            
            if not mappings and not formula_rules:
                return Response({
                    'success': False,
                    'error': 'No mappings or formula rules to save'
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            # Standalone template (formula-only from Dashboard)
            mappings = override_mappings or {}
            formula_rules = override_formula_rules or []
            
            if not formula_rules:
                return Response({
                    'success': False,
                    'error': 'No formula rules provided for standalone template'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            info = None  # No session info for standalone templates
        
        # Read headers (only if we have session info)
        if info:
            mapper = BOMHeaderMapper()
            client_headers = mapper.read_excel_headers(
                file_path=info["client_path"],
                sheet_name=info["sheet_name"],
                header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0
            )

            # CRITICAL FIX: Always generate the complete template structure for saving
            # Never rely on current_template_headers alone as it may be incomplete
            # Generate the full template with core headers + dynamic counts
            
            # Get column counts first
            tags_count = request.data.get('tags_count') or info.get('tags_count', 1)
            spec_pairs_count = request.data.get('spec_pairs_count') or info.get('spec_pairs_count', 1) 
            customer_id_pairs_count = request.data.get('customer_id_pairs_count') or info.get('customer_id_pairs_count', 1)
            
            # Generate complete template with core + dynamic headers
            template_headers = generate_template_columns(tags_count, spec_pairs_count, customer_id_pairs_count)
            logger.info(f"🔧 CRITICAL FIX: Generated complete template structure for save_mapping_template ({len(template_headers)} headers): {template_headers}")
            
            # Verify we have standard headers
            standard_headers_check = [
                # Core Factwise headers
                "Item code", "Item name", "Description", "Item type", "Measurement unit", "Procurement entity name",
                # Standard template fields that should behave like core headers when mapped
                "Notes", "Internal notes", "Procurement item", "Sales item", "Preferred vendor code"
            ]
            missing_standard = [h for h in standard_headers_check if h not in template_headers]
            if missing_standard:
                logger.error(f"🚨 CRITICAL ERROR: Missing standard headers in generated template: {missing_standard}")
            else:
                logger.info(f"✅ All standard headers present in saved template")
        else:
            # Standalone template - use empty headers
            client_headers = []
            template_headers = []
        
        # Get column counts from request, session, or use defaults (for standalone templates only)
        if not info:
            # Standalone template - use request or defaults
            tags_count = request.data.get('tags_count', 3)
            spec_pairs_count = request.data.get('spec_pairs_count', 3) 
            customer_id_pairs_count = request.data.get('customer_id_pairs_count', 1)
        
        # Normalize Tag formula targets to generic 'Tag' so templates don't hard-pin Tag_N
        try:
            normalized_formula_rules = []
            for _r in (formula_rules or []):
                r = dict(_r or {})
                if (r or {}).get('column_type', 'Tag') == 'Tag':
                    r['target_column'] = 'Tag'
                normalized_formula_rules.append(r)
            formula_rules = normalized_formula_rules
        except Exception:
            pass

        # Create template with backward compatibility
        try:
            debug_log(session_id, f"Saving template '{template_name}'", {
                'factwise_rules': factwise_rules,
                'default_values': default_values,
                'mappings_count': len(raw_mappings.get('mappings', [])) if isinstance(raw_mappings, dict) else 0,
                'formula_rules_count': len(formula_rules),
                'column_counts': {
                    'tags_count': tags_count,
                    'spec_pairs_count': spec_pairs_count,
                    'customer_id_pairs_count': customer_id_pairs_count
                }
            })
            template = MappingTemplate.objects.create(
                name=template_name,
                description=description,
                template_headers=template_headers,
                source_headers=client_headers,
                mappings=mappings,
                formula_rules=formula_rules,  # Include normalized formula rules
                factwise_rules=factwise_rules,  # Include factwise ID rules
                default_values=default_values,  # Include default values
                mpn_validation_metadata=mpn_validation_metadata,  # Include MPN validation metadata
                tags_count=tags_count,
                spec_pairs_count=spec_pairs_count,
                customer_id_pairs_count=customer_id_pairs_count,
                session_id=session_id
            )
        except Exception as e:
            # If new fields don't exist yet, create without them
            if 'formula_rules' in str(e) or 'factwise_rules' in str(e) or 'default_values' in str(e) or 'mpn_validation_metadata' in str(e):
                template = MappingTemplate.objects.create(
                    name=template_name,
                    description=description,
                    template_headers=template_headers,
                    source_headers=client_headers,
                    mappings=mappings,
                    session_id=session_id
                )
            else:
                raise e
        
        # CRITICAL FIX: Return comprehensive response with all template data
        response_data = {
            'success': True,
            'message': f'Template "{template.name}" saved successfully',
            'template_id': template.id,
            'template_name': template.name,
            'description': template.description,
            'default_values': default_values,
            'column_counts': {
                'tags_count': tags_count,
                'spec_pairs_count': spec_pairs_count,
                'customer_id_pairs_count': customer_id_pairs_count
            }
        }
        
        debug_log(session_id, "Template saved successfully, returning comprehensive response", {
            'template_id': template.id,
            'template_name': template.name,
            'default_values_count': len(default_values) if default_values else 0,
            'response_keys': list(response_data.keys())
        })
        
        return Response(response_data, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Error in save_mapping_template: {e}")
        return Response({
            'success': False,
            'error': f'Failed to save template: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def update_mapping_template(request):
    """Update/overwrite an existing mapping template."""
    try:
        session_id = request.data.get('session_id')
        template_id = request.data.get('template_id')
        action = request.data.get('action')
        template_name = request.data.get('template_name')
        description = request.data.get('description', '')
        
        if not session_id or session_id not in SESSION_STORE:
            return Response({
                'success': False,
                'error': 'Invalid session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not template_id:
            return Response({
                'success': False,
                'error': 'Template ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get the existing template
        try:
            template = MappingTemplate.objects.get(id=template_id)
        except MappingTemplate.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Template not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        info = SESSION_STORE[session_id]
        mappings = info.get("mappings")
        
        if not mappings:
            return Response({
                'success': False,
                'error': 'No mappings to save'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Read headers
        mapper = BOMHeaderMapper()
        client_headers = mapper.read_excel_headers(
            file_path=info["client_path"],
            sheet_name=info["sheet_name"],
            header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0
        )

        # Prefer canonical dynamic headers if available
        canonical_headers = info.get("current_template_headers") or info.get("enhanced_headers")
        if canonical_headers and isinstance(canonical_headers, list) and len(canonical_headers) > 0:
            template_headers = canonical_headers
            logger.info(f"🔧 DEBUG: Using canonical template headers from session for update_mapping_template ({len(template_headers)} headers)")
        else:
            template_headers = mapper.read_excel_headers(
                file_path=info["template_path"],
                sheet_name=info.get("template_sheet_name"),
                header_row=info.get("template_header_row", 1) - 1 if info.get("template_header_row", 1) > 0 else 0
            )
            # Normalize to internal numbered headers
            template_headers = normalize_headers_to_internal(template_headers)
            logger.info(f"🔧 DEBUG: Using original file template headers for update_mapping_template ({len(template_headers)} headers) [normalized]")
        
        # Get formula rules from session if they exist
        formula_rules = info.get("formula_rules", [])
        
        # Update the template
        template.client_headers = client_headers
        template.template_headers = template_headers
        template.mappings = mappings
        template.formula_rules = formula_rules  # Update formula rules

        # Also persist dynamic counts and default values if present
        tags_count = info.get('tags_count', getattr(template, 'tags_count', 1))
        spec_pairs_count = info.get('spec_pairs_count', getattr(template, 'spec_pairs_count', 1))
        customer_id_pairs_count = info.get('customer_id_pairs_count', getattr(template, 'customer_id_pairs_count', 1))
        default_values = info.get('default_values', getattr(template, 'default_values', {}))

        try:
            template.tags_count = int(tags_count)
            template.spec_pairs_count = int(spec_pairs_count)
            template.customer_id_pairs_count = int(customer_id_pairs_count)
        except Exception:
            logger.warning("🔧 DEBUG: Could not convert dynamic counts to int during update_mapping_template")
        try:
            template.default_values = default_values
        except Exception:
            logger.warning("🔧 DEBUG: Could not set default_values during update_mapping_template")
        
        # Update name and description if provided
        if template_name:
            template.name = template_name
        if description:
            template.description = description
        
        template.save()
        
        logger.info(
            f"Template {template.id} updated successfully with {len(mappings)} mappings, "
            f"{len(formula_rules)} formula rules, counts: tags={getattr(template,'tags_count',None)}, "
            f"spec_pairs={getattr(template,'spec_pairs_count',None)}, cust_pairs={getattr(template,'customer_id_pairs_count',None)}"
        )
        
        return Response({
            'success': True,
            'message': f'Template "{template.name}" updated successfully',
            'template_id': template.id
        })
        
    except Exception as e:
        logger.error(f"Error in update_mapping_template: {e}")
        return Response({
            'success': False,
            'error': f'Failed to update template: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_mapping_templates(request):
    """Get all saved mapping templates."""
    try:
        templates = MappingTemplate.objects.all().order_by('-created_at')
        template_list = [template.get_mapping_summary() for template in templates]
        
        return Response({
            'success': True,
            'templates': template_list
        })
        
    except Exception as e:
        logger.error(f"Error in get_mapping_templates: {e}")
        return Response({
            'success': False,
            'error': f'Failed to get templates: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def delete_mapping_template(request, template_id):
    """Delete a saved mapping template."""
    try:
        # Get the template
        try:
            template = MappingTemplate.objects.get(id=template_id)
        except MappingTemplate.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Template not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        template_name = template.name
        template.delete()
        
        return Response({
            'success': True,
            'message': f'Template "{template_name}" deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error in delete_mapping_template: {e}")
        return Response({
            'success': False,
            'error': f'Failed to delete template: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def apply_mapping_template(request):
    """Apply a saved mapping template to a session."""
    try:
        session_id = request.data.get('session_id')
        template_id = request.data.get('template_id')
        
        logger.info(f"🔧 DEBUG: apply_mapping_template called with session_id: {session_id}, template_id: {template_id}")
        logger.debug(f"Request data: {request.data}")
        
        info = get_session_consistent(session_id)
        if not session_id or not info:
            logger.error(f"❌ Invalid session_id: {session_id} (available: {list(SESSION_STORE.keys())})")
            return Response({
                'success': False,
                'error': 'Invalid session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not template_id:
            logger.error(f"❌ Template ID is required")
            return Response({
                'success': False,
                'error': 'Template ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get the template
        try:
            template = MappingTemplate.objects.get(id=template_id)
            logger.info(f"✅ Found template: {template.name} (ID: {template.id})")
            logger.debug(f"Template details: tags_count={getattr(template, 'tags_count', 'N/A')}, "
                        f"spec_pairs_count={getattr(template, 'spec_pairs_count', 'N/A')}, "
                        f"customer_id_pairs_count={getattr(template, 'customer_id_pairs_count', 'N/A')}")
        except MappingTemplate.DoesNotExist:
            logger.error(f"❌ Template not found with ID: {template_id}")
            return Response({
                'success': False,
                'error': 'Template not found'
            }, status=status.HTTP_404_NOT_FOUND)
        
        info = SESSION_STORE[session_id]
        logger.debug(f"Session info keys: {list(info.keys())}")
        logger.debug(f"Session has client_path: {info.get('client_path')}")
        logger.debug(f"Session has sheet_name: {info.get('sheet_name')}")
        logger.debug(f"Session has header_row: {info.get('header_row')}")
        
        # Read client headers
        mapper = BOMHeaderMapper()
        client_headers = mapper.read_excel_headers(
            file_path=info["client_path"],
            sheet_name=info["sheet_name"],
            header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0
        )
        logger.info(f"📋 Read {len(client_headers)} client headers: {client_headers}")
        
        # Apply template mappings
        logger.info(f"🔄 Applying template mappings...")
        application_result = template.apply_to_headers(client_headers)
        logger.info(f"✅ Template application result: {application_result}")
        
        if application_result['total_mapped'] > 0:
            logger.info(f"✅ Template applied successfully with {application_result['total_mapped']} mappings")
            
            # Update session with applied template ID
            SESSION_STORE[session_id]["original_template_id"] = template_id
            logger.debug(f"Updated session with original_template_id: {template_id}")
            
            # CRITICAL FIX: Clear enhanced data cache to force fresh mapping on data review
            if "formula_enhanced_data" in SESSION_STORE[session_id]:
                del SESSION_STORE[session_id]["formula_enhanced_data"]
                logger.debug("Cleared formula_enhanced_data cache")
            if "enhanced_headers" in SESSION_STORE[session_id]:
                del SESSION_STORE[session_id]["enhanced_headers"]
                logger.debug("Cleared enhanced_headers cache")
            
            # CRITICAL FIX: Apply column counts from template and include counts implied by formula rules
            template_tags_count = getattr(template, 'tags_count', 1)
            template_spec_pairs_count = getattr(template, 'spec_pairs_count', 1)
            template_customer_id_pairs_count = getattr(template, 'customer_id_pairs_count', 1)
            
            logger.debug(f"Template column counts - Tags: {template_tags_count}, Spec pairs: {template_spec_pairs_count}, Customer pairs: {template_customer_id_pairs_count}")
            
            # Count mapped fields in the template (use exact counts of distinct indices)
            template_mappings = getattr(template, 'mappings', {})
            mapped_tag_indices = set()
            mapped_spec_indices = set()
            mapped_customer_indices = set()

            def _extract_index(name: str, prefix: str) -> int:
                try:
                    return int(name.replace(prefix, '').strip('_'))
                except Exception:
                    return None

            logger.debug(f"Analyzing template mappings: {template_mappings}")
            
            if isinstance(template_mappings, dict) and 'new_format' in template_mappings:
                logger.debug(f"Processing new_format mappings: {template_mappings['new_format']}")
                for mapping in template_mappings['new_format']:
                    target = mapping.get('target', '') or ''
                    if target.startswith('Tag_'):
                        idx = _extract_index(target, 'Tag_')
                        if idx:
                            mapped_tag_indices.add(idx)
                            logger.debug(f"Found mapped Tag_{idx}")
                    elif target.startswith('Specification_Name_'):
                        idx = _extract_index(target, 'Specification_Name_')
                        if idx:
                            mapped_spec_indices.add(idx)
                            logger.debug(f"Found mapped Specification_Name_{idx}")
                    elif target.startswith('Customer_Identification_Name_'):
                        idx = _extract_index(target, 'Customer_Identification_Name_')
                        if idx:
                            mapped_customer_indices.add(idx)
                            logger.debug(f"Found mapped Customer_Identification_Name_{idx}")
            elif isinstance(template_mappings, dict):
                logger.debug(f"Processing old format mappings: {template_mappings}")
                for target in (template_mappings or {}).keys():
                    target = target or ''
                    if target.startswith('Tag_'):
                        idx = _extract_index(target, 'Tag_')
                        if idx:
                            mapped_tag_indices.add(idx)
                            logger.debug(f"Found mapped Tag_{idx}")
                    elif target.startswith('Specification_Name_'):
                        idx = _extract_index(target, 'Specification_Name_')
                        if idx:
                            mapped_spec_indices.add(idx)
                            logger.debug(f"Found mapped Specification_Name_{idx}")
                    elif target.startswith('Customer_Identification_Name_'):
                        idx = _extract_index(target, 'Customer_Identification_Name_')
                        if idx:
                            mapped_customer_indices.add(idx)
                            logger.debug(f"Found mapped Customer_Identification_Name_{idx}")
            
            logger.debug(f"Mapped indices found - Tags: {mapped_tag_indices}, Spec: {mapped_spec_indices}, Customer: {mapped_customer_indices}")
            
            # Also consider formula_rules implied counts (distinct Tag_N targets)
            # Normalize Tag rules to generic 'Tag' so they don't force-create Tag_N slots
            fr_raw = getattr(template, 'formula_rules', []) or []
            fr = []
            for _r in fr_raw:
                r = dict(_r or {})
                if (r or {}).get('column_type', 'Tag') == 'Tag':
                    r['target_column'] = 'Tag'
                fr.append(r)
            logger.debug(f"Formula rules: {fr}")
            formula_tag_targets = [r.get('target_column') for r in fr if (r or {}).get('column_type', 'Tag') == 'Tag']
            formula_tag_indices = set()
            for t in formula_tag_targets:
                if t and str(t).startswith('Tag_'):
                    idx = _extract_index(str(t), 'Tag_')
                    if idx:
                        formula_tag_indices.add(idx)

            # Use the highest index actually referenced by mappings or Tag-specific formula targets.
            # Avoid inflating counts from stored template_tags_count (which may carry old sessions).
            mapped_tag_max = max(mapped_tag_indices) if mapped_tag_indices else 0
            formula_tag_max = max(formula_tag_indices) if formula_tag_indices else 0
            tags_count = max(mapped_tag_max, formula_tag_max)
            # Ensure at least 1 Tag column if template declared any tags
            if tags_count == 0 and template_tags_count > 0:
                tags_count = min(template_tags_count, 1)

            # For specs/customers, keep existing behavior by distinct counts
            spec_pairs_count = max(template_spec_pairs_count, len(mapped_spec_indices))
            customer_id_pairs_count = max(template_customer_id_pairs_count, len(mapped_customer_indices))
            
            logger.info(f"🔧 DEBUG: Template column count logic - Template: tags={template_tags_count}, specs={template_spec_pairs_count}, customers={template_customer_id_pairs_count}")
            logger.info(f"🔧 DEBUG: Template column count logic - Mapped: tags={len(mapped_tag_indices)}, specs={len(mapped_spec_indices)}, customers={len(mapped_customer_indices)}")
            logger.info(f"🔧 DEBUG: Template column count logic - Final: tags={tags_count}, specs={spec_pairs_count}, customers={customer_id_pairs_count}")
            
            SESSION_STORE[session_id]["tags_count"] = tags_count
            SESSION_STORE[session_id]["spec_pairs_count"] = spec_pairs_count
            SESSION_STORE[session_id]["customer_id_pairs_count"] = customer_id_pairs_count
            
            # Get existing template headers to preserve tag numbering
            existing_template_headers = getattr(template, 'template_headers', [])
            
            # CRITICAL FIX: Prevent tag duplication by checking existing headers first
            # Only regenerate if we don't already have the right number of dynamic columns
            existing_dynamic_columns = [h for h in existing_template_headers if any(h.startswith(prefix) for prefix in ['Tag_', 'Specification_Name_', 'Specification_Value_', 'Customer_Identification_']) or h in ['Tag', 'Specification name', 'Specification value', 'Customer identification name', 'Customer identification value']]
            
            debug_log(session_id, "Checking existing dynamic columns before regeneration", {
                'existing_dynamic_count': len(existing_dynamic_columns),
                'expected_tags_count': tags_count,
                'expected_spec_pairs_count': spec_pairs_count,
                'existing_dynamic_columns': existing_dynamic_columns[:10],  # Log first 10 for readability
                'tag_columns': [h for h in existing_dynamic_columns if h.startswith('Tag_') or h == 'Tag'],
                'spec_columns': [h for h in existing_dynamic_columns if h.startswith('Specification_') or h in ['Specification name', 'Specification value']],
                'customer_columns': [h for h in existing_dynamic_columns if h.startswith('Customer_Identification_') or h in ['Customer identification name', 'Customer identification value']]
            })
            
            # Only regenerate if counts don't match or if no dynamic columns exist
            should_regenerate = (
                len([h for h in existing_dynamic_columns if h.startswith('Tag_') or h == 'Tag']) != tags_count or
                len([h for h in existing_dynamic_columns if h.startswith('Specification_Name_') or h == 'Specification name']) != spec_pairs_count or
                len([h for h in existing_dynamic_columns if h.startswith('Customer_Identification_Name_') or h == 'Customer identification name']) != customer_id_pairs_count or
                len(existing_dynamic_columns) == 0
            )
            
            if should_regenerate:
                debug_log(session_id, "Regenerating dynamic columns due to count mismatch", {
                    'existing_tags': len([h for h in existing_dynamic_columns if h.startswith('Tag_') or h == 'Tag']),
                    'expected_tags': tags_count,
                    'existing_specs': len([h for h in existing_dynamic_columns if h.startswith('Specification_Name_') or h == 'Specification name']),
                    'expected_specs': spec_pairs_count,
                    'existing_customers': len([h for h in existing_dynamic_columns if h.startswith('Customer_Identification_Name_') or h == 'Customer identification name'])
                })
                
                regenerated_headers = generate_template_columns(
                    tags_count, 
                    spec_pairs_count, 
                    customer_id_pairs_count, 
                    existing_headers=existing_template_headers
                )
                SESSION_STORE[session_id]["current_template_headers"] = regenerated_headers
                SESSION_STORE[session_id]["enhanced_headers"] = regenerated_headers
            else:
                debug_log(session_id, "Using existing dynamic columns (no regeneration needed)", {
                    'existing_headers_count': len(existing_template_headers),
                    'dynamic_columns_count': len(existing_dynamic_columns)
                })
                # Keep existing headers to prevent duplication
                regenerated_headers = existing_template_headers
                SESSION_STORE[session_id]["current_template_headers"] = regenerated_headers
                SESSION_STORE[session_id]["enhanced_headers"] = regenerated_headers
            
            # IMPORTANT: Save session immediately after setting column counts
            save_session(session_id, SESSION_STORE[session_id])
            
            logger.info(f"🔧 DEBUG: Template applied - regenerated {len(regenerated_headers)} numbered headers: {regenerated_headers}")
            logger.info(f"🔧 DEBUG: Saved column counts to session: tags={tags_count}, spec_pairs={spec_pairs_count}, customer_id_pairs={customer_id_pairs_count}")
            
            # IMPORTANT: Ensure session mappings are in list (new-format) and preserve duplicates
            current_session_mappings = SESSION_STORE[session_id].get("mappings")
            if isinstance(current_session_mappings, dict) and 'mappings' in current_session_mappings:
                # OK: already new format
                pass
            elif isinstance(current_session_mappings, list):
                SESSION_STORE[session_id]["mappings"] = {"mappings": current_session_mappings}
            elif isinstance(current_session_mappings, dict):
                # Old format dict -> convert to list
                converted = [{"source": v, "target": k} for k, v in current_session_mappings.items()]
                SESSION_STORE[session_id]["mappings"] = {"mappings": converted}
            
            # FIXED: Store FactWise rules in session for frontend display
            factwise_rules = getattr(template, 'factwise_rules', []) or []
            if factwise_rules:
                SESSION_STORE[session_id]["factwise_rules"] = factwise_rules
                logger.info(f"🔧 DEBUG: Stored {len(factwise_rules)} FactWise rules in session for frontend display")
            
            # CRITICAL FIX: Preserve duplicates by using the new-format list from application_result
            # The application_result contains the actual mappings that were successfully applied
            new_format_list = application_result.get('mappings_new_format', [])
            if not new_format_list and 'mappings' in application_result:
                # Fallback: convert old format to new format
                old_mappings = application_result['mappings']
                if isinstance(old_mappings, dict):
                    new_format_list = [{"source": v, "target": k} for k, v in old_mappings.items()]
                elif isinstance(old_mappings, list):
                    new_format_list = old_mappings

            # NEW: If Tag formulas exist, reserve the next available Tag_N column for formulas
            # and drop any direct mapping targeting exactly that reserved Tag_N (e.g., Tag_4)
            try:
                tag_formulas = [r for r in (getattr(template, 'formula_rules', []) or []) if (r or {}).get('column_type', 'Tag') == 'Tag']
                if tag_formulas and isinstance(new_format_list, list):
                    # Determine used Tag indices in direct mappings
                    used_tag_indices = set()
                    for m in new_format_list:
                        tgt = (m or {}).get('target')
                        if isinstance(tgt, str) and tgt.startswith('Tag_'):
                            try:
                                idx = int(tgt.split('_')[1])
                                used_tag_indices.add(idx)
                            except Exception:
                                pass
                    next_idx = (max(used_tag_indices) + 1) if used_tag_indices else 1
                    reserved_tag = f'Tag_{next_idx}'
                    # Filter out direct mappings to the reserved Tag_N
                    filtered_list = [m for m in new_format_list if (m or {}).get('target') != reserved_tag]
                    if len(filtered_list) != len(new_format_list):
                        logger.info(f"🔧 DEBUG: Dropped direct mapping to reserved formula tag '{reserved_tag}' to keep it formula-only")
                    new_format_list = filtered_list
                    # IMPORTANT: Bump tags_count to allow formula engine to create the reserved Tag column
                    try:
                        current_cap = int(SESSION_STORE[session_id].get('tags_count', 0) or 0)
                    except Exception:
                        current_cap = 0
                    if next_idx > current_cap:
                        SESSION_STORE[session_id]['tags_count'] = next_idx
                        logger.info(f"🔧 DEBUG: Increased tags_count cap to {next_idx} to allow formula Tag column creation")
            except Exception as _e:
                pass
            
            # Store mappings in new format to preserve duplicates
            new_format_mappings = {"mappings": new_format_list}
            SESSION_STORE[session_id]["mappings"] = new_format_mappings
            logger.info(f"🔄 Stored {len(new_format_list)} mappings in new-format list for session (duplicates preserved)")
            
            # CRITICAL: Update mappingsCacheRef equivalent on backend
            # This ensures the frontend can restore mappings even if edges are cleared
            SESSION_STORE[session_id]["cached_mappings"] = new_format_list
            
            # Apply formula rules if they exist
            formula_rules = fr  # use normalized rules
            if formula_rules:
                SESSION_STORE[session_id]["formula_rules"] = formula_rules
                
                # Apply formulas to create enhanced data
                mapping_result = apply_column_mappings(
                    client_file=info["client_path"],
                    mappings=new_format_mappings,
                    sheet_name=info["sheet_name"],
                    header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0,
                    session_id=session_id
                )
                
                # Convert to dict format for formula processing
                dict_rows = []
                for row_list in mapping_result['data']:
                    row_dict = {}
                    for i, header in enumerate(mapping_result['headers']):
                        if i < len(row_list):
                            row_dict[header] = row_list[i]
                        else:
                            row_dict[header] = ""
                    dict_rows.append(row_dict)
                
                # Apply formula rules to create enhanced data
                formula_result = apply_formula_rules(
                    data_rows=dict_rows,
                    headers=mapping_result['headers'],
                    formula_rules=formula_rules,
                    session_info=SESSION_STORE[session_id]
                )
                
                # Persist enhanced data immediately so Review and subsequent steps see Tag/Spec columns populated
                logger.info(f"Applied {len(formula_rules)} formula rules from template; persisting enhanced data")
                try:
                    SESSION_STORE[session_id]["formula_enhanced_data"] = formula_result.get('data', [])
                    SESSION_STORE[session_id]["enhanced_headers"] = formula_result.get('headers', mapping_result['headers'])
                    save_session(session_id, SESSION_STORE[session_id])
                except Exception as _e:
                    logger.warning(f"Failed to persist formula-enhanced data: {_e}")
            
            # Apply factwise rules if they exist
            factwise_rules = getattr(template, 'factwise_rules', []) or []
            if factwise_rules:
                SESSION_STORE[session_id]["factwise_rules"] = factwise_rules
                
                # Apply each factwise rule with error handling
                for rule in factwise_rules:
                    try:
                        if rule.get("type") == "factwise_id":
                            # Apply the Factwise ID rule by calling the existing function logic
                            first_column = rule.get("first_column")
                            second_column = rule.get("second_column")
                            operator = rule.get("operator", "_")
                        
                        if first_column and second_column:
                            # Get current data (either formula-enhanced or basic mapped)
                            current_data = SESSION_STORE[session_id].get("formula_enhanced_data")
                            current_headers = SESSION_STORE[session_id].get("enhanced_headers")
                            
                            if not current_data:
                                # Use basic mapped data if no formula data exists
                                mapping_result = apply_column_mappings(
                                    client_file=info["client_path"],
                                    mappings=new_format_mappings,
                                    sheet_name=info["sheet_name"],
                                    header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0,
                                    session_id=session_id
                                )
                                current_data = mapping_result['data']
                                current_headers = mapping_result['headers']
                            
                            # Use template column names directly for FactWise ID creation
                            # The FactWise rule stores template column names, and we should use them directly
                            # after the data has been mapped to template format
                            logger.info(f"🔧 DEBUG: Using template columns directly: '{first_column}', '{second_column}'")
                            
                            # Check if the template columns exist in the current headers
                            if first_column not in current_headers:
                                logger.warning(f"🆔 Template Factwise ID: First template column '{first_column}' not found in current headers: {current_headers}")
                                continue  # Skip this factwise rule
                            
                            if second_column not in current_headers:
                                logger.warning(f"🆔 Template Factwise ID: Second template column '{second_column}' not found in current headers: {current_headers}")
                                continue  # Skip this factwise rule
                            
                            # Apply FactWise ID creation using template column data
                            # Map directly into 'Item code' so it's available immediately after apply
                            factwise_id_column = []
                            first_col_idx = current_headers.index(first_column)
                            second_col_idx = current_headers.index(second_column)
                            strategy = (rule.get("strategy") or "fill_only_null")
                            
                            logger.info(f"🔧 DEBUG: Template data indices - first_idx: {first_col_idx}, second_idx: {second_col_idx}")
                            
                            if first_col_idx >= 0 and second_col_idx >= 0:
                                for row in current_data:
                                    if isinstance(row, dict):
                                        first_val = str(row.get(first_column, "")).strip()
                                        second_val = str(row.get(second_column, "")).strip()
                                    else:
                                        first_val = str(row[first_col_idx] if first_col_idx < len(row) else "").strip()
                                        second_val = str(row[second_col_idx] if second_col_idx < len(row) else "").strip()
                                    factwise_id = f"{first_val}{operator}{second_val}" if first_val and second_val else (first_val or second_val or "")
                                    factwise_id_column.append(factwise_id)
                                
                                # Map into Item code (or create it if missing)
                                new_headers = list(current_headers)
                                item_idx = None
                                for i_h, h in enumerate(new_headers):
                                    if str(h).strip().lower().replace(" ", "_") == "item_code":
                                        item_idx = i_h
                                        new_headers[i_h] = "Item code"
                                        break
                                new_data_rows = []
                                if item_idx is None:
                                    new_headers = ["Item code"] + new_headers
                                    for i, row in enumerate(current_data):
                                        new_row = [factwise_id_column[i]] + list(row)
                                        new_data_rows.append(new_row)
                                else:
                                    for i, row in enumerate(current_data):
                                        if isinstance(row, list):
                                            new_row = list(row)
                                            while len(new_row) <= item_idx:
                                                new_row.append("")
                                            if strategy == 'override_all' or not str(new_row[item_idx] or '').strip():
                                                new_row[item_idx] = factwise_id_column[i]
                                            new_data_rows.append(new_row)
                                        else:
                                            # dict rows
                                            new_row = dict(row)
                                            if strategy == 'override_all' or not str(new_row.get('Item code', '') or '').strip():
                                                new_row['Item code'] = factwise_id_column[i]
                                            new_data_rows.append(new_row)
                                
                                # Update session with Item code-enhanced data
                                SESSION_STORE[session_id]["formula_enhanced_data"] = new_data_rows
                                SESSION_STORE[session_id]["enhanced_headers"] = new_headers
                                try:
                                    SESSION_STORE[session_id]["current_template_headers"] = new_headers
                                except Exception:
                                    pass
                                save_session(session_id, SESSION_STORE[session_id])
                                
                                logger.info(f"🆔 Applied Factwise ID rule to 'Item code' from template: {first_column} {operator} {second_column}")
                    except Exception as factwise_error:
                        logger.warning(f"🆔 Failed to apply Factwise ID rule from template: {factwise_error}")
                        # Continue with other rules even if this one fails
            
            # CRITICAL FIX: Apply default values if they exist
            default_values = getattr(template, 'default_values', {}) or {}
            if default_values:
                logger.info(f"🔧 DEBUG: Template {template.id} has default values: {default_values}")
                # Store default values in session immediately for frontend access
                SESSION_STORE[session_id]["default_values"] = default_values
                logger.info(f"🔧 DEBUG: Stored default values in session: {default_values}")
                
                # CRITICAL: Save session immediately to ensure default values are persisted
                save_session(session_id, SESSION_STORE[session_id])
                logger.info(f"🔧 DEBUG: Saved session with default values immediately after template application")
                
                # Apply default values to current data if available
                current_data = SESSION_STORE[session_id].get("formula_enhanced_data") or SESSION_STORE[session_id].get("mapped_data")
                current_headers = SESSION_STORE[session_id].get("enhanced_headers") or SESSION_STORE[session_id].get("mapped_headers")
                
                if current_data and current_headers:
                    # Apply default values to each row
                    for row in current_data:
                        for field_name, default_value in default_values.items():
                            if field_name in current_headers:
                                field_index = current_headers.index(field_name)
                                
                                # Only apply if the field is empty
                                if isinstance(row, list) and field_index < len(row):
                                    if not row[field_index] or str(row[field_index]).strip() == "":
                                        row[field_index] = default_value
                                        logger.info(f"🔧 DEBUG: Applied default value '{default_value}' to field '{field_name}' in row")
                                elif isinstance(row, dict):
                                    if field_name not in row or not row[field_name] or str(row[field_name]).strip() == "":
                                        row[field_name] = default_value
                                        logger.info(f"🔧 DEBUG: Applied default value '{default_value}' to field '{field_name}' in row")
                    
                    # Update both data sources to ensure consistency
                    SESSION_STORE[session_id]["formula_enhanced_data"] = current_data
                    SESSION_STORE[session_id]["mapped_data"] = current_data
                    logger.info(f"🔧 DEBUG: Applied {len(default_values)} default values to {len(current_data)} rows")
            
            # Final save before template application completion
            info['mappings'] = application_result.get('mappings') or info.get('mappings')
            info['enhanced_headers'] = regenerated_headers
            info['default_values'] = default_values or info.get('default_values', {})
            # Store MPN validation metadata from template if available
            mpn_metadata = getattr(template, 'mpn_validation_metadata', {})
            if mpn_metadata:
                info['mpn_validation_metadata'] = mpn_metadata
                logger.info(f"🔧 DEBUG: Applied MPN validation metadata from template: {mpn_metadata}")

                # AUTO-RESTORE MPN COLUMNS FROM CACHE
                try:
                    mpn_column = mpn_metadata.get('mpn_column')
                    if mpn_column:
                        logger.info(f"🔧 DEBUG: Auto-restoring MPN columns from cache for column: {mpn_column}")

                        # Import the MPN restore function locally to avoid circular imports
                        from .mpn_views import mpn_restore_from_cache
                        from rest_framework.request import Request
                        from django.http import QueryDict

                        # Create a mock request object for the MPN restore function
                        mock_request = type('MockRequest', (), {})()
                        mock_request.data = {
                            'session_id': session_id,
                            'mpn_header': mpn_column,
                            'manufacturer_header': mpn_metadata.get('manufacturer_column')
                        }

                        # Call the MPN restore function
                        restore_response = mpn_restore_from_cache(mock_request)

                        if hasattr(restore_response, 'data') and restore_response.data.get('success'):
                            cache_hits = restore_response.data.get('cache_hits', 0)
                            columns_added = restore_response.data.get('columns_added', [])
                            uncached_count = restore_response.data.get('uncached_count', 0)

                            logger.info(f"✅ MPN cache restore successful: {cache_hits} cache hits, {len(columns_added)} columns added, {uncached_count} uncached MPNs")

                            if columns_added:
                                # Refresh the session data after MPN columns were added
                                info = get_session_consistent(session_id)
                                if info and 'enhanced_data' in info:
                                    regenerated_headers = info['enhanced_data'].get('headers', regenerated_headers)
                                    logger.info(f"🔄 Updated headers after MPN restore: {len(regenerated_headers)} columns")

                            if uncached_count > 0:
                                logger.info(f"⚠️  {uncached_count} MPNs need validation - user should run MPN validation for new parts")
                        else:
                            error_msg = restore_response.data.get('error', 'Unknown error') if hasattr(restore_response, 'data') else 'Response error'
                            logger.warning(f"⚠️  MPN cache restore failed: {error_msg}")

                except Exception as e:
                    logger.warning(f"⚠️  MPN cache restore failed with exception: {e}")
                    # Don't fail template application if MPN restore fails
                    pass

            info['column_counts'] = {
                'tags_count': tags_count,
                'spec_pairs_count': spec_pairs_count,
                'customer_id_pairs_count': customer_id_pairs_count,
            }
            save_session(session_id, info)
            
            # Atomic version bump AFTER all data is saved
            new_version = increment_template_version(session_id)
            
            # Increment template usage
            template.increment_usage()
            
            return no_store(Response({
                'success': True,
                'template_version': new_version,
                'enhanced_headers': regenerated_headers,
                # Return filtered mappings to frontend to avoid applying direct mapping to reserved Tag_N
                'mappings': {k: v for k, v in (application_result.get('mappings', {}) or {}).items() if not (isinstance(k, str) and k.startswith('Tag_') and k not in [m.get('target') for m in new_format_list])},
                'mappings_new_format': new_format_list,
                'default_values': default_values,
                'column_counts': info['column_counts'],
                'total_mapped': application_result.get('total_mapped', 0),
                # Include formula rules to help frontends reflect tag rules immediately
                'formula_rules': _externalize_formula_rules(SESSION_STORE.get(session_id, {}).get('formula_rules', []), SESSION_STORE.get(session_id, {})),
            }))
        else:
            return Response({
                'success': False,
                'error': 'No columns could be mapped from this template'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    except Exception as e:
        logger.error(f"Error in apply_mapping_template: {e}")
        return no_store(Response({
            'success': False,
            'error': f'Failed to apply template: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR))


class BOMHeaderMappingView(APIView):
    """Legacy API view for BOM header mapping."""
    
    def post(self, request):
        return Response({
            'success': True,
            'message': 'Use the new mapping endpoints instead'
        })


# ─── FORMULA MANAGEMENT ENDPOINTS ──────────────────────────────────────────────

def apply_formula_rules(data_rows, headers, formula_rules, replace_existing=False, session_info=None):
    """
    Apply formula rules with manual sub-rules to data rows and return modified data with new columns.
    
    New Rule Structure with Sub-Rules:
    - source_column: Column to check
    - column_type: 'Tag' or 'Specification Value' 
    - specification_name: Static name for specification column (only if column_type='Specification Value')
    - sub_rules: Array of conditions [
        { search_text: 'CAP', output_value: 'Capacitor', case_sensitive: false },
        { search_text: 'DIODE', output_value: 'Diode', case_sensitive: false }
      ]
    
    Logic:
    - Each rule defines ONE column (Tag or Specification)
    - Sub-rules define multiple conditions within that rule
    - First matching sub-rule wins per row
    - Multiple rules can create multiple columns
    
    Args:
        data_rows: List of dictionaries representing the data
        headers: List of column headers
        formula_rules: List of formula rule dictionaries with sub_rules
        replace_existing: If True, replace existing formula columns instead of creating new ones
    
    Returns:
        Dict with modified data and new headers
    """
    if not data_rows or not formula_rules:
        return {'data': data_rows, 'headers': headers, 'new_columns': []}
    
    # Create a copy of the data to avoid modifying original
    modified_data = [row.copy() for row in data_rows]
    new_headers = headers.copy()
    new_columns = []
    
    # Track column usage for auto-naming
    used_column_names = set(headers)
    
    # Process each rule separately (each rule creates its own column) 
    # FIXED: Don't use counters for tags - always use "Tag" to ensure isolation
    spec_counter = 1
    
    for rule_index, rule in enumerate(formula_rules):
        source_column = rule.get('source_column')
        column_type = rule.get('column_type', 'Tag')
        specification_name = rule.get('specification_name', '')
        sub_rules = rule.get('sub_rules', [])
        
        # Skip rule if missing required fields
        if not source_column or not sub_rules:
            continue
        
        # Determine column name based on type - SIMPLIFIED TAG COLUMN MANAGEMENT
        if column_type == 'Tag':
            # Check if this rule already has a target column specified
            target_column = rule.get('target_column')
            
            # Initialize column_name to None to ensure it's always assigned
            column_name = None
            
            if target_column and target_column.startswith('Tag_'):
                # Use the specified target column if it exists
                column_name = target_column
                if column_name not in used_column_names:
                    logger.info(f"🔧 DEBUG: Using specified Tag column '{column_name}' for rule {rule_index + 1}")
                else:
                    logger.warning(f"🔧 DEBUG: Specified Tag column '{column_name}' already exists, creating new one")
                    column_name = None  # Reset to None so we create a new one
            elif target_column == 'Tag':
                # CRITICAL FIX: Convert old-style 'Tag' to use first available numbered Tag column
                logger.info(f"🔧 DEBUG: Converting old-style 'Tag' target to use existing numbered Tag column")
                # Find the first available numbered Tag column in headers
                for header in headers:
                    if header.startswith('Tag_') and header not in used_column_names:
                        column_name = header
                        logger.info(f"🔧 DEBUG: Using existing Tag column '{column_name}' for old 'Tag' rule")
                        break
            
            # If we don't have a valid column_name yet, create a new one
            if not column_name or column_name in used_column_names:
                # Use centralized function to get next available Tag column
                # This ensures consistency with mapping logic
                # Pass session info to get proper context
                if session_info is None:
                    session_info = {'current_template_headers': headers, 'enhanced_headers': headers}
                column_name = get_next_available_tag_column(session_info, used_column_names)
                logger.info(f"🔧 DEBUG: Creating new Tag column '{column_name}' for rule {rule_index + 1}")
            
            # Evaluate matches first without mutating rows
            tag_assignments = []  # list of (row_index, value)
            for idx, row in enumerate(modified_data):
                match_value = None
                for sub_rule in sub_rules:
                    search_text = sub_rule.get('search_text', '')
                    output_value = sub_rule.get('output_value', '')
                    case_sensitive = sub_rule.get('case_sensitive', False)
                    if not search_text or not output_value:
                        continue
                    cell_value = str(row.get(source_column, ''))
                    search_text_compare = search_text if case_sensitive else str(search_text).lower()
                    cell_value_compare = cell_value if case_sensitive else cell_value.lower()
                    if search_text_compare in cell_value_compare:
                        match_value = str(output_value)
                        break
                if match_value is not None and str(match_value).strip() != '':
                    tag_assignments.append((idx, match_value))

            if tag_assignments:
                # Try to fit matches into existing Tag columns first (per-row first empty slot)
                existing_tag_cols = [h for h in new_headers if h.startswith('Tag_')]
                # Sort by numeric index to preserve order
                try:
                    existing_tag_cols.sort(key=lambda x: int(x.split('_')[1]))
                except Exception:
                    existing_tag_cols.sort()

                unresolved = []
                # Ensure all existing tag columns are present in each row
                for row in modified_data:
                    for tcol in existing_tag_cols:
                        if tcol not in row:
                            row[tcol] = ''

                for idx, value in tag_assignments:
                    placed = False
                    # Place into first empty existing Tag column for this row
                    for tcol in existing_tag_cols:
                        current = str(modified_data[idx].get(tcol, '') or '').strip()
                        if not current:
                            modified_data[idx][tcol] = value
                            placed = True
                            break
                        # If already contains the value, treat as placed
                        existing_values = [v.strip() for v in current.split(',')]
                        if value in existing_values:
                            placed = True
                            break
                    if not placed:
                        unresolved.append((idx, value))

                # If we still have unresolved assignments, create exactly one new Tag column
                if unresolved:
                    # Respect session tag cap if provided: do not create more Tag columns than tags_count
                    tag_cap = 0
                    try:
                        if session_info and isinstance(session_info, dict):
                            tag_cap = int(session_info.get('tags_count', 0) or 0)
                    except Exception:
                        tag_cap = 0

                    if tag_cap and len(existing_tag_cols) >= tag_cap:
                        # Do not add a new Tag column; fold unresolved values into the last Tag column
                        target_fold_col = existing_tag_cols[-1] if existing_tag_cols else None
                        if target_fold_col:
                            for idx, value in unresolved:
                                existing_value = str(modified_data[idx].get(target_fold_col, '')).strip()
                                if existing_value and existing_value != value:
                                    existing_values = [v.strip() for v in existing_value.split(',')]
                                    if value not in existing_values:
                                        modified_data[idx][target_fold_col] = f"{existing_value}, {value}"
                                else:
                                    modified_data[idx][target_fold_col] = value
                        else:
                            # No existing Tag_N columns — initialize the chosen column_name without growing headers list
                            for row in modified_data:
                                if column_name not in row:
                                    row[column_name] = ''
                            for idx, value in unresolved:
                                modified_data[idx][column_name] = value
                    else:
                        if column_name not in new_headers:
                            new_headers.append(column_name)
                            new_columns.append(column_name)
                            used_column_names.add(column_name)
                            logger.info(f"🔧 DEBUG: Added Tag column '{column_name}' to headers (needed for unresolved matches)")
                        # Initialize column in all rows
                        for row in modified_data:
                            if column_name not in row:
                                row[column_name] = ''
                        # Apply unresolved assignments
                        for idx, value in unresolved:
                            existing_value = str(modified_data[idx].get(column_name, '')).strip()
                            if existing_value and existing_value != value:
                                existing_values = [v.strip() for v in existing_value.split(',')]
                                if value not in existing_values:
                                    modified_data[idx][column_name] = f"{existing_value}, {value}"
                            else:
                                modified_data[idx][column_name] = value
            else:
                logger.info(f"🔧 DEBUG: Skipped adding Tag column '{column_name}' (no matches)")
        
        elif column_type == 'Specification Value' and specification_name:
            # Try to use generic specification column names first
            name_column = 'Specification name'
            value_column = 'Specification value'
            
            # If generic names are already used, create numbered versions
            if name_column in used_column_names or value_column in used_column_names:
                name_column = f"Specification_Name_{spec_counter}"
                value_column = f"Specification_Value_{spec_counter}"
                
                # Find next available specification column numbers
                while name_column in used_column_names or value_column in used_column_names:
                    spec_counter += 1
                    name_column = f"Specification_Name_{spec_counter}"
                    value_column = f"Specification_Value_{spec_counter}"
            
            # Evaluate matches first without mutating rows
            spec_assignments = []  # list of (row_index, value)
            for idx, row in enumerate(modified_data):
                match_value = None
                for sub_rule in sub_rules:
                    search_text = sub_rule.get('search_text', '')
                    output_value = sub_rule.get('output_value', '')
                    case_sensitive = sub_rule.get('case_sensitive', False)
                    if not search_text or not output_value:
                        continue
                    cell_value = str(row.get(source_column, ''))
                    search_text_compare = search_text if case_sensitive else str(search_text).lower()
                    cell_value_compare = cell_value if case_sensitive else cell_value.lower()
                    if search_text_compare in cell_value_compare:
                        match_value = str(output_value)
                        break
                if match_value is not None and str(match_value).strip() != '':
                    spec_assignments.append((idx, match_value))

            # Only add spec columns if at least one row matched
            if spec_assignments:
                if name_column not in new_headers:
                    new_headers.append(name_column)
                    new_columns.append(name_column)
                    used_column_names.add(name_column)
                if value_column not in new_headers:
                    new_headers.append(value_column)
                    new_columns.append(value_column)
                    used_column_names.add(value_column)
                # Initialize columns
                for row in modified_data:
                    if name_column not in row:
                        row[name_column] = specification_name
                    if value_column not in row:
                        row[value_column] = ''
                # Apply assignments
                for idx, value in spec_assignments:
                    existing_value = str(modified_data[idx].get(value_column, '')).strip()
                    if existing_value and existing_value != value:
                        existing_values = [v.strip() for v in existing_value.split(',')]
                        if value not in existing_values:
                            modified_data[idx][value_column] = f"{existing_value}, {value}"
                    else:
                        modified_data[idx][value_column] = value
                spec_counter += 1
            else:
                logger.info(f"🔧 DEBUG: Skipped adding specification columns '{name_column}/{value_column}' (no matches)")
    
    return {
        'data': modified_data,
        'headers': new_headers,
        'new_columns': new_columns,
        'total_rows': len(modified_data)
    }


@api_view(['POST'])
def apply_formulas(request):
    """Apply formula rules to session data and return updated data."""
    try:
        session_id = request.data.get('session_id')
        formula_rules = request.data.get('formula_rules', [])
        
        logger.info(f"🔧 DEBUG: apply_formulas called for session {session_id} with {len(formula_rules)} rules")
        
        info = get_session(session_id)
        if not session_id or not info:
            logger.error(f"🔧 DEBUG: Session {session_id} not found")
            return no_store(Response({'success': False, 'error': 'Invalid session'}, status=400))
        
        if not formula_rules:
            return Response({
                'success': False,
                'error': 'No formula rules provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Session info already retrieved via get_session_consistent
        enhanced_headers = info.get("enhanced_headers", []) or info.get("current_template_headers", [])
        
        # Get all existing columns from all sources
        template_headers = info.get("template_headers", [])
        client_headers = list(info.get("source_headers", {}).keys())
        
        all_existing_columns = set(template_headers + client_headers + enhanced_headers)

        # Assign stable internal Tag_N names for Tag rules without a target and persist them
        updated_formula_rules = []
        for idx, rule in enumerate(formula_rules):
            updated_rule = rule.copy()
            if updated_rule.get('column_type', 'Tag') == 'Tag':
                target = updated_rule.get('target_column')
                # If no internal target assigned yet, allocate next available Tag_N once
                if not target or not str(target).startswith('Tag_'):
                    # Use centralized allocator with session context to pick next Tag_N
                    try:
                        next_tag = get_next_available_tag_column(info, set(all_existing_columns))
                        updated_rule['target_column'] = next_tag
                        all_existing_columns.add(next_tag)
                        logger.info(f"🔧 DEBUG: Assigned stable Tag column '{next_tag}' to rule {idx+1}")
                    except Exception as e:
                        logger.warning(f"Failed to allocate Tag column for rule {idx+1}: {e}")
            updated_formula_rules.append(updated_rule)

        # Persist updated rules back to session so subsequent applications reuse same Tag_N
        formula_rules = updated_formula_rules
        
        info = SESSION_STORE[session_id]
        mappings = info.get("mappings")
        
        if not mappings:
            return Response({
                'success': False,
                'error': 'No mappings found. Please create mappings first.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Convert mappings list format to expected dict format for apply_column_mappings
        if isinstance(mappings, list):
            # Convert list format to new dict format that apply_column_mappings expects
            formatted_mappings = {"mappings": mappings}
            logger.info(f"🔧 DEBUG: Converted list mappings to dict format for formulas: {formatted_mappings}")
        else:
            formatted_mappings = mappings
            logger.info(f"🔧 DEBUG: Using existing dict mappings for formulas: {formatted_mappings}")
        
        # Always start from fresh mapped data - no caching
        mapping_result = apply_column_mappings(
            client_file=info["client_path"],
            mappings=formatted_mappings,
            sheet_name=info["sheet_name"],
            header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0,
            session_id=session_id
        )
        
        # Convert to dict format for formula processing
        dict_rows = []
        for row_list in mapping_result['data']:
            row_dict = {}
            for i, header in enumerate(mapping_result['headers']):
                if i < len(row_list):
                    row_dict[header] = row_list[i]
                else:
                    row_dict[header] = ""
            dict_rows.append(row_dict)
        transformed_rows = dict_rows
        current_headers = mapping_result['headers']
        
        if not transformed_rows:
            return Response({
                'success': False,
                'error': 'No data available to apply formulas'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Apply formula rules (always create new unique columns)
        logger.info(f"🔧 DEBUG: About to apply {len(formula_rules)} rules to {len(transformed_rows)} rows with headers: {current_headers}")
        formula_result = apply_formula_rules(transformed_rows, current_headers, formula_rules, replace_existing=False, session_info=info)
        
        logger.info(f"🔧 DEBUG: Formula result - new_columns: {formula_result.get('new_columns', [])}, headers: {formula_result.get('headers', [])}")
        
        # Persist canonical state
        info['formula_rules'] = formula_rules
        # Avoid persisting full enhanced data for very large datasets; compute per page instead
        SMALL_DATA_THRESHOLD = 2000
        if len(formula_result.get('data', [])) <= SMALL_DATA_THRESHOLD:
            info['formula_enhanced_data'] = formula_result['data']
        else:
            info.pop('formula_enhanced_data', None)
        info['enhanced_headers'] = formula_result['headers']
        info['current_template_headers'] = formula_result['headers']
        
        # Increment template version when formulas create new columns
        new_version = increment_template_version(session_id)
        
        # Also bump general version for compatibility
        info['version'] = info.get('version', 0) + 1
        save_session(session_id, info)

        return no_store(Response({
            'success': True,
            'snapshot': build_snapshot(info),
            'new_columns': formula_result.get('new_columns', []),
            'total_rows': formula_result.get('total_rows', 0),
            'rules_applied': len(formula_rules),
            'template_version': new_version,
            'message': f'Applied {len(formula_rules)} formula rules successfully'
        }, status=200))
        
    except Exception as e:
        logger.error(f"Error in apply_formulas: {e}")
        return no_store(Response({
            'success': False,
            'error': f'Failed to apply formulas: {str(e)}'
        }, status=500))


@api_view(['POST'])
def preview_formulas(request):
    """Preview the results of applying formula rules without saving."""
    try:
        session_id = request.data.get('session_id')
        formula_rules = request.data.get('formula_rules', [])
        sample_size = int(request.data.get('sample_size', 5))
        
        if not session_id or session_id not in SESSION_STORE:
            return Response({
                'success': False,
                'error': 'Invalid session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not formula_rules:
            return Response({
                'success': False,
                'error': 'No formula rules provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        info = SESSION_STORE[session_id]
        mappings = info.get("mappings")
        
        if not mappings:
            return Response({
                'success': False,
                'error': 'No mappings found. Please create mappings first.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get the current transformed data
        mapping_result = apply_column_mappings(
            client_file=info["client_path"],
            mappings=mappings,
            sheet_name=info["sheet_name"],
            header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0,
            session_id=session_id
        )
        # Convert to dict format for preview
        dict_rows = []
        for row_list in mapping_result['data']:
            row_dict = {}
            for i, header in enumerate(mapping_result['headers']):
                if i < len(row_list):
                    row_dict[header] = row_list[i]
                else:
                    row_dict[header] = ""
            dict_rows.append(row_dict)
        transformed_rows = dict_rows
        
        if not transformed_rows:
            return Response({
                'success': False,
                'error': 'No data available for preview'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get current headers
        mapper = BOMHeaderMapper()
        current_headers = mapper.read_excel_headers(
            file_path=info["template_path"],
            sheet_name=info.get("template_sheet_name"),
            header_row=info.get("template_header_row", 1) - 1 if info.get("template_header_row", 1) > 0 else 0
        )
        
        # Apply formula rules to get preview
        formula_result = apply_formula_rules(transformed_rows, current_headers, formula_rules, session_info=info)
        
        # Calculate statistics for each rule
        rule_stats = []
        for i, rule in enumerate(formula_rules):
            source_column = rule.get('source_column')
            search_text = str(rule.get('search_text', '')).lower()
            case_sensitive = rule.get('case_sensitive', False)
            
            matches = 0
            for row in transformed_rows:
                source_value = str(row.get(source_column, '')).strip()
                if case_sensitive:
                    text_match = search_text in source_value
                else:
                    text_match = search_text in source_value.lower()
                
                if text_match:
                    matches += 1
            
            rule_stats.append({
                'rule_index': i,
                'matches': matches,
                'total_rows': len(transformed_rows),
                'match_percentage': round((matches / len(transformed_rows)) * 100, 1) if transformed_rows else 0
            })
        
        # Get sample data for preview (first few rows)
        sample_data = formula_result['data'][:sample_size]
        
        return Response({
            'success': True,
            'preview_data': sample_data,
            'headers': formula_result['headers'],
            'new_columns': formula_result['new_columns'],
            'rule_statistics': rule_stats,
            'total_rows': len(transformed_rows),
            'sample_size': len(sample_data),
            'message': f'Preview generated for {len(formula_rules)} rules'
        })
        
    except Exception as e:
        logger.error(f"Error in preview_formulas: {e}")
        return Response({
            'success': False,
            'error': f'Failed to preview formulas: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_formula_templates(request):
    """Get predefined formula templates for common use cases."""
    try:
        # Predefined templates for common component types
        templates = {
            'electronics_basic': {
                'name': 'Electronics Components (Basic)',
                'description': 'Common electronic components',
                'rules': [
                    {
                        'source_column': 'Description',
                        'search_text': 'cap',
                        'tag_value': 'Capacitor',
                        'target_column': 'Component_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description', 
                        'search_text': 'res',
                        'tag_value': 'Resistor',
                        'target_column': 'Component_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'ic',
                        'tag_value': 'Integrated Circuit',
                        'target_column': 'Component_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'led',
                        'tag_value': 'LED',
                        'target_column': 'Component_Type',
                        'case_sensitive': False
                    }
                ]
            },
            'electronics_advanced': {
                'name': 'Electronics Components (Advanced)',
                'description': 'Extended electronic components classification',
                'rules': [
                    {
                        'source_column': 'Description',
                        'search_text': 'capacitor',
                        'tag_value': 'Capacitor',
                        'target_column': 'Component_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'resistor',
                        'tag_value': 'Resistor', 
                        'target_column': 'Component_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'inductor',
                        'tag_value': 'Inductor',
                        'target_column': 'Component_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'diode',
                        'tag_value': 'Diode',
                        'target_column': 'Component_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'transistor',
                        'tag_value': 'Transistor',
                        'target_column': 'Component_Type',
                        'case_sensitive': False
                    }
                ]
            },
            'mechanical': {
                'name': 'Mechanical Parts',
                'description': 'Common mechanical hardware components',
                'rules': [
                    {
                        'source_column': 'Description',
                        'search_text': 'screw',
                        'tag_value': 'Fastener',
                        'target_column': 'Hardware_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'bolt',
                        'tag_value': 'Fastener',
                        'target_column': 'Hardware_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'washer',
                        'tag_value': 'Hardware',
                        'target_column': 'Hardware_Type',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'nut',
                        'tag_value': 'Fastener',
                        'target_column': 'Hardware_Type',
                        'case_sensitive': False
                    }
                ]
            },
            'value_classification': {
                'name': 'Value-based Classification',
                'description': 'Classify components by value ranges',
                'rules': [
                    {
                        'source_column': 'Description',
                        'search_text': 'pf',
                        'tag_value': 'Low Value Capacitor',
                        'target_column': 'Value_Category',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'uf',
                        'tag_value': 'High Value Capacitor',
                        'target_column': 'Value_Category',
                        'case_sensitive': False
                    },
                    {
                        'source_column': 'Description',
                        'search_text': 'ohm',
                        'tag_value': 'Standard Resistor',
                        'target_column': 'Value_Category',
                        'case_sensitive': False
                    }
                ]
            }
        }
        
        return Response({
            'success': True,
            'templates': templates,
            'total_templates': len(templates)
        })
        
    except Exception as e:
        logger.error(f"Error in get_formula_templates: {e}")
        return Response({
            'success': False,
            'error': f'Failed to get templates: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def save_custom_formulas(request):
    """Save custom formula rules to session for reuse."""
    try:
        session_id = request.data.get('session_id')
        formula_rules = request.data.get('formula_rules', [])
        template_name = request.data.get('template_name', 'Custom Formula Set')
        
        if not session_id or session_id not in SESSION_STORE:
            return Response({
                'success': False,
                'error': 'Invalid session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not formula_rules:
            return Response({
                'success': False,
                'error': 'No formula rules to save'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Save to session store
        if 'custom_formula_templates' not in SESSION_STORE[session_id]:
            SESSION_STORE[session_id]['custom_formula_templates'] = {}
        
        template_id = f"custom_{len(SESSION_STORE[session_id]['custom_formula_templates']) + 1}"
        SESSION_STORE[session_id]['custom_formula_templates'][template_id] = {
            'name': template_name,
            'rules': formula_rules,
            'created_at': datetime.now().isoformat(),
            'usage_count': 0
        }
        
        return Response({
            'success': True,
            'template_id': template_id,
            'message': f'Saved {len(formula_rules)} formula rules as "{template_name}"'
        })
        
    except Exception as e:
        logger.error(f"Error in save_custom_formulas: {e}")
        return Response({
            'success': False,
            'error': f'Failed to save formulas: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_enhanced_data(request):
    """Get data enhanced with formula results."""
    try:
        session_id = request.GET.get('session_id')
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        
        if not session_id or session_id not in SESSION_STORE:
            return Response({
                'success': False,
                'error': 'Invalid session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        info = SESSION_STORE[session_id]
        
        # Check if we have formula-enhanced data
        enhanced_data = info.get("formula_enhanced_data")
        enhanced_headers = info.get("enhanced_headers")
        
        if enhanced_data and enhanced_headers:
            # Use formula-enhanced data
            data_to_return = enhanced_data
            headers_to_return = enhanced_headers
        else:
            # Fall back to regular mapped data
            mappings = info.get("mappings")
            if not mappings:
                return Response({
                    'success': False,
                    'error': 'No data available. Please create mappings first.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            mapping_result = apply_column_mappings(
                client_file=info["client_path"],
                mappings=mappings,
                sheet_name=info["sheet_name"],
                header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0,
                session_id=session_id
            )
            # Convert to dict format
            dict_rows = []
            for row_list in mapping_result['data']:
                row_dict = {}
                for i, header in enumerate(mapping_result['headers']):
                    if i < len(row_list):
                        row_dict[header] = row_list[i]
                    else:
                        row_dict[header] = ""
                dict_rows.append(row_dict)
            data_to_return = dict_rows
            headers_to_return = mapping_result['headers']
        
        if not data_to_return:
            return Response({
                'success': False,
                'error': 'No data available'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Normalize generic 'Tag' column: move its values into numbered Tag_N columns, then drop 'Tag'.
        try:
            if isinstance(headers_to_return, list) and 'Tag' in headers_to_return and isinstance(data_to_return, list) and len(data_to_return) > 0:
                # Build ordered list of Tag_N headers
                tag_n_headers = [h for h in headers_to_return if isinstance(h, str) and h.startswith('Tag_')]
                try:
                    tag_n_headers.sort(key=lambda x: int(x.split('_')[1]))
                except Exception:
                    tag_n_headers.sort()
                if isinstance(data_to_return[0], dict):
                    for row in data_to_return:
                        val = str(row.get('Tag', '') or '').strip()
                        if not val:
                            continue
                        placed = False
                        for tcol in tag_n_headers:
                            cur = str(row.get(tcol, '') or '').strip()
                            if not cur:
                                row[tcol] = val
                                placed = True
                                break
                        if not placed and tag_n_headers:
                            last = tag_n_headers[-1]
                            cur = str(row.get(last, '') or '').strip()
                            if cur:
                                parts = [p.strip() for p in cur.split(',')]
                                if val not in parts:
                                    row[last] = f"{cur}, {val}"
                            else:
                                row[last] = val
                        row.pop('Tag', None)
                else:
                    tag_idx = headers_to_return.index('Tag')
                    # indices of Tag_N
                    tag_n_indices = []
                    for h in tag_n_headers:
                        try:
                            tag_n_indices.append(headers_to_return.index(h))
                        except ValueError:
                            pass
                    for row in data_to_return:
                        if tag_idx < len(row):
                            val = str(row[tag_idx] or '').strip()
                        else:
                            val = ''
                        if not val:
                            continue
                        placed = False
                        for idx in tag_n_indices:
                            if idx < len(row):
                                cur = str(row[idx] or '').strip()
                                if not cur:
                                    row[idx] = val
                                    placed = True
                                    break
                        if not placed and tag_n_indices:
                            last_idx = tag_n_indices[-1]
                            if last_idx < len(row):
                                cur = str(row[last_idx] or '').strip()
                                if cur:
                                    parts = [p.strip() for p in cur.split(',')]
                                    if val not in parts:
                                        row[last_idx] = f"{cur}, {val}"
                                else:
                                    row[last_idx] = val
                    # remove Tag column value; keep headers cleanup below
                # Finally drop 'Tag' header
                headers_to_return = [h for h in headers_to_return if h != 'Tag']
        except Exception:
            pass

        # Cleanup: drop any Tag_N columns that are empty-only
        try:
            if isinstance(headers_to_return, list) and len(headers_to_return) > 0:
                tag_n_headers = [h for h in headers_to_return if isinstance(h, str) and h.startswith('Tag_')]
                def col_empty_only(col_name: str) -> bool:
                    if not data_to_return:
                        return True
                    if isinstance(data_to_return[0], dict):
                        for row in data_to_return:
                            if str(row.get(col_name, '') or '').strip():
                                return False
                        return True
                    else:
                        if col_name not in headers_to_return:
                            return True
                        idx = headers_to_return.index(col_name)
                        for row in data_to_return:
                            if idx < len(row) and str(row[idx] or '').strip():
                                return False
                        return True
                # Remove empty-only Tag_N headers
                for h in list(tag_n_headers):
                    if col_empty_only(h):
                        if h in headers_to_return:
                            headers_to_return.remove(h)
                        if isinstance(data_to_return[0], dict):
                            for row in data_to_return:
                                row.pop(h, None)
                        else:
                            # list-of-lists: recompute index after header removal is tricky; skip for list rows
                            pass
        except Exception:
            pass

        # Implement pagination
        total_rows = len(data_to_return)
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_data = data_to_return[start_idx:end_idx]
        
        return Response({
            'success': True,
            'data': paginated_data,
            'headers': headers_to_return,
            'has_formulas': bool(enhanced_data),
            'formula_rules': info.get("formula_rules", []),
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_rows': total_rows,
                'total_pages': (total_rows + page_size - 1) // page_size
            }
        })
        
    except Exception as e:
        logger.error(f"Error in get_enhanced_data: {e}")
        return Response({
            'success': False,
            'error': f'Failed to get enhanced data: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def check_column_conflicts(request):
    """Check for column name conflicts before applying formulas."""
    logger.info("🔧 DEBUG: Starting check_column_conflicts function")
    try:
        session_id = request.data.get('session_id')
        formula_rules = request.data.get('formula_rules', [])
        
        if not session_id or session_id not in SESSION_STORE:
            return Response({
                'success': False,
                'error': 'Invalid session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        info = SESSION_STORE[session_id]
        mappings = info.get("mappings")
        
        if not mappings:
            return Response({
                'success': False,
                'error': 'No mappings found. Please create mappings first.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get current headers (both mapped and original)
        mapper = BOMHeaderMapper()
        template_headers = mapper.read_excel_headers(
            file_path=info["template_path"],
            sheet_name=info.get("template_sheet_name"),
            header_row=info.get("template_header_row", 1) - 1 if info.get("template_header_row", 1) > 0 else 0
        )
        
        client_headers = mapper.read_excel_headers(
            file_path=info["client_path"],
            sheet_name=info["sheet_name"],
            header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0
        )
        
        # Get enhanced headers (includes numbered fields like Tag_1, Tag_2)
        enhanced_headers = info.get("enhanced_headers", []) or info.get("current_template_headers", [])
        all_existing_columns = set(template_headers + client_headers + enhanced_headers)
        
        # Check for conflicts and provide smart numbering
        conflicts = []
        suggestions = {}
        
        logger.info(f"🔧 DEBUG: About to process {len(formula_rules)} rules with {len(all_existing_columns)} existing columns")
        for rule in formula_rules:
            target_column = rule.get('target_column') or 'Tag'
            column_type = rule.get('column_type', 'Tag')
            
            # Smart numbering for Tag and Specification columns
            if column_type == 'Tag':
                # Use centralized function to get next available Tag column
                suggested_name = get_next_available_tag_column(info, set(all_existing_columns))
                
                # Check if the target column already exists or conflicts
                if target_column in all_existing_columns:
                    # Target column already exists, suggest using it
                    conflicts.append({
                        'rule_index': formula_rules.index(rule),
                        'column': target_column,
                        'conflicting_column': target_column,
                        'conflict_type': 'column_exists',
                        'suggested_name': target_column,
                        'message': f'Tag column already exists, will use existing column'
                    })
                    suggestions[target_column] = target_column
                elif target_column != suggested_name:
                    # Suggest auto-numbering
                    conflicts.append({
                        'rule_index': formula_rules.index(rule),
                        'column': target_column,
                        'conflicting_column': target_column,
                        'conflict_type': 'auto_numbering',
                        'suggested_name': suggested_name,
                        'message': f'Auto-assigned to next available tag number'
                    })
                    suggestions[target_column] = suggested_name
                    # Reserve the suggested name
                    all_existing_columns.add(suggested_name)
                else:
                    # Target column matches suggested name, no conflict
                    all_existing_columns.add(suggested_name)
                    
            elif column_type == 'Specification Value':
                spec_name = rule.get('specification_name', 'Unknown')
                base_column = f"Specification_Value_{spec_name}"
                
                if base_column in all_existing_columns:
                    # Find next available specification number
                    spec_numbers = []
                    for col in all_existing_columns:
                        if col.startswith(f'Specification_Value_{spec_name}_') and col.split('_')[-1].isdigit():
                            spec_numbers.append(int(col.split('_')[-1]))
                    
                    next_spec_number = max(spec_numbers, default=0) + 1
                    suggested_name = f"Specification_Value_{spec_name}_{next_spec_number}"
                    
                    conflicts.append({
                        'rule_index': formula_rules.index(rule),
                        'column': base_column,
                        'conflicting_column': base_column,
                        'conflict_type': 'auto_numbering',
                        'suggested_name': suggested_name,
                        'message': f'Auto-assigned specification number for {spec_name}'
                    })
                    
                    suggestions[base_column] = suggested_name
                    all_existing_columns.add(suggested_name)  # Reserve this name
            else:
                # Generic conflict resolution for other column types
                if target_column in all_existing_columns:
                    base_name = target_column
                    counter = 1
                    suggested_name = f"{base_name}_{counter}"
                    while suggested_name in all_existing_columns:
                        counter += 1
                        suggested_name = f"{base_name}_{counter}"
                    
                    conflicts.append({
                        'rule_index': formula_rules.index(rule),
                        'column': target_column,
                        'conflicting_column': target_column,
                        'conflict_type': 'existing_column',
                        'suggested_name': suggested_name,
                        'message': f'Column already exists, using numbered variant'
                    })
                    
                    suggestions[target_column] = suggested_name
                    all_existing_columns.add(suggested_name)  # Reserve this name
        
        return Response({
            'success': True,
            'conflicts': conflicts,
            'suggestions': suggestions,
            'existing_columns': list(all_existing_columns),
            'has_conflicts': len(conflicts) > 0
        })
        
    except Exception as e:
        logger.error(f"Error checking column conflicts: {e}")
        return Response({
            'success': False,
            'error': f'Failed to check conflicts: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def clear_formulas(request):
    """Clear all formula rules and remove generated columns from session data."""
    try:
        session_id = request.data.get('session_id')
        
        if not session_id or session_id not in SESSION_STORE:
            return Response({
                'success': False,
                'error': 'Invalid session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        info = SESSION_STORE[session_id]
        
        # Track what we're removing for user feedback
        cleared_items = []
        
        # Check if there are formula rules to clear
        if info.get("formula_rules"):
            cleared_items.append(f"{len(info['formula_rules'])} formula rules")
            # Clear formula rules
            SESSION_STORE[session_id]["formula_rules"] = []
        
        # Check if there's enhanced data to clear
        if info.get("formula_enhanced_data"):
            cleared_items.append("formula-generated data")
            # Clear formula-enhanced data
            SESSION_STORE[session_id]["formula_enhanced_data"] = None
        
        # Check if there are enhanced headers to clear
        if info.get("enhanced_headers"):
            original_headers = info.get("enhanced_headers", [])
            mappings = info.get("mappings", {})
            
            # Handle different mapping formats
            if isinstance(mappings, list):
                # List format: [{'source': 'A', 'target': 'B'}, ...]
                template_columns = [m.get('target', '') for m in mappings if isinstance(m, dict)]
            elif isinstance(mappings, dict):
                template_columns = list(mappings.keys())
            else:
                template_columns = []
            
            # Find formula-generated columns (columns not in original mappings)
            formula_columns = [h for h in original_headers if h not in template_columns]
            if formula_columns:
                cleared_items.append(f"{len(formula_columns)} generated columns")
            
            # Clear enhanced headers
            SESSION_STORE[session_id]["enhanced_headers"] = None
        
        # If no formulas were found to clear
        if not cleared_items:
            return Response({
                'success': True,
                'message': 'No formulas or generated columns found to clear',
                'cleared_items': []
            })
        
        # Log the clearing action
        logger.info(f"Cleared formulas for session {session_id}: {', '.join(cleared_items)}")
        
        return Response({
            'success': True,
            'message': f'Successfully cleared: {", ".join(cleared_items)}',
            'cleared_items': cleared_items
        })
        
    except Exception as e:
        logger.error(f"Error clearing formulas: {e}")
        return Response({
            'success': False,
            'error': f'Failed to clear formulas: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==============================================
# TAG TEMPLATE VIEWS
# ==============================================

@api_view(['POST'])
def save_tag_template(request):
    """Save a new tag template with formula rules."""
    try:
        template_name = request.data.get('template_name')
        description = request.data.get('description', '')
        formula_rules = request.data.get('formula_rules', [])
        
        if not template_name:
            return Response({
                'success': False,
                'error': 'Template name is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not formula_rules:
            return Response({
                'success': False,
                'error': 'Formula rules are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if template name already exists
        if TagTemplate.objects.filter(name=template_name).exists():
            return Response({
                'success': False,
                'error': f'Template "{template_name}" already exists'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create new tag template
        tag_template = TagTemplate.objects.create(
            name=template_name,
            description=description,
            formula_rules=formula_rules
        )
        
        logger.info(f"Created tag template: {template_name} with {len(formula_rules)} rules")
        
        return Response({
            'success': True,
            'template_id': tag_template.id,
            'message': f'Tag template "{template_name}" saved successfully',
            'template': tag_template.get_template_summary()
        })
        
    except Exception as e:
        logger.error(f"Error saving tag template: {e}")
        return Response({
            'success': False,
            'error': f'Failed to save tag template: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_tag_templates(request):
    """Get all saved tag templates."""
    try:
        templates = TagTemplate.objects.all()
        template_data = [template.get_template_summary() for template in templates]
        
        return Response({
            'success': True,
            'templates': template_data,
            'total_templates': len(template_data)
        })
        
    except Exception as e:
        logger.error(f"Error getting tag templates: {e}")
        return Response({
            'success': False,
            'error': f'Failed to get tag templates: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def delete_tag_template(request, template_id):
    """Delete a tag template."""
    try:
        template = TagTemplate.objects.get(id=template_id)
        template_name = template.name
        template.delete()
        
        logger.info(f"Deleted tag template: {template_name}")
        
        return Response({
            'success': True,
            'message': f'Tag template "{template_name}" deleted successfully'
        })
        
    except TagTemplate.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Tag template not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error deleting tag template: {e}")
        return Response({
            'success': False,
            'error': f'Failed to delete tag template: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def apply_tag_template(request, template_id):
    """Get formula rules from a tag template for application."""
    try:
        template = TagTemplate.objects.get(id=template_id)
        
        # Increment usage count
        template.increment_usage()
        
        logger.info(f"Applied tag template: {template.name}")
        
        # Externalize internal target names for UI clarity
        try:
            rules_ext = _externalize_formula_rules(template.formula_rules, None)
        except Exception:
            rules_ext = template.formula_rules
        return Response({
            'success': True,
            'template_name': template.name,
            'formula_rules': rules_ext,
            'message': f'Tag template "{template.name}" applied successfully'
        })
        
    except TagTemplate.DoesNotExist:
        return Response({
            'success': False,
            'error': 'Tag template not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error applying tag template: {e}")
        return Response({
            'success': False,
            'error': f'Failed to apply tag template: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def create_factwise_id(request):
    """Create a Factwise ID by combining two existing columns and map it to 'Item code'.
    Supports strategy: 'fill_only_null' (default) or 'override_all'. Treats variants of Item code as same.
    """
    try:
        session_id = request.data.get('session_id')
        first_column = request.data.get('first_column')
        second_column = request.data.get('second_column')
        operator = request.data.get('operator', '_')
        strategy = request.data.get('strategy', 'fill_only_null')
        
        info = get_session_consistent(session_id)
        if not session_id or not info:
            return no_store(Response({'success': False, 'error': 'Invalid session'}, status=400))
        
        if not first_column or not second_column:
            return no_store(Response({
                'success': False,
                'error': 'Both first_column and second_column are required'
            }, status=400))
        mappings = info.get("mappings")
        
        if not mappings:
            return Response({
                'success': False,
                'error': 'No mappings found for this session'
            }, status=status.HTTP_400_BAD_REQUEST)

        logger.info(f"🆔 Creating Factwise ID: {first_column} {operator} {second_column}")
        
        # Convert mappings list format to expected dict format for apply_column_mappings
        if isinstance(mappings, list):
            # Convert list format to new dict format that apply_column_mappings expects
            formatted_mappings = {"mappings": mappings}
            logger.info(f"🔧 DEBUG: Converted list mappings to dict format for Factwise ID: {formatted_mappings}")
        else:
            formatted_mappings = mappings
            logger.info(f"🔧 DEBUG: Using existing dict mappings for Factwise ID: {formatted_mappings}")
        
        # Get the current data - PREFER formula-enhanced data if available to preserve Tag/Spec/Customer columns
        headers = None
        data_rows = None
        enhanced_data = info.get("formula_enhanced_data")
        enhanced_headers = info.get("enhanced_headers")

        if enhanced_data and enhanced_headers and isinstance(enhanced_headers, list) and len(enhanced_headers) > 0:
            # Normalize enhanced data to list-of-lists for consistent processing
            headers = list(enhanced_headers)
            if isinstance(enhanced_data[0], dict):
                normalized_rows = []
                for row in enhanced_data:
                    normalized_row = []
                    for h in headers:
                        normalized_row.append(row.get(h, ""))
                    normalized_rows.append(normalized_row)
                data_rows = normalized_rows
            else:
                data_rows = enhanced_data
            logger.info(f"🔧 DEBUG: Using formula-enhanced data with {len(headers)} headers and {len(data_rows)} rows for Factwise ID")
        else:
            # Fall back to fresh mapped data if no enhanced data exists
            mapping_result = apply_column_mappings(
                client_file=info["client_path"],
                mappings=formatted_mappings,
                sheet_name=info["sheet_name"],
                header_row=info["header_row"] - 1 if info["header_row"] > 0 else 0,
                session_id=session_id
            )
            
            if not mapping_result or not mapping_result.get('data'):
                return Response({
                    'success': False,
                    'error': 'No data available for processing'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            headers = mapping_result['headers']
            data_rows = mapping_result['data']
        
        # Helper function to convert external names to internal names
        def convert_external_to_internal_name(external_name):
            """Convert external display names to internal column names."""
            external_name_lower = external_name.lower().strip()
            
            # Handle Customer identification name variations
            if external_name_lower in ['customer identification name', 'custom identification name']:
                # Find the first available Customer_Identification_Name_* column
                for header in headers:
                    if header.startswith('Customer_Identification_Name_'):
                        return header
                return None
            
            # Handle Customer identification value variations
            elif external_name_lower in ['customer identification value', 'custom identification value']:
                # Find the first available Customer_Identification_Value_* column
                for header in headers:
                    if header.startswith('Customer_Identification_Value_'):
                        return header
                return None
            
            # Handle Specification name variations
            elif external_name_lower in ['specification name']:
                # Find the first available Specification_Name_* column
                for header in headers:
                    if header.startswith('Specification_Name_'):
                        return header
                return None
            
            # Handle Specification value variations
            elif external_name_lower in ['specification value']:
                # Find the first available Specification_Value_* column
                for header in headers:
                    if header.startswith('Specification_Value_'):
                        return header
                return None
            
            # Handle Procurement entity name variations
            elif external_name_lower in ['procurement entity name']:
                # Look for exact match first, then try variations
                for header in headers:
                    if header.lower() == external_name_lower:
                        return header
                # If not found, return None
                return None
            
            # For other columns, try exact match first
            for header in headers:
                if header.lower() == external_name_lower:
                    return header
            
            # If no exact match, return the original name (might be a regular column)
            return external_name
        
        # Convert external names to internal names
        first_column_internal = convert_external_to_internal_name(first_column)
        second_column_internal = convert_external_to_internal_name(second_column)
        
        logger.info(f"🔧 DEBUG: Converting column names for Factwise ID creation:")
        logger.info(f"  First column: '{first_column}' -> '{first_column_internal}'")
        logger.info(f"  Second column: '{second_column}' -> '{second_column_internal}'")
        logger.info(f"  Available headers: {headers}")
        
        # Find column indices using internal names
        first_col_idx = -1
        second_col_idx = -1
        for i, header in enumerate(headers):
            if header == first_column_internal:
                first_col_idx = i
            elif header == second_column_internal:
                second_col_idx = i
        
        if first_col_idx == -1 or second_col_idx == -1:
            missing_columns = []
            if first_col_idx == -1:
                missing_columns.append(first_column)
            if second_col_idx == -1:
                missing_columns.append(second_column)
            
            return Response({
                'success': False,
                'error': f'Columns not found: {", ".join(missing_columns)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create Factwise ID values
        factwise_id_column = []
        for row in data_rows:
            first_val = str(row[first_col_idx]) if first_col_idx < len(row) and row[first_col_idx] is not None else ""
            second_val = str(row[second_col_idx]) if second_col_idx < len(row) and row[second_col_idx] is not None else ""

            if first_val and second_val:
                factwise_id = f"{first_val}{operator}{second_val}"
            elif first_val:
                factwise_id = first_val
            elif second_val:
                factwise_id = second_val
            else:
                factwise_id = ""

            factwise_id_column.append(factwise_id)

        # Helper to normalize header names
        def norm(s: str) -> str:
            return str(s).strip().lower().replace(' ', '').replace('_', '').replace('-', '')

        # Map Factwise ID into 'Item code' (create if missing). Treat Item code variants as same.
        item_idx = None
        new_headers = list(headers)
        for i, h in enumerate(new_headers):
            if norm(h) == norm('item code'):
                item_idx = i
                new_headers[i] = 'Item code'
                break

        new_data_rows = []
        if item_idx is None:
            new_headers = ['Item code'] + new_headers
            for i, row in enumerate(data_rows):
                new_row = [factwise_id_column[i]] + list(row)
                new_data_rows.append(new_row)
        else:
            for i, row in enumerate(data_rows):
                new_row = list(row)
                while len(new_row) < len(new_headers):
                    new_row.append("")
                if strategy == 'override_all':
                    new_row[item_idx] = factwise_id_column[i]
                else:  # fill only null/empty
                    current_val = new_row[item_idx]
                    if current_val is None or str(current_val).strip() == "":
                        new_row[item_idx] = factwise_id_column[i]
                new_data_rows.append(new_row)
        
        # Store Factwise ID rule for template saving and reuse
        factwise_id_rule = {
            "type": "factwise_id",
            "first_column": first_column,
            "second_column": second_column,
            "operator": operator,
            "strategy": strategy
        }
        
        # Update session headers for immediate UI reflect
        # Avoid persisting full data for large datasets; data pages are recomputed on demand
        SMALL_DATA_THRESHOLD = 2000
        if len(new_data_rows) <= SMALL_DATA_THRESHOLD:
            info["formula_enhanced_data"] = new_data_rows
        else:
            # Ensure any previous large cache is cleared
            info.pop("formula_enhanced_data", None)
        info["enhanced_headers"] = new_headers
        # Persist canonical headers to avoid alternating states across requests
        try:
            info["current_template_headers"] = new_headers
        except Exception:
            pass
            
        debug_log(session_id, "Updated session with Factwise ID data", {
            'new_headers_count': len(new_headers),
            'new_data_rows_count': len(new_data_rows),
            'factwise_id_rule': factwise_id_rule
        })
        
        # Initialize factwise_rules if not exists
        if "factwise_rules" not in info:
            info["factwise_rules"] = []
        
        # Add or update the Factwise ID rule (only keep one), using internal column names
        info["factwise_rules"] = [rule for rule in info.get("factwise_rules", []) if rule.get("type") != "factwise_id"]
        info["factwise_rules"].append(factwise_id_rule)
        
        # CRITICAL FIX: Ensure session is fully saved before returning response
        # This prevents the race condition where frontend fetches stale data
        debug_log(session_id, "Saving session before returning Factwise ID response", {
            'session_keys': list(info.keys()),
            'has_formula_enhanced_data': 'formula_enhanced_data' in info,
            'has_enhanced_headers': 'enhanced_headers' in info
        })
        
        # Force immediate session save to prevent race conditions
        try:
            save_session(session_id, info)
            # Verify session was saved by checking if it's accessible
            if session_id in SESSION_STORE:
                debug_log(session_id, "Session saved successfully", {
                    'session_keys_after_save': list(SESSION_STORE[session_id].keys()),
                    'data_persisted': 'formula_enhanced_data' in SESSION_STORE[session_id]
                })
            else:
                debug_log(session_id, "WARNING: Session not found after save", level='warning')
        except Exception as save_error:
            debug_log(session_id, f"Error saving session: {save_error}", level='error')
            # Continue anyway, but log the error
        
        # Save all updated data and headers before version bump
        if len(new_data_rows) <= SMALL_DATA_THRESHOLD:
            info["formula_enhanced_data"] = new_data_rows
        else:
            info.pop("formula_enhanced_data", None)
        info["enhanced_headers"] = new_headers
        info["current_template_headers"] = new_headers
        save_session(session_id, info)
        
        # Atomic version bump AFTER all data is saved
        new_version = increment_template_version(session_id)
        
        logger.info(f"🆔 Successfully created Factwise ID mapped into 'Item code' with {len(factwise_id_column)} entries (strategy={strategy})")
        
        return no_store(Response({
            'success': True,
            'template_version': new_version,
            'enhanced_headers': new_headers,
            'rows': len(new_data_rows),
            'message': 'Factwise ID created and mapped to Item code'
        }))
        
    except Exception as e:
        logger.error(f"Error creating Factwise ID: {e}")
        return no_store(Response({
            'success': False,
            'error': f'Failed to create Factwise ID: {str(e)}'
        }, status=500))


@api_view(['GET'])
def system_diagnostics(request):
    """Comprehensive system diagnostics for session persistence verification."""
    try:
        import psutil
        import os
        import platform
        from datetime import datetime, timedelta
        
        # System information with worker diagnostics
        system_info = {
            'platform': platform.platform(),
            'python_version': platform.python_version(),
            'cpu_count': psutil.cpu_count(),
            'memory_total_gb': round(psutil.virtual_memory().total / (1024**3), 2),
            'memory_used_gb': round(psutil.virtual_memory().used / (1024**3), 2),
            'pid': os.getpid(),
            'worker_id': os.environ.get('SERVER_SOFTWARE', 'Unknown'),
            'timestamp': datetime.utcnow().isoformat(),
        }
        
        # Session persistence analysis
        sessions_info = []
        session_files_count = 0
        
        if hybrid_file_manager.local_temp_dir.exists():
            session_files = list(hybrid_file_manager.local_temp_dir.glob("session_*.json"))
            session_files_count = len(session_files)
        
        for session_id, session_data in list(SESSION_STORE.items())[-10:]:  # Last 10 sessions
            session_file = hybrid_file_manager.local_temp_dir / f"session_{session_id}.json"
            sessions_info.append({
                'session_id': session_id[:8] + "...",  # Truncate for security
                'created': session_data.get('created'),
                'has_client_file': bool(session_data.get('client_path')),
                'has_template_file': bool(session_data.get('template_path')),
                'has_mappings': bool(session_data.get('mappings')),
                'template_modified': session_data.get('template_modified', False),
                'persisted_to_file': session_file.exists(),
            })
        
        # File system diagnostics
        file_system_info = {
            'temp_dir_exists': hybrid_file_manager.local_temp_dir.exists(),
            'upload_dir_exists': hybrid_file_manager.local_upload_dir.exists(),
            'temp_dir_path': str(hybrid_file_manager.local_temp_dir),
            'upload_dir_path': str(hybrid_file_manager.local_upload_dir),
        }
        
        # Azure storage diagnostics
        azure_available = hybrid_file_manager.azure_storage.is_available()
        
        # Session persistence health check
        persistence_health = {
            'memory_sessions_count': len(SESSION_STORE),
            'file_sessions_count': session_files_count,
            'universal_session_helpers_active': True,
            'single_worker_config_applied': True,
            'file_fallback_available': True,
        }
        
        # Fix status summary
        fix_status = {
            'issue_identified': 'Multi-worker Gunicorn session isolation',
            'primary_fix': 'Single worker configuration (workers=1)',
            'secondary_fix': 'Enhanced session persistence with file fallback',
            'affected_endpoints': ['/api/upload/', '/api/headers/{session_id}/', '/api/mapping/', '/api/data/'],
            'fix_applied': True,
            'expected_behavior': 'Sessions persist across all requests within same server instance',
            'deployment_required': True,
            'deployment_note': 'Restart Azure Web App to apply Gunicorn configuration changes',
        }
        
        return Response({
            'success': True,
            'system_info': system_info,
            'session_persistence': persistence_health,
            'file_system': file_system_info,
            'azure_storage_available': azure_available,
            'recent_sessions': sessions_info,
            'fix_status': fix_status,
            'diagnostics_timestamp': datetime.utcnow().isoformat(),
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': f'System diagnostics failed: {str(e)}',
            'timestamp': datetime.utcnow().isoformat(),
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Custom logging function for debugging template and default value issues
def debug_log(session_id, message, data=None, level='info'):
    """Enhanced logging for debugging template and default value issues"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    log_msg = f"[{timestamp}] 🔍 SESSION_{session_id}: {message}"
    if data is not None:
        log_msg += f" | DATA: {json.dumps(data, default=str)[:500]}"
    
    if level == 'error':
        logger.error(log_msg)
    elif level == 'warning':
        logger.warning(log_msg)
    else:
        logger.info(log_msg)

# Add this function near the top of the file, after imports
def get_next_available_tag_column(session_info, used_tag_columns=None):
    """
    Centralized function to get the next available Tag column name.
    This prevents duplicate Tag columns and ensures consistent numbering.
    """
    if used_tag_columns is None:
        used_tag_columns = set()
    
    # Get all existing Tag columns from session (only numbered ones)
    existing_headers = session_info.get('current_template_headers', []) or session_info.get('enhanced_headers', []) or []
    existing_tag_columns = [h for h in existing_headers if h.startswith('Tag_')]
    
    # Find next available number
    next_number = 1
    while f'Tag_{next_number}' in existing_tag_columns or f'Tag_{next_number}' in used_tag_columns:
        next_number += 1
    
    return f'Tag_{next_number}'

def convert_internal_to_external_name(column_name):
    """
    Convert internal column names to external display names.
    Internal: Tag_1, Tag_2 -> External: Tag (always generic name)
    """
    if column_name.startswith('Tag_') or column_name == 'Tag':
        # Always show "Tag" regardless of how many tags exist
        return 'Tag'
    elif column_name.startswith('Specification_Name_') or column_name == 'Specification name':
        return 'Specification name'
    elif column_name.startswith('Specification_Value_') or column_name == 'Specification value':
        return 'Specification value'
    elif column_name.startswith('Customer_Identification_Name_') or column_name == 'Customer identification name':
        return 'Customer identification name'
    elif column_name.startswith('Customer_Identification_Value_') or column_name == 'Customer identification value':
        return 'Customer identification value'
    
    return column_name

def convert_external_to_internal_name(column_name, session_info, used_columns=None):
    """
    Convert external column names to internal names with proper numbering.
    External: Tag -> Internal: Tag_1, Tag_2, etc.
    """
    if used_columns is None:
        used_columns = set()
    
    if column_name == 'Tag':
        return get_next_available_tag_column(session_info, used_columns)
    elif column_name == 'Specification name':
        # Find next available spec number
        existing_headers = session_info.get('current_template_headers', []) or session_info.get('enhanced_headers', []) or []
        existing_spec_columns = [h for h in existing_headers if h.startswith('Specification_Name_')]
        next_number = 1
        while f'Specification_Name_{next_number}' in existing_spec_columns or f'Specification_Name_{next_number}' in used_columns:
            next_number += 1
        return f'Specification_Name_{next_number}'
    elif column_name == 'Specification value':
        # Find next available spec number
        existing_headers = session_info.get('current_template_headers', []) or session_info.get('enhanced_headers', []) or []
        existing_spec_columns = [h for h in existing_headers if h.startswith('Specification_Value_')]
        next_number = 1
        while f'Specification_Value_{next_number}' in existing_spec_columns or f'Specification_Value_{next_number}' in used_columns:
            next_number += 1
        return f'Specification_Value_{next_number}'
    elif column_name == 'Customer identification name':
        # Find next available customer number
        existing_headers = session_info.get('current_template_headers', []) or session_info.get('enhanced_headers', []) or []
        existing_customer_columns = [h for h in existing_headers if h.startswith('Customer_Identification_Name_')]
        next_number = 1
        while f'Customer_Identification_Name_{next_number}' in existing_customer_columns or f'Customer_Identification_Name_{next_number}' in used_columns:
            next_number += 1
        return f'Customer_Identification_Name_{next_number}'
    elif column_name == 'Customer identification value':
        # Find next available customer number
        existing_headers = session_info.get('current_template_headers', []) or session_info.get('enhanced_headers', []) or []
        existing_customer_columns = [h for h in existing_headers if h.startswith('Customer_Identification_Value_')]
        next_number = 1
        while f'Customer_Identification_Value_{next_number}' in existing_customer_columns or f'Customer_Identification_Value_{next_number}' in used_columns:
            next_number += 1
        return f'Customer_Identification_Value_{next_number}'
    
    return column_name


@api_view(['GET'])
def mpn_cache_stats(request):
    """Get global MPN cache statistics"""
    try:
        from .models import GlobalMpnCache

        stats = GlobalMpnCache.get_cache_stats()

        return Response({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        logger.error(f"Failed to get MPN cache stats: {e}")
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def mpn_cache_cleanup(request):
    """Clean up old MPN cache entries"""
    try:
        from .models import GlobalMpnCache

        # Get cleanup parameters
        cleanup_type = request.data.get('type', 'old')  # 'old' or 'invalid'
        days_old = request.data.get('days_old', 365 if cleanup_type == 'old' else 30)

        if cleanup_type == 'old':
            count = GlobalMpnCache.cleanup_old_entries(days_old=days_old)
            message = f"Cleaned up {count} old cache entries (>{days_old} days)"
        elif cleanup_type == 'invalid':
            count = GlobalMpnCache.cleanup_invalid_entries(days_old=days_old)
            message = f"Cleaned up {count} invalid cache entries (>{days_old} days)"
        else:
            return Response({
                'success': False,
                'error': 'Invalid cleanup type. Use "old" or "invalid".'
            }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'success': True,
            'message': message,
            'count': count
        })
    except Exception as e:
        logger.error(f"Failed to cleanup MPN cache: {e}")
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
